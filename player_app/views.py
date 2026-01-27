import csv
from socket import TCP_NODELAY

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse, HttpResponseForbidden
from openpyxl import Workbook
from openpyxl.styles import Font
from .models import Player, Player_Group, CampTournament, CampActivity, Program, AssignedProgram, WorkoutData, Injury, \
    MedicalDocument
from player_app.forms import *
from django.http import JsonResponse
import json
import pandas as pd
from django.shortcuts import render, redirect
from django.contrib.auth import login, authenticate, get_user_model
from django.contrib.auth import logout
from django.contrib import messages
from openpyxl import load_workbook
from django.core.exceptions import ValidationError
from accounts.models import Organization, Staff
from form.forms import AssignForm
from form.models import FormAssignment
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist


# from ..app.models import DailyWellnessForm
# from player_app.ams.app.models import DailyWellnessForm


# View to list all players
def player_list(request):
    # Get all organizations for the superadmin
    organizations = Organization.objects.all()

    # Get the logged-in user's organization (if not a superuser)
    selected_organization = None
    players = []

    if request.user.is_superuser:
        # If superuser, allow organization selection
        selected_organization_id = request.GET.get('organization')
        if selected_organization_id:
            selected_organization = get_object_or_404(Organization, id=selected_organization_id)
            players = Player.objects.filter(organization=selected_organization)
    else:
        # Handle case where the staff user has no assigned organization
        try:
            if request.user.organization:  # Ensure organization exists
                selected_organization = request.user.organization
                players = Player.objects.filter(organization=selected_organization)
            else:
                players = []  # No players to show if no organization
        except ObjectDoesNotExist:
            players = []  # Handle missing organization safely

    # Pass organizations, selected organization, and players to the template
    return render(request, 'player_app/player_list.html', {
        'organizations': organizations,
        'selected_organization': selected_organization,
        'players': players,
    })


# View to display a single player's details
def player_detail(request, pk):
    player = get_object_or_404(Player, pk=pk)
    return render(request, 'player_app/player_detail.html', {'player': player})


from django.contrib.auth import get_user_model


def player_create(request):
    if request.method == 'POST':
        form = PlayerForm(request.POST, request.FILES)
        if form.is_valid():
            ins = form.save(commit=False)

            # Set the organization automatically if the user is an OrganizationAdmin
            if request.user.role == "OrganizationAdmin":
                # Automatically assign the logged-in organization's instance to this player
                organization = get_object_or_404(Organization, user=request.user)
                ins.organization = organization
            elif request.user.role == "Staff":
                # Fetch the organization of the logged-in Staff user
                staff = get_object_or_404(Staff, user=request.user)
                organization = staff.organization  # Assign the organization tied to the Staff user
                ins.organization = organization

            email = request.POST.get('email')
            firstname = request.POST.get('name').split(" ")[0]
            password = form.cleaned_data['password']

            User = get_user_model()
            user = User.objects.create(username=email, email=email, first_name=firstname)
            user.set_password(password)  # Set the password securely
            user.is_super_admin = False
            user.role = "Player"
            user.save()

            ins.user = user
            ins.save()
            messages.success(request, 'Player created successfully!')
            return redirect('player_list')
        else:
            messages.error(request, 'Please correct the errors below.')
            print("Form errors:", form.errors)
    else:
        form = PlayerForm()

    return render(request, 'player_app/player_form.html', {'form': form, 'title': 'Create Player'})


def player_update(request, pk):
    player = get_object_or_404(Player, pk=pk)
    original_organization = player.organization  # ✅ Store the original organization before update

    if request.method == 'POST':
        form = PlayerForm(request.POST, request.FILES, instance=player)
        files = request.FILES.getlist('documents')  # Get multiple uploaded documents

        if form.is_valid():
            password = form.cleaned_data.get('password')
            if password:
                user = player.user
                user.set_password(password)
                user.save()

            player = form.save(commit=False)  # ⛔️ Don't save yet

            # ✅ Preserve Organization: If missing, restore original organization
            if not player.organization:
                player.organization = original_organization

            player.save()  # ✅ Now save the updated player with the correct organization

            # Save uploaded medical documents
            for file in files:
                MedicalDocument.objects.create(player=player, document=file)

            messages.success(request, 'Player updated successfully!')
            return redirect('player_list')
        else:
            messages.error(request, 'Please correct the errors below.')
            print("Form errors:", form.errors)

    else:
        form = PlayerForm(instance=player)

    medical_documents = player.medical_documents.all()

    return render(request, 'player_app/player_form.html', {
        'form': form,
        'title': 'Update Player',
        'player': player,
        'medical_documents': medical_documents
    })


# View to delete an existing player
def player_delete(request, pk):
    player = get_object_or_404(Player, pk=pk)
    player.delete()
    messages.success(request, 'Player deleted successfully!')
    return redirect('player_list')




# View to list all players


def organization_player_list(request):
    if request.user.role == "Staff":
        org = request.user.staff.organization

    if request.user.role == "OrganizationAdmin":
        org = get_object_or_404(Organization, user=request.user)

    players = Player.objects.filter(organization=org)

    # Collect filter params
    age_categories = request.GET.getlist('age_category')
    handednesses = request.GET.getlist('handedness')
    roles = request.GET.getlist('role')
    player_statuses = request.GET.getlist('player_status')
    sort_gender = request.GET.get('sort') == 'gender'

    # Filtering logic
    if age_categories:
        players = players.filter(age_category__in=age_categories)
    if handednesses:
        players = players.filter(handedness__in=handednesses)
    if roles:
        players = players.filter(role__in=roles)
    if player_statuses:
        players = players.filter(player_status__in=player_statuses)

    # Calculate filters count including player_statuses
    filters_count = len(age_categories) + len(handednesses) + len(roles) + len(player_statuses)

    # Use your actual lowercase class attribute names for choices
    AGE_CHOICES = getattr(Player, "Age_category_choices", [])
    HAND_CHOICES = getattr(Player, "handedness_choices", [])
    ROLE_CHOICES = (
        Player.objects.filter(organization=org)
        .values_list('role', flat=True)
        .distinct()
        .order_by('role')
    )

    # Sorting by gender
    if sort_gender:
        gender_order = {'F': 0, 'M': 1, 'O': 2}
        players = sorted(
            players,
            key=lambda p: (
                gender_order.get(getattr(p, "gender", ""), 9),
                (getattr(p, "name", "") or '').lower()
            )
        )

    request_getlist = {k: request.GET.getlist(k) for k in request.GET}

    context = {
        'players': players,
        'AGE_CHOICES': AGE_CHOICES,
        'HAND_CHOICES': HAND_CHOICES,
        'ROLE_CHOICES': ROLE_CHOICES,
        'active_filters': {
            'age_category': age_categories,
            'handedness': handednesses,
            'role': roles,
            'player_status': player_statuses,
        },
        'filters_count': filters_count,
        'sort_gender': sort_gender,
        'request_getlist': request_getlist,
        'request_obj': request,
    }
    return render(request, "player_app/organization/organization_player_list.html", context)

@login_required
def organization_player_add(request):
    age_category_choices = getattr(Player, "Age_category_choices", [])
    return render(request, 'player_app/organization/organization_player_form.html', {'age_category_choices':age_category_choices})

@login_required
def player_create_view(request):
    if request.method == 'POST':
        # Extract fields manually from request.POST and request.FILES
        name = request.POST.get('name')
        image = request.FILES.get('image')  # File input
        email = request.POST.get('email')
        date_of_birth = request.POST.get('date_of_birth')
        primary_contact_number = request.POST.get('primary_contact_number')
        secondary_contact_number = request.POST.get('secondary_contact_number')
        gender = request.POST.get('gender')
        state = request.POST.get('state')
        district = request.POST.get('district')
        role = request.POST.get('role')
        batting_style = request.POST.get('batting_style')
        bowling_style = request.POST.get('bowling_style')
        handedness = request.POST.get('handedness')
        age_category = request.POST.get('age_category')
        guardian_name = request.POST.get('guardian_name')
        relation = request.POST.get('relation')
        guardian_mobile_number = request.POST.get('guardian_mobile_number')
        password = "admin"  # Or generate/set your own logic

        # Create Player/user instances and link
        User = get_user_model()

        # Basic Email uniqueness check to avoid duplicate users
        if User.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return render(request, 'player_app/organization/organization_player_form.html', {'title': 'Create Player'})

        user = User.objects.create(username=email, email=email, first_name=name.split(' ')[0])
        user.set_password(password)
        user.is_super_admin = False
        user.role = "Player"
        user.save()

        player = Player(
            name=name,
            email=email,
            date_of_birth=date_of_birth or None,
            primary_contact_number=primary_contact_number,
            secondary_contact_number=secondary_contact_number,
            gender=gender,
            state=state,
            district=district,
            role=role,
            batting_style=batting_style,
            bowling_style=bowling_style,
            handedness=handedness,
            age_category=age_category,
            guardian_name=guardian_name,
            relation=relation,
            guardian_mobile_number=guardian_mobile_number,
            user=user,
        )

        if image:
            player.image = image

        # Assign organization from user role as in your original logic
        if hasattr(request.user, 'role') and request.user.role == "OrganizationAdmin":
            organization = get_object_or_404(Organization, user=request.user)
            player.organization = organization
        elif hasattr(request.user, 'role') and request.user.role == "Staff":
            staff = get_object_or_404(Staff, user=request.user)
            player.organization = staff.organization

        player.save()

        # If you have M2M fields, handle here like player.m2mfield.set([...])

        PlayerActivityLog.objects.create(
            player=player,
            actor=request.user,
            action='created',
            details=f"Player '{player.name}' was created."
        )

        messages.success(request, 'Player created successfully!')
        return redirect('organization_player_list')

    
    return render(request, 'player_app/organization/organization_player_form.html', {'title': 'Create Player'})


from django.utils.timezone import now

@login_required
def organization_player_edit(request, pk):
    if request.user.role == "Staff":
        player = get_object_or_404(Player, pk=pk, organization=request.user.staff.organization)
        
        return redirect('organization_player_detail', pk=pk)
        
    org = get_object_or_404(Organization, user=request.user)
    player = get_object_or_404(Player, pk=pk, organization=org)
    user = player.user

    if request.method == 'POST':
        # Store old values for comparison
        old_data = {
            'name': player.name,
            'email': player.email,
            'primary_contact_number': player.primary_contact_number,
            'image':player.image.url if player.image else None,
            'date_of_birth': player.date_of_birth,
            'primary_contact_number': player.primary_contact_number,
            'secondary_contact_number': player.secondary_contact_number,
            'gender': player.gender,
            'state': player.state,
            'role': player.role,
            'batting_style': player.batting_style,
            'bowling_style': player.bowling_style,
            'handedness': player.handedness,
            'age_category': player.age_category,
            'guardian_name': player.guardian_name,
            'relation': player.relation,
            'guardian_mobile_number': player.guardian_mobile_number,
            
        }

        form = OrganizationPlayerFormUpdate(request.POST, request.FILES, instance=player)
        if form.is_valid():
            player = form.save()

            # Compare and prepare change log
            changes = []
            for field, old_value in old_data.items():
                new_value = getattr(player, field)
                if new_value != old_value:
                    changes.append(f"{field} changed from '{old_value}' to '{new_value}'")

            # Sync user first_name and email with updated player info
            user.first_name = player.name.split(' ')[0]
            user.email = player.email

            # Password update logic
            new_password = form.cleaned_data.get('new_password')
            if new_password:
                user.set_password(new_password)
                messages.info(request, "Password updated for this player.")
            user.save()

            # Log changes only if there are changes
            if changes:
                PlayerActivityLog.objects.create(
                    player=player,  # Use player, not injury
                    actor=request.user,
                    action='updated',
                    details="; ".join(changes)
                )

            messages.success(request, "Player updated successfully.")
            return redirect('organization_player_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = OrganizationPlayerFormUpdate(instance=player)

    return render(request, 'player_app/organization/organization_player_edit.html', {
        'form': form,
        'player': player,
        'title': 'Edit Player'
    })


@login_required
def organization_player_delete(request, pk):
    org = get_object_or_404(Organization, user=request.user)
    player = get_object_or_404(Player, pk=pk, organization=org)
    player.user.delete()  # OnDelete CASCADE deletes Player too
    messages.success(request, "Player deleted successfully!")
    return redirect('organization_player_list')

from django.db.models import Q
from django.db.models import Q
from collections import defaultdict

@login_required
def organization_player_detail(request, pk):
    player = get_object_or_404(Player, pk=pk)
    injuries = player.injuries.select_related('reported_by').all().order_by('-injury_date')
    documents = (
        player.medical_documents
        .filter(Q(view_option="profile") | Q(view_option="injury_profile"))
        .order_by('-date', '-uploaded_at')
    )
    
    # NEW: Group test results by test type
    test_results = defaultdict(list)
    test_choices_dict = dict(TestAndResult.TEST_CHOICES)
    
    player_tests = player.testandresult_set.all().order_by('-created_at')
    for test in player_tests:
        test_name = test_choices_dict.get(test.test, test.test)
        test_results[test_name].append(test)
    
    # Dynamic status
    injury_status = "Injured" if player.injuries.filter(status='open').exists() else "Fit"
    participation_status = "Benched" if injury_status == "Injured" else "Available"

    # Doc upload logic
    if request.method == "POST":
        doc_form = MedicalDocumentForm(request.POST, request.FILES, player=player)
        if doc_form.is_valid():
            doc = doc_form.save(commit=False)
            doc.player = player
            doc.user = request.user                
            doc.save()

            MedicalActivityLog.objects.create(
                player=player,
                document=doc,
                user=request.user,
                activity_type='UPLOAD',
                description=f"{request.user.get_username()} uploaded medical document '{doc.title}'"
            )
            messages.success(request, "Medical document uploaded.")
            return redirect(request.path)
    else:
        doc_form = MedicalDocumentForm(player=player)

    context = {
        "player": player,
        "injuries": injuries,
        "documents": documents,
        "test_results": dict(test_results),  # Convert defaultdict to dict
        "injury_status": injury_status,
        "participation_status": participation_status,
        "doc_form": doc_form,
    }
    return render(request, "player_app/organization/organization_player_detail.html", context)

import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required

@login_required
def organization_player_export(request):
    org = get_object_or_404(Organization, user=request.user)
    players = Player.objects.filter(organization=org)

    # Apply same filters as your list view!
    age_categories = request.GET.getlist('age_category')
    handednesses = request.GET.getlist('handedness')
    roles = request.GET.getlist('role')
    if age_categories:
        players = players.filter(age_category__in=age_categories)
    if handednesses:
        players = players.filter(handedness__in=handednesses)
    if roles:
        players = players.filter(role__in=roles)

    # If you want to keep column order/export identical, use the ALL_SWAP_COLS and FIXED_COLS in your JS.
    columns = [
        ("S.No", None),
        ("Player ID", "id"),
        ("Name", "name"),
        ("Email", "email"),
        ("D.O.B", "date_of_birth"),
        ("Primary Contact", "primary_contact_number"),
        ("Secondary Contact", "secondary_contact_number"),
        ("Gender", "get_gender_display"),
        ("State", "state"),
        ("Role", "role"),
        ("Batting Style", "batting_style"),
        ("Bowling Style", "bowling_style"),
        ("Handedness", "get_handedness_display"),
        ("Age Category", "get_age_category_display"),
        ("Guardian Name", "guardian_name"),
        ("Relation", "relation"),
        ("Guardian Mobile", "guardian_mobile_number")
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Players"

    # Header row
    ws.append([col[0] for col in columns])

    # Data rows
    for idx, player in enumerate(players, start=1):
        row = []
        for col_name, field in columns:
            if field is None:
                row.append(idx)  # S.No
            else:
                attr = getattr(player, field, "")
                if callable(attr):
                    val = attr()
                else:
                    val = attr
                row.append(val if val is not None else "")
        ws.append(row)

    # Fit width
    for i, col in enumerate(ws.columns, 1):
        length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = min(length + 3, 35)

    # Return as download
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=Players_Export.xlsx'
    wb.save(response)
    return response



#  Injury Management Views
from django.shortcuts import render, get_object_or_404, redirect
from .models import Organization, Player, Staff
from .forms import InjuryForm
from django.views.decorators.http import require_GET


from datetime import date, timedelta
from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.db.models import Count
from player_app.models import Injury, Organization, Player

from datetime import date, timedelta
from django.shortcuts import render, get_object_or_404
from django.db.models import Count
from .models import Injury, Organization, Player

def organization_injury_list(request):
    # Identify organization
    if request.user.role == "Staff":
        org = request.user.staff.organization
    elif request.user.role == "OrganizationAdmin":
        org = get_object_or_404(Organization, user=request.user)
    else:
        org = None

    injuries = Injury.objects.filter(player__organization=org).select_related("player", "reported_by")
    today = date.today()

    # Filters from GET
    severity_vals = request.GET.getlist('severity')
    status_vals = request.GET.getlist('status')
    body_parts = request.GET.getlist('body_region')
    player_ids = request.GET.getlist('player_id')
    search_name = request.GET.get('name', '').strip()
    sort = request.GET.get('sort', '')

    month = request.GET.get('month')
    year = request.GET.get('year')
    range_filter = request.GET.get('range', '')
    season = request.GET.get('season')
    categories = request.GET.getlist('category')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    MONTH_CHOICES = [
        (1, "January"), (2, "February"), (3, "March"), (4, "April"),
        (5, "May"), (6, "June"), (7, "July"), (8, "August"),
        (9, "September"), (10, "October"), (11, "November"), (12, "December")
    ]

    # Generate seasons last 10 years (e.g. 2025-2026)
    current_season_start = today.year if today.month >= 4 else today.year - 1
    season_choices = [f"{y}-{y+1}" for y in range(current_season_start, current_season_start - 10, -1)]

    # Filtering logic
    if search_name:
        injuries = injuries.filter(player__name__icontains=search_name)
    else:
        if start_date and end_date:
            injuries = injuries.filter(injury_date__range=[start_date, end_date])
        elif range_filter == "last3":
            min_date = today - timedelta(days=90)
            injuries = injuries.filter(injury_date__gte=min_date)
        elif range_filter == "last6":
            min_date = today - timedelta(days=182)
            injuries = injuries.filter(injury_date__gte=min_date)
        elif month and year:
            injuries = injuries.filter(injury_date__month=int(month), injury_date__year=int(year))
        elif season and year:
            try:
                year_val = int(year)
                sy, ey = map(int, season.split('-'))
                start = date(sy, 4, 1)
                end = date(ey, 3, 31)
                injuries = injuries.filter(injury_date__gte=start, injury_date__lte=end, injury_date__year=year_val)
            except Exception:
                pass
        elif season:
            try:
                sy, ey = map(int, season.split('-'))
                start = date(sy, 4, 1)
                end = date(ey, 3, 31)
                injuries = injuries.filter(injury_date__gte=start, injury_date__lte=end)
            except Exception:
                pass

        if categories:
            injuries = injuries.filter(nature_of_injury__in=categories)
        if severity_vals:
            injuries = injuries.filter(severity__in=severity_vals)
        if status_vals:
            injuries = injuries.filter(status__in=status_vals)
        if body_parts:
            injuries = injuries.filter(affected_body_part__in=body_parts)
        if player_ids:
            injuries = injuries.filter(player__id__in=player_ids)

    # Sorting
    if sort == 'severity':
        injuries = injuries.order_by('severity')
    elif sort == 'date':
        injuries = injuries.order_by('injury_date')
    elif sort == 'status':
        injuries = injuries.order_by('status')

    # Stats counts
    injury_total = injuries.count()
    injury_open = injuries.filter(status='open').count()
    injury_closed = injuries.filter(status='closed').count()

    SEVERITY_CHOICES = Injury.SEVERITY_CHOICES
    STATUS_CHOICES = Injury.STATUS_CHOICES
    BODY_PART_CHOICES = list(
        Injury.objects.filter(player__organization=org)
        .values_list('affected_body_part', flat=True)
        .distinct()
        .exclude(affected_body_part__isnull=True)
        .exclude(affected_body_part__exact='')
    )
    BODY_PART_CHOICES = [(bp, bp.title()) for bp in BODY_PART_CHOICES if bp]

    PLAYER_CHOICES = Player.objects.filter(organization=org).values_list('id', 'name')

    CATEGORY_CHOICES = (
        Player.objects.filter(organization=org)
        .exclude(age_category__isnull=True)
        .exclude(age_category='')
        .values_list('age_category', flat=True)
        .distinct()
    )

    # Count active filters
    filters_count = (
        len(severity_vals) + len(status_vals) + len(body_parts) +
        len(player_ids) + len(categories)
    )
    if search_name:
        filters_count += 1
    if month or year:
        filters_count += 1
    if range_filter:
        filters_count += 1
    if season:
        filters_count += 1
    if start_date or end_date:
        filters_count += 1

    # Player participation stats
    # player_status_counts = injuries.values('player_status').annotate(count=Count('player_status'))
    # status_map = {"full participation": 0, "limited participation": 0, "no participation": 0}
    # for ps in player_status_counts:
    #     key = ps['player_status']
    #     if key in status_map:
    #         status_map[key] = ps['count']
    
    context = {
        'year_choices': list(range(2020, today.year + 1)),
        'injuries': injuries,
        'SEVERITY_CHOICES': SEVERITY_CHOICES,
        'STATUS_CHOICES': STATUS_CHOICES,
        'BODY_PART_CHOICES': BODY_PART_CHOICES,
        'PLAYER_CHOICES': PLAYER_CHOICES,
        'CATEGORY_CHOICES': CATEGORY_CHOICES,
        'season_choices': season_choices,
        'MONTH_CHOICES': MONTH_CHOICES,
        'injury_total': injury_total,
        'injury_open': injury_open,
        'injury_closed': injury_closed,
        'filters_count': filters_count,
        'today': today,
        'start_date': start_date,
        'end_date': end_date,
        'filter_month': month,
        'filter_year': year,
        'filter_range': range_filter,
        'filter_season': season,
        'filter_categories': categories,
       
        'active_filters': {
            'severity': severity_vals,
            'status': status_vals,
            'body_region': body_parts,
            'player_id': player_ids,
            'name': search_name,
            'category': categories,
            'month': month,
            'year': year,
            'range': range_filter,
            'season': season,
            'start_date': start_date,
            'end_date': end_date,
        },
        'request_getlist': {k: request.GET.getlist(k) for k in request.GET},
        'sort': sort,
    }

    return render(request, "player_app/organization/organization_injury_list.html", context)



@login_required
def activity_log_combined_view(request):
    # Fetch all medical logs
    medical_logs = MedicalActivityLog.objects.select_related('player', 'user', 'document').order_by('-timestamp')

    # Fetch all injury logs
    injury_logs = InjuryActivityLog.objects.select_related('injury', 'actor').order_by('-created_at')

    # Fetch all player logs
    player_logs = PlayerActivityLog.objects.select_related('player', 'actor').order_by('-created_at')

    logs = []

    for log in medical_logs:
        logs.append({
            'log_type': 'Medical',
            'time': log.timestamp,
            'user': log.user.username if log.user else '',
            'target': log.player.name,
            'action': log.activity_type,
            'desc': log.description,
        })

    for log in injury_logs:
        logs.append({
            'log_type': 'Injury',
            'time': log.created_at,
            'user': log.actor.username if log.actor else '',
            'target': str(log.injury),
            'action': log.action,
            'desc': log.details,
        })

    for log in player_logs:
        logs.append({
            'log_type': 'Player',
            'time': log.created_at,
            'user': log.actor.username if log.actor else '',
            'target': log.player.name,
            'action': log.action,
            'desc': log.details,
        })

    # Sort all logs by time descending
    logs.sort(key=lambda x: x['time'], reverse=True)

    return render(request, "player_app/organization/activity_log_combined.html", {
        'logs': logs,
    })

@login_required
def organization_create_injury(request):
    # Get the current user's organization
    if request.user.role == "Staff":
            organization = request.user.staff.organization

    if request.user.role == "OrganizationAdmin":
          organization = get_object_or_404(Organization, user=request.user)
   
    players_qs = Player.objects.filter(organization=organization)
    physios_qs = Staff.objects.filter(organization=organization, role__iexact='Physio')
    
    if request.method == 'POST':
        # Pass the filtered querysets to the form so the select fields are filtered
        form = InjuryForm(request.POST, players_qs=players_qs, physios_qs=physios_qs)
        if form.is_valid():
           
            injury = form.save()
            # Log injury creation
            InjuryActivityLog.objects.create(
                injury=injury,
                actor=request.user,
                action='created',
                details=f'Injury reported by {injury.reported_by} for player {injury.player}'
            )
            return redirect('organization_injury_list')  # Update to your desired redirect
        else:
            print(form.errors)# Print all form errors in console for debugging
    else:
        # GET: Create empty form with filtered choices
        
        form = InjuryForm(players_qs=players_qs, physios_qs=physios_qs)
    
    context = {
        'form': form,
    }
    return render(request, 'player_app/organization/injury_create.html', context)

@require_GET
def get_player_info(request, player_id):
    # AJAX view to return player info as JSON
    try:
        player = Player.objects.get(id=player_id)
        data = {
            'name': player.name,
            'gender': player.gender,
            'age': player.age,
            'email': player.email,
            'role': player.role,
            'bowling_style': player.bowling_style,
            'player_id': player.id,
            'date_of_birth': player.date_of_birth.strftime('%b %d, %Y') if player.date_of_birth else '',
            'contact_number': player.primary_contact_number,
            'handedness': player.handedness,
            'batting_style': player.batting_style,
            'district_cricket_association_id': player.state,
            'photo_url': player.image.url if player.image else 'https://randomuser.me/api/portraits/men/32.jpg',
        }
        return JsonResponse({'success': True, 'player': data})
    except Player.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Player not found.'}, status=404)


@login_required
def organization_injury_edit(request, pk):
    organization = get_object_or_404(Organization, user=request.user)
    injury = get_object_or_404(Injury, pk=pk, player__organization=organization)
    players_qs = Player.objects.filter(organization=organization)
    physios_qs = Staff.objects.filter(organization=organization, role__iexact='Physio')

    # Store original values for fields you want to track (add more fields as needed)
    original_data = {
        "name": injury.name,
        # "injury_type": injury.injury_type,
        "severity": injury.severity,
        "status": injury.status,
        "injury_date": injury.injury_date,
        "expected_date_of_return": injury.expected_date_of_return,
        "affected_body_part": injury.affected_body_part,
        "nature_of_injury": injury.nature_of_injury,
        # "body_segment": injury.body_segment,
        "player_status": injury.player_status,
        "venue": injury.venue,
        "notes": injury.notes,
        "reported_by_id": injury.reported_by_id,
        # Add other fields you want to track changes for
    }

    if request.method == "POST":
        form = InjuryFormUpdate(request.POST, instance=injury, players_qs=players_qs, physios_qs=physios_qs)
        if form.is_valid():
            updated_injury = form.save(commit=False)

            changed_fields = []
            for field, old_value in original_data.items():
                new_value = getattr(updated_injury, field)
                # For foreign keys, compare IDs
                if field.endswith('_id'):
                    if new_value != old_value:
                        changed_fields.append(f"{field[:-3]} changed from '{old_value}' to '{new_value}'")
                else:
                    if new_value != old_value:
                        changed_fields.append(f"{field} changed from '{old_value}' to '{new_value}'")

            updated_injury.save()

            # Only create log if something changed
            if changed_fields:
                InjuryActivityLog.objects.create(
                    injury=updated_injury,
                    actor=request.user,
                    action='updated',
                    details="; ".join(changed_fields)
                )
            else:
                pass

            return redirect('organization_injury_list')
    else:
        form = InjuryFormUpdate(instance=injury, players_qs=players_qs, physios_qs=physios_qs)
    return render(request, "player_app/organization/organization_injury_edit.html", {"form": form, "injury": injury})

from itertools import chain
from operator import attrgetter
@login_required
def organization_injury_detail(request, pk):
    organization = get_object_or_404(Organization, user=request.user)
    injury = get_object_or_404(Injury, pk=pk, player__organization=organization)
    player = get_object_or_404(Player, pk=injury.player.pk, organization=organization)

    # Fetch medical documents related to this injury
    medical_document = injury.documents.select_related('user').order_by('-date', '-uploaded_at')
    medical_documents = medical_document.filter(view_option__in=["injury_profile", "injury_only"])

    availability_form = PlayerAvailabilityForm(instance=injury)

    # Fetch activity logs: injury logs and medical logs combined and sorted by their date fields
    injury_logs = injury.activity_logs.select_related('actor').all()
    medical_logs = player.activity_logs.select_related('user', 'document').all()
    combined_logs = sorted(
        chain(injury_logs, medical_logs),
        key=lambda log: getattr(log, 'created_at', getattr(log, 'timestamp', None)),
        reverse=True
    )
    old_injury_status = injury.status
    old_player_statuss = injury.player.player_status
    oldabc = injury.player_status

    # print(f"Initial Injury Status: {old_injury_status}, Player Status: {old_player_status}")
    if request.method == 'POST':    
        if "availability_submit" in request.POST:
            availability_form = PlayerAvailabilityForm(request.POST, instance=injury)
            if availability_form.is_valid():
                old_status = injury.status
                old_player_status = injury.player_status
                availability_form.save()
                injury.refresh_from_db()  # <- refresh injury instance to get updated values
                new_status = injury.status
                new_player_status = injury.player.player_status
                
                status_changed = old_injury_status != new_status
                player_status_changed = old_player_statuss != new_player_status
                
                # Then build your message_parts and create log based on these refreshed values
                message_parts = []
                
                if status_changed:
                    message_parts.append(f"Status changed from '{old_injury_status}' to '{new_status}'")
                if player_status_changed:
                    message_parts.append(f"Participation changed from '{old_player_statuss}' to '{new_player_status}'")
                if message_parts:
                    InjuryActivityLog.objects.create(
                        injury=injury,
                        actor=request.user,
                        action="updated availability",
                        details='; '.join(message_parts)
                    )
                return redirect('organization_injury_detail', pk=injury.pk)

        else:
            form = MedicalDocumentFormN(request.POST, request.FILES, injury=injury)
            if form.is_valid():
                doc = form.save(commit=False)
                doc.injury = injury
                doc.player = player
                doc.user = request.user
                doc.save()
                MedicalActivityLog.objects.create(
                    player=player,
                    document=doc,
                    user=request.user,
                    activity_type='UPLOAD',
                    description=f"Uploaded document '{doc.title}'"
                )
                return redirect('organization_injury_detail', pk=injury.pk)
    else:
        form = MedicalDocumentFormN(injury=injury)

    return render(
        request,
        "player_app/organization/organization_injury_detail.html",
        {
            "injury": injury,
            "activity_logs": combined_logs,
            "player": player,
            "medical_documents": medical_documents,
            "form": form,
            "availability_form": availability_form,
        }
    )



@login_required
def organization_injury_delete(request, pk):
    organization = get_object_or_404(Organization, user=request.user)
    injury = get_object_or_404(Injury, pk=pk, player__organization=organization)
    if request.method == "POST":
        InjuryActivityLog.objects.create(
            injury=injury,
            actor=request.user,
            action='deleted',
            details='Injury record deleted'
        )
        injury.delete()
        return redirect("organization_injury_list")
    return render(request, "player_app/organization/organization_injury_confirm_delete.html", {"injury": injury})

@login_required
def organization_injury_export(request):
    organization = get_object_or_404(Organization, user=request.user)
    injuries = Injury.objects.filter(player__organization=organization).select_related("player", "reported_by")
    # _Apply filters as in your list view if you want filtered export_
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="injuries.csv"'
    writer = csv.writer(response)
    writer.writerow([
        "ID", "Player", "Severity", "Type", "Status", "Title", "Injury Date",
        "Expected Return", "Reported By", "Body Region", "Body Segment", "Venue", "Notes"
    ])
    for inj in injuries:
        writer.writerow([
            inj.id, inj.player.name, inj.get_severity_display(), inj.injury_type, inj.get_status_display(),
            inj.name, inj.injury_date, inj.expected_date_of_return, getattr(inj.reported_by, 'name', ''),
            inj.affected_body_part, inj.body_segment, inj.venue, inj.notes
        ])
    return response

# Camps & Tournaments views
def organization_camps_tournaments(request):
    """
    Displays a list of camps/tournaments.
    - Super Admins can view all camps.
    - Staff can see camps based on their permissions.
    - Users in an organization can see only camps from their organization.
    """

    # Super Admin: Sees all camps
    if request.user.is_superuser:
        organizations = Organization.objects.all()
        selected_org = request.GET.get('organization')

        if selected_org:
            camps = CampTournament.objects.filter(organization_id=selected_org, is_deleted=False)
            male_players = Player.objects.filter(organization_id=selected_org, gender='Male')
            female_players = Player.objects.filter(organization_id=selected_org, gender='Female')
        else:
            camps = CampTournament.objects.filter(is_deleted=False)
            male_players = Player.objects.filter(gender='Male')
            female_players = Player.objects.filter(gender='Female')

    # Staff with permission: Sees all camps
    elif hasattr(request.user, 'staff') and request.user.staff.view_camps_tournaments:
        org = request.user.staff.organization
        camps = CampTournament.objects.filter(is_deleted=False)
        male_players = Player.objects.filter(organization=org, gender='Male')
        female_players = Player.objects.filter(organization=org, gender='Female')

    # Regular users: See only camps in their organization
    elif hasattr(request.user, 'organization') and request.user.organization:
        org = request.user.organization  # Fixed: was request.user.staff.organization
        camps = CampTournament.objects.filter(organization=org, is_deleted=False)
        male_players = Player.objects.filter(organization=org, gender='Male')
        female_players = Player.objects.filter(organization=org, gender='Female')

    # No Access: Deny access
    else:
        return HttpResponseForbidden(
            "You must belong to an organization or have the necessary permissions to view camps and tournaments.")
    
    camps_qs = camps.annotate(
        year=ExtractYear('start_date')
    ).order_by('-start_date')

    # Group camps by year range (year to year+1)
    camps_by_range = defaultdict(list)
    for camp in camps_qs:
        year = camp.year
        range_key = f"{year}-{year + 1}" if year else "No Year"
        camps_by_range[range_key].append(camp)

    # Sort year ranges descending
    sorted_ranges = sorted(camps_by_range.items(), key=lambda x: int(x[0].split('-')[0]), reverse=True)
    
    return render(request, 'player_app/organization/organization_camps_tournaments.html', {
        'camps': camps,
        'camps_by_range': dict(sorted_ranges),  # Changed from camps_by_year
        'male_players': male_players,
        'female_players': female_players,
        'organizations': organizations if request.user.is_superuser else None
    })



def organization_edit_camp(request, camp_id):
    """
    Handles editing a specific camp/tournament, including participants.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)

    if request.method == 'POST':
        camp.name = request.POST.get('name')
        camp.camp_type = request.POST.get('camp_type')

        # Keep existing start date (only allow editing end date)
        end_date = request.POST.get('end_date')
        if end_date:
            camp.end_date = end_date

        camp.venue = request.POST.get('venue')

        # Update participants (only from the same organization)
        participant_ids = request.POST.getlist('participants')
        camp.participants.set(participant_ids)

        camp.save()

        # Log the update activity
        CampActivity.objects.create(
            camp=camp,
            action='updated',
            performed_by=request.user,
            details=f"Camp/Tournament '{camp.name}' was updated."
        )

        messages.success(request, "Camp/Tournament updated successfully!")
        return redirect('organization_camp_detail', camp_id=camp.id)

    # Get only players from the same organization
    participants = Player.objects.filter(organization=camp.organization)

    return render(request, 'player_app/organization/organization_edit_camp.html', {
        'camp': camp,
        'participants': participants
    })

@login_required
def organization_create_camp(request):
    """
    Create a new Camp/Tournament.
    """
    organizations = None
    players = Player.objects.none()  # Default to no players

    # Super Admin: Get all organizations
    if request.user.is_superuser:
        organizations = Organization.objects.all()
        players = Player.objects.all()  # Super Admins can see all players

    # Organization Admins: Check if user has an organization directly
    elif hasattr(request.user, "organization") and request.user.organization:
        organization = request.user.organization
        players = Player.objects.filter(organization=organization)

    # Staff Members: Ensure they have a staff profile before accessing
    elif hasattr(request.user, "staff") and request.user.staff:
        organization = request.user.staff.organization
        players = Player.objects.filter(organization=organization)

    else:
        return HttpResponseForbidden(
            "You must be a Super Admin, Organization Admin, or a Staff member to create a camp.")

    players_grouped = {'M': {}, 'F': {}}
    for player in players:
        gender = player.gender
        age_cat = player.age_category
        if gender in ['M', 'F']:
            if age_cat not in players_grouped[gender]:
                players_grouped[gender][age_cat] = []
            players_grouped[gender][age_cat].append({'id': player.id, 'name': player.name})

    players_grouped_json = json.dumps(players_grouped)

    if request.method == "POST":
        name = request.POST.get("name")
        camp_type = request.POST.get("camp_type")
        start_date = request.POST.get("start_date")
        gender = request.POST.get("gender")
        age_category = request.POST.get("age_category")
        venue = request.POST.get("venue")


        # Super Admin: Allow selecting an organization
        if request.user.is_superuser:
            organization_id = request.POST.get("organization")
            organization = get_object_or_404(Organization, id=organization_id)
        elif hasattr(request.user, "organization") and request.user.organization:
            # Organization Admins: Auto-set organization
            organization = request.user.organization
        elif hasattr(request.user, "staff") and request.user.staff:
            # Staff: Auto-set organization from staff profile
            organization = request.user.staff.organization
        else:
            return HttpResponseForbidden("You do not have permission to create a camp.")

        # Create the camp/tournament
        camp = CampTournament.objects.create(
            name=name,
            camp_type=camp_type,
            start_date=start_date,
            gender=gender,
            age_category=age_category,
            venue=venue,
            organization=organization,
            created_by=request.user
        )

        # Add participants (Only players from the same organization)
        selected_participants = request.POST.getlist("participants")
        valid_participants = players.filter(id__in=selected_participants)  # Ensure only valid participants
        camp.participants.set(valid_participants)

        messages.success(request, "Camp/Tournament created successfully!")
        return redirect("organization_camps_tournaments")  # Redirect after creation
    
    return render(request, "player_app/organization/organization_create_camp.html", {
        "organizations": organizations,  # Super Admin can select
        "players": players,    # Filtered players for the user
        "players_grouped_json": players_grouped_json,
    })

def organization_delete_camp(request, camp_id):
    """
    Allows authorized users to soft-delete a camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)  # Remove organization filtering
    camp.is_deleted = True
    camp.save()
    # Log the deletion activity
    CampActivity.objects.create(
        camp=camp,
        action='deleted',
        performed_by=request.user,
        details=f"Camp/Tournament '{camp.name}' was deleted."
    )
    messages.success(request, 'Camp/Tournament deleted successfully.')
    return redirect('organization_camps_tournaments')

def organization_camp_detail(request, camp_id):
    """
    Displays details of a specific camp/tournament with S&C Logs, Injuries, and Tests.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)
    phase = get_object_or_404(CampTournament, id=camp_id, 
                              organization=request.user.organization)
    
    # Dict of all test querysets by phase
    test_data = {
        '10m': TenMeterTest.objects.filter(phase=phase).select_related('player')[:50],
        '20m': TwentyMeterTest.objects.filter(phase=phase).select_related('player')[:50],
        '40m': FortyMeterTest.objects.filter(phase=phase).select_related('player')[:50],
        'YoYo': YoYoTest.objects.filter(phase=phase).select_related('player')[:50],
        'SBJ': SBJTest.objects.filter(phase=phase).select_related('player')[:50],
        'Run A 3': RunA3Test.objects.filter(phase=phase).select_related('player')[:50],
        '1 Mile': OneMileTest.objects.filter(phase=phase).select_related('player')[:50],
        'Push-ups': PushUpsTest.objects.filter(phase=phase).select_related('player')[:50],
        '2 KM': TwoKmTest.objects.filter(phase=phase).select_related('player')[:50],
    }
    all_empty = not any(qs.exists() for qs in test_data.values())

    TEST_SPECIAL = {
        'S/L Glute Bridges': SLGluteBridges.objects.filter(phase=phase).select_related('player')[:50],
        'S/L Lunge Calf Raises': SLLungeCalfRaises.objects.filter(phase=phase).select_related('player')[:50],
        'MB Rotational Throws': MBRotationalThrows.objects.filter(phase=phase).select_related('player')[:50],
        'Copenhagen': CopenhagenTest.objects.filter(phase=phase).select_related('player')[:50],
        'S/L Hop': SLHopTest.objects.filter(phase=phase).select_related('player')[:50],
    }
    
    other_test_models = {
        "CMJ Scores": CMJTest,
        "Anthropometry Test": AnthropometryTest,
        "Blood Work": BloodTest,
        "DEXA Scan Test": DexaScanTest,
        "MSK Injury Assessment": MSKInjuryAssessment,
    }

    OTHER_TEST = {}
    for label, model in other_test_models.items():
        qs = model.objects.filter(phase=phase).select_related('player__organization')[:50]

        rows = []
        headers = None

        for obj in qs:
            row_dict = obj_to_row(obj)
            
            # ✅ Fix player name display (handles both player/player_id)
            if 'player_id' in row_dict:
                row_dict['player_id'] = obj.player.name if obj.player else 'Unknown Player'
            elif 'player' in row_dict:
                row_dict['player'] = obj.player.name if obj.player else 'Unknown Player'
            
            if headers is None:
                headers = list(row_dict.keys())
            rows.append([row_dict[h] for h in headers])

        OTHER_TEST[label] = {
            "headers": headers or [],
            "rows": rows,
        }

    # 🔥 NEW: S&C Logs (TOP section) - Optimized with prefetch_related for activities
    snc_logs = DailySncLogCamps.objects.filter(team=phase).prefetch_related('activities')[:10]

    # 🔥 NEW: Injuries (MIDDLE section) - Optimized
    camp_injuries = Injury.objects.filter(
        camp_tournament=phase,
        player__organization=request.user.organization
    ).select_related('player').order_by('-injury_date')[:10]

    context = {
        'snc_logs': snc_logs,
        'camp_injuries': camp_injuries,
        "phase": phase,
        "test_data": test_data,
        "total_tests": sum(len(tests) for tests in test_data.values()),
        "all_empty": all_empty,
        "TEST_SPECIAL": TEST_SPECIAL,
        "OTHER_TEST": OTHER_TEST,
        'camp': camp,
    }
    return render(request, 'player_app/organization/organization_camp_detail.html', context)



def obj_to_row(obj):
    EXCLUDED_KEYS = {"id", "created_at", "reported_by_designation", "phase", "updated_at","gender","category"}

    data = obj.__dict__.copy()
    # Remove private + excluded keys
    for key in list(data.keys()):
        if key.startswith("_") or key in EXCLUDED_KEYS:
            data.pop(key, None)
    return data

@login_required
def phase_tests_view(request, id):
    phase = get_object_or_404(CampTournament, id=id, 
                             organization=request.user.organization)
    
    # Dict of all test querysets by phase
    test_data = {
        '10m': TenMeterTest.objects.filter(phase=phase).select_related('player')[:50],
        '20m': TwentyMeterTest.objects.filter(phase=phase).select_related('player')[:50],
        '40m': FortyMeterTest.objects.filter(phase=phase).select_related('player')[:50],
        'YoYo':YoYoTest.objects.filter(phase=phase).select_related('player')[:50],
        'SBJ': SBJTest.objects.filter(phase=phase).select_related('player')[:50],
        'Run A 3': RunA3Test.objects.filter(phase=phase).select_related('player')[:50],
        # 'Run A 3x6':
        '1 Mile': OneMileTest.objects.filter(phase=phase).select_related('player')[:50],
        'Push-ups': PushUpsTest.objects.filter(phase=phase).select_related('player')[:50],
        '2 KM': TwoKmTest.objects.filter(phase=phase).select_related('player')[:50],
    }
    all_empty = not any(qs.exists() for qs in test_data.values())

    TEST_SPECIAL = {
        'S/L Glute Bridges': SLGluteBridges.objects.filter(phase=phase).select_related('player')[:50],
        'S/L Lunge Calf Raises': SLLungeCalfRaises.objects.filter(phase=phase).select_related('player')[:50],
        'MB Rotational Throws': MBRotationalThrows.objects.filter(phase=phase).select_related('player')[:50],
        'Copenhagen': CopenhagenTest.objects.filter(phase=phase).select_related('player')[:50],
        'S/L Hop': SLHopTest.objects.filter(phase=phase).select_related('player')[:50],
    }
    
    other_test_models = {
        "CMJ Scores": CMJTest,
        "Anthropometry Test": AnthropometryTest,
        "Blood Work": BloodTest,
        "DEXA Scan Test": DexaScanTest,
        "MSK Injury Assessment": MSKInjuryAssessment,
    }

    OTHER_TEST = {}
    for label, model in other_test_models.items():
        qs = model.objects.filter(phase=phase).select_related("player")[:50]

        rows = []
        headers = None

        for obj in qs:
            row_dict = obj_to_row(obj)          # dict without excluded keys
            if headers is None:
                headers = list(row_dict.keys())  # order fixed from first row
            rows.append([row_dict[h] for h in headers])

        OTHER_TEST[label] = {
            "headers": headers or [],
            "rows": rows,
        }

    context = {
        "phase": phase,
        "test_data": test_data,
        "total_tests": sum(len(tests) for tests in test_data.values()),
        "all_empty": all_empty,
        "TEST_SPECIAL": TEST_SPECIAL,
        "OTHER_TEST": OTHER_TEST,
    }
    return render(request, "player_app/tests/phase_test_data.html", context)

def phase_test(request,id):
    phase = get_object_or_404(CampTournament, id=id, 
                            organization=request.user.organization)
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']

    request.session['phase_id_test'] = id
    print("session id :",request.session['phase_id_test'])
    
    return render(request,"player_app/tests/phase_test.html",{"phase":phase})

# Daily Activities of S&C Coach's Log
ACTIVITY_NAMES = [
    "Match",
    "Match Simulation",
    "Practice",
    "Fielding",
    "Strength",
    "Conditioning",
    "Rest",
    "Travelling",
    "Fitness Testing",
    "Screening / Assessments",
]

# Function for the view
import re
def _slugify_activity(name: str) -> str:
    slug = name.lower()
    slug = re.sub(r"\s+/\s+", "_", slug)
    slug = re.sub(r"\s+", "_", slug)
    slug = re.sub(r"[^a-z0-9_]", "", slug)
    return slug

# Daily Activity log view for S&C Coach's
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from .models import CampTournament, Organization, DailySncLogCamps, DailyActivityCamps, Injury, Player, Staff

def _slugify_activity(activity_name):
    """Helper to slugify activity names (same as template JS)"""
    return activity_name.lower().replace(" / ", "_").replace(" ", "_").replace("/", "_")

def daily_activity_coach_log(request, id):
    camp = get_object_or_404(CampTournament, id=id)
    organization = get_object_or_404(Organization, user=request.user)
    players_qs = Player.objects.filter(organization=organization)
    physios_qs = Staff.objects.filter(organization=organization, role__iexact='Physio')
    
    if request.method == "POST":
        with transaction.atomic():  # Ensure all data saves atomically
            team = request.POST.get("team", "").strip()
            coach_name = request.POST.get("coach_name", "").strip()
            date = request.POST.get("session_date")
            end_date = request.POST.get("session_date_end")
            concerns = request.POST.get("concerns", "").strip()
            
            niggles_value = request.POST.get("niggles")
            niggles = niggles_value == "yes"
            
            recovery_list = request.POST.getlist("recovery")
            recovery_sessions = ",".join(recovery_list)
            
            if not (team and coach_name and date):
                messages.error(request, "Team, coach name and date are required.")
                return render(
                    request,
                    "player_app/camps/daily_activity.html",
                    {
                        "activities": ACTIVITY_NAMES, 
                        "camp": camp, 
                        "organization": organization,
                        "players_qs": players_qs,
                        "physios_qs": physios_qs,
                    },
                )
            
            teams = CampTournament.objects.get(name=team)
            
            # Create/update main log entry
            log = DailySncLogCamps.objects.create(
                team=teams,
                date=date,
                end_date=end_date,
                user=request.user,
                coach_name=coach_name,
                concerns=concerns,
                niggles=niggles,
                recovery_sessions=recovery_sessions,
            )
                        
            # Save activities
            for activity_name in ACTIVITY_NAMES:
                slug = _slugify_activity(activity_name)
                duration_key = f"activities[{slug}][duration]"
                intensity_key = f"activities[{slug}][intensity]"
                
                duration = request.POST.get(duration_key, "").strip()
                intensity = request.POST.get(intensity_key, "").strip()
                
                if duration:  # Only save if duration selected
                    DailyActivityCamps.objects.create(
                        log=log,
                        activity_name=activity_name,
                        duration=duration,
                        intensity=intensity,
                    )
            
            # **NEW: Handle Injury Data when niggles == True**
            if niggles:
                player_id = request.POST.get('player')
                reported_by_id = request.POST.get('reported_by')
                injury_date = request.POST.get('injury_date')
                name = request.POST.get('title')  # Title field
                nature_of_injury = request.POST.get('nature_of_injury')
                diagnosis_date = request.POST.get('diagnosis_date')
                severity_rating = request.POST.get('severity')
                venue = request.POST.get('venue')
                type_of_activity = request.POST.get('type_of_activity')
                notes = request.POST.get('notes')
                action_taken = request.POST.get('action_taken')
                player_status = request.POST.get('player_status')
                expected_date_of_return = request.POST.get('expected_date_of_return')
                side = request.POST.get('side')
                camp_tournament = camp.id
                team = teams
                organization = organization
                affected_body_parts = request.POST.getlist('affected_body_part')
                
                print(player_id, reported_by_id, injury_date, name, nature_of_injury, diagnosis_date, severity_rating,
                      venue, type_of_activity, notes, action_taken, player_status, expected_date_of_return, side, affected_body_parts)
                phase_obj = get_object_or_404(CampTournament, id=camp.id)
                player = get_object_or_404(Player, id=player_id)
                
                print(phase_obj, player)
                # Staff → User
                staff_obj = get_object_or_404(Staff, id=reported_by_id)
                reported_by_user = staff_obj.user 
                # Get affected body parts (handles multiple selections)
                print("Reported by user:", reported_by_user)
                
                # Validate required injury fields
                if player_id:  # Player & Title required
                    try:
                        injury = Injury.objects.create(
                            player_id=player_id,
                            reported_by_id=reported_by_id or None,
                            injury_date=injury_date or None,
                            name=name,
                            nature_of_injury=nature_of_injury or '',
                            diagnosis_date=diagnosis_date or None,
                            severity_rating=severity_rating or 0,
                            venue=venue or '',
                            type_of_activity=type_of_activity or '',
                            notes=notes or '',
                            action_taken=action_taken or '',
                            player_status=player_status or 'no participation',
                            expected_date_of_return=expected_date_of_return or None,
                            side=side or 'bilateral',
                            camp_tournament_id=camp.id,
                            affected_body_part=affected_body_parts,
                            team=team,
                        )
                        
                        # Save body parts (assuming Injury model has ManyToMany or JSONField)
                        print("Injury created with ID:", injury.id)
                        messages.success(request, f"Daily log AND injury report saved successfully!")
                        
                    except Exception as e:
                        print("Error saving injury:", e)
                        messages.error(request, f"Injury save failed: {str(e)}")
                else:
                    messages.warning(request, "Injury reported but missing required fields (Player/Title). Log saved.")
            
            else:
                messages.success(request, "Daily S&C camp log saved.")
            
            return redirect("daily_snc_camp_detail", pk=log.pk)
    
    # GET request - render form
    return render(
        request,
        "player_app/camps/daily_activity.html",
        {
            "activities": ACTIVITY_NAMES,
            "camp": camp,
            "organization": organization,
            "players_qs": players_qs,
            "physios_qs": physios_qs,
        },
    )


# Daily Activity log view page
def daily_snc_camp_detail(request, pk):
    log = get_object_or_404(DailySncLogCamps, pk=pk)
    activities = log.activities.all().order_by("id")

    recovery_list = []
    if log.recovery_sessions:
        raw_items = [item for item in log.recovery_sessions.split(",") if item]
        for item in raw_items:
            label = item.replace("_", " ").strip().title()
            recovery_list.append(label)
    return render(request,"player_app/camps/daily_log_camp_detail.html", {"log": log, "activities": activities,"recovery_list": recovery_list,},)


def daily_snc_camp_logs_list(request, camp_id):
    """List all daily S&C logs for a camp with sorting by date."""
    camp = get_object_or_404(CampTournament, id=camp_id)
    organization = get_object_or_404(Organization, user=request.user)
    
    # Get all logs for this camp's teams
    camp_teams = CampTournament.objects.filter(id=camp_id).values_list('name', flat=True)
    logs = DailySncLogCamps.objects.filter(
        team__name__in=camp_teams
    ).select_related('user').prefetch_related('activities').order_by('-date')
    
    # Sorting by date (newest first by default)
    sort_order = request.GET.get('sort', '-date')
    logs = logs.order_by(sort_order)
    
    context = {
        'camp': camp,
        'organization': organization,
        'logs': logs,
        'sort_order': sort_order,
    }
    return render(request, 'player_app/camps/daily_snc_camp_logs_list.html', context)

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Prefetch
from datetime import datetime, timedelta
from .models import DailySncLogCamps, CampTournament, DailyActivityCamps

from datetime import datetime, date  # Add 'date' import

@login_required
def snc_camps_dashboard(request):
    user_org = getattr(request.user, "organization", None)
    
    logs = CampTournament.objects.filter(organization=user_org).order_by('name')
    
    context = {
        'logs': logs,
        'report_date': '',
        'camp': '',
        'selected_camp': None,
        'report_data': None,
    }
    
    if request.method == 'POST':
        report_date = request.POST.get('report_date')
        camp_id = request.POST.get('camp')
        report_date_end = request.POST.get('report_date_end')
        

        context['report_date'] = report_date
        context['camp'] = camp_id
        
        if report_date and camp_id:
            selected_camp = get_object_or_404(CampTournament, id=camp_id, organization=user_org)
            
            # ✅ FIXED: report_date → TODAY (Jan 12, 2026)
            start_date = datetime.strptime(report_date, '%Y-%m-%d').date()
            end_date = date.today()  # CURRENT DATE (2026-01-12)

            
            # Filter logs between selected date and today
            report_data = DailySncLogCamps.objects.filter(
                team=selected_camp,
                date__range=[start_date, report_date_end]  # 2025-12-01 → 2026-01-12
            ).prefetch_related('activities').select_related('user').order_by('-date')
            niggles_count = report_data.filter(niggles=True).count()
            
            
            context['selected_camp'] = selected_camp
            context['report_data'] = report_data
            context['date_range'] = f"{start_date} → Today ({report_date_end})"
            context['report_date_end'] = report_date_end
            context['niggles_count']=niggles_count
    
    return render(request, 'player_app/camps/snc_dashboard.html', context)

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseBadRequest
from datetime import datetime, date
from dateutil import parser

def wellness_dashboard(request):
    """
    Daily Wellness Report Dashboard - Single Date Filter Only
    """
    user_org = getattr(request.user, "organization", None)
    camps = CampTournament.objects.filter(organization=user_org).order_by('name')
    
    report_data = None
    selected_camp = None
    report_date = None
    
    if request.method == 'POST':
        try:
            report_date = request.POST.get('report_date')
            camp_id = request.POST.get('camp')
            print("Received POST data:", report_date, camp_id)
            
            if not all([report_date, camp_id]):
                return HttpResponseBadRequest("Missing required parameters")
            
            # Parse single date
            selected_date = parser.parse(report_date).date()
            selected_camp = get_object_or_404(CampTournament, id=camp_id)
            print(f"Parsed date: {selected_date}")
            
            # Filter wellness logs for selected CAMP and EXACT date - ALL PLAYERS
            report_data = DailyWellnessTest.objects.filter(
                phase=selected_camp,
                date=selected_date
                # ✅ Shows ALL players who submitted wellness for this camp/date
            ).select_related('player', 'phase').order_by('player__name')
            
            print(f"Fetched {report_data.count()} wellness logs for camp {selected_camp.name} on {selected_date}")
            
        except Exception as e:
            logger.error(f"Error processing report request: {e}")
            report_data = []
    
    context = {
        'logs': camps,  # renamed from 'logs'
        'report_data': report_data,
        'selected_camp': selected_camp,
        'report_date': report_date,
    }
    
    return render(request, 'player_app/camps/daily_wellness_all.html', context)

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import DailyWellnessTest, CampTournament, Player


@login_required
def daily_wellness_camp_report(request, camp_id):
    """
    Show Daily Wellness Tests for SPECIFIC CAMP - players who are in camp AND completed tests
    """
    user_org = getattr(request.user, "organization", None)
    
    # Get the specific camp
    camp = get_object_or_404(
        CampTournament, 
        id=camp_id, 
        organization=user_org
    )
    
    # ✅ KEY FILTER: Players IN THIS CAMP who have DailyWellnessTest records
    wellness_tests = DailyWellnessTest.objects.filter(
        phase=camp  # Direct FK filter ✅
    ).select_related(
        'player', 'phase', 'created_by'
    ).order_by('-date', 'player__name')
    
    context = {
        'camp': camp,
        'wellness_tests': wellness_tests,
        'total_tests': wellness_tests.count(),
        'unique_players': wellness_tests.values('player').distinct().count(),
        'camp_id': camp_id,
    }

    
    return render(request, 'player_app/camps/wellness_report.html', context)


def player_wellness_report(request):
    """
    Player Wellness Report - Single Player + Date (All Camps/Tournaments)
    """
    user_org = getattr(request.user, "organization", None)
    players = Player.objects.filter(organization=user_org).order_by('name')
    
    report_data = None
    selected_player = None
    report_date = None
    
    if request.method == 'POST':
        try:
            report_date = request.POST.get('report_date')
            player_id = request.POST.get('player_id')
            print("Received POST data:", report_date, player_id)
            
            if not all([report_date, player_id]):
                return HttpResponseBadRequest("Missing required parameters")
            
            # Parse single date
            selected_date = parser.parse(report_date).date()
            selected_player = get_object_or_404(Player, id=player_id)
            print(f"Parsed date: {selected_date}, player: {selected_player}")
            
            # Filter wellness logs for selected PLAYER and EXACT date (ALL camps)
            report_data = DailyWellnessTest.objects.filter(
                player=selected_player,
                date=selected_date
                # ✅ No phase/camp filter - shows ALL camps/tournaments for player
            ).select_related('player', 'phase')
            
            print(f"Fetched {report_data.count()} wellness logs for player {selected_player} on {selected_date}")
            
        except Exception as e:
            logger.error(f"Error processing report request: {e}")
            report_data = []
    
    context = {
        'players': players,
        'report_data': report_data,
        'selected_player': selected_player,
        'report_date': report_date,
    }
    
    return render(request, 'player_app/camps/player-wellness-report.html', context)



# -----------------------------------------------------------------------------------------------------------

from django.db.models import Avg
from django.db.models import Min, Max
from collections import defaultdict
# @login_required
# def test_dashboard(request):
#     # Get user's organization (adjust as per your user model)
#     user_organization = getattr(request.user, 'organization', None)
#     if not user_organization:
#         return render(request, 'player_app/organization/test_dashboard.html', {
#             'error_message': "Your account is not linked to any organization.",
#         })

#     # Get players in user's organization
#     players_in_org = Player.objects.filter(organization=user_organization)

#     # Handle new test result form (restrict player queryset to org players)
#     if request.method == 'POST':
#         add_form = TestAndResultForm(request.POST)
#         add_form.fields['player'].queryset = players_in_org
#         if add_form.is_valid():
#             add_form.save()
#             return redirect('test_dashboard')
#     else:
#         add_form = TestAndResultForm()
#         add_form.fields['player'].queryset = players_in_org

#     # Handle filter form (restrict player queryset to org players)
#     filter_form = TestSummaryFilterForm(request.GET or None)
#     filter_form.fields['player'].queryset = players_in_org

#     # Base queryset filtered by org players only
#     qs = TestAndResult.objects.select_related('player').filter(player__in=players_in_org).order_by('player__name', 'test', 'date', 'id')

#     # Apply filters if valid form submitted
#     if filter_form.is_valid():
#         if filter_form.cleaned_data.get('player'):
#             qs = qs.filter(player=filter_form.cleaned_data['player'])
#         if filter_form.cleaned_data.get('test'):
#             qs = qs.filter(test=filter_form.cleaned_data['test'])

#     # Group trials by (player_id, test)
#     player_test_trials = defaultdict(list)
#     for trial in qs:
#         key = (trial.player.id, trial.test)
#         player_test_trials[key].append(trial)

#     # Build summary rows without serial yet
#     summary_rows = []
#     for (player_id, test), trials in player_test_trials.items():
#         if not trials:
#             continue

#         # Last two trials chronologically
#         last_two_trials = trials[-2:] if len(trials) >= 2 else trials[-1:]

#         trial_1 = last_two_trials[0].trial if len(last_two_trials) == 2 else None
#         trial_2 = last_two_trials[-1].trial

#         best_trial = min(t.trial for t in trials)

#         last_trial_obj = last_two_trials[-1]
#         # Individual Average: mean of all trials for this player and test
        
#         indv_average = sum(t.trial for t in trials) / len(trials) if trials else None
#         group_average = TestAndResult.objects.filter(test=test).aggregate(Avg('trial'))['trial__avg']
        
#         summary_rows.append({
#             'player_name': last_trial_obj.player.name,
#             'test': test,
#             'last_date': last_trial_obj.date,
#             'last_phase': last_trial_obj.phase,
#             'trial_1': trial_1,
#             'trial_2': trial_2,
#             'best_trial': best_trial,
#             'indv_average': indv_average,
#             'group_average': group_average,
#         })

#     # Group rows by test
#     summary_by_test = defaultdict(list)
#     for row in summary_rows:
#         summary_by_test[row['test']].append(row)

#     # Assign serial numbers per test table starting at 1
#     for test_name, rows in summary_by_test.items():
#         for idx, row in enumerate(rows, start=1):
#             row['serial'] = idx

#     context = {
#         'add_form': add_form,
#         'form': filter_form,
#         'summary_by_test': dict(summary_by_test),
#     }
#     return render(request, 'player_app/organization/test_dashboard.html', context)

@login_required
def add_test_result(request):
    organization = getattr(request.user, 'organization', None)

    if request.method == 'POST':
        form = TestAndResultForm(request.POST, organization=organization)
        if form.is_valid():
            final_level_value = form.cleaned_data.get('best')
            result_id = NomativeData.objects.get(final_level=float(final_level_value))
            result = model_to_dict(result_id)
            instance = form.save(commit=False)
            instance.distance_covered = float(result["total_distance"])
            instance.predicted_vo2max = float(result["approximately_vo2max"])
            instance.save()
            return redirect('test_results_main') 
    else:
        form = TestAndResultForm(organization=organization)

    return render(request, 'player_app/organization/test_add.html', {'form': form})



from django.shortcuts import render
from django.db.models import Count, Q
from .models import Player, Injury, Staff
from django.contrib.auth.decorators import login_required
from itertools import chain
from collections import OrderedDict
from itertools import chain
from django.db.models import Count, Q
from django.db.models.functions import ExtractYear

@login_required
def organization_dashboard_org(request):
    selected_category = request.GET.get('category', 'all')
    selected_gender = request.GET.get('gender', 'all')

    # Determine organization
    if request.user.role == "Staff":
        organization = request.user.staff.organization
    elif request.user.role == "OrganizationAdmin":
        organization = get_object_or_404(Organization, user=request.user)
    else:
        organization = None

    if not organization:
        return render(request, 'error.html', {'message': 'No organization access'})

    # Base querysets
    all_players = Player.objects.filter(organization=organization)
    all_injuries = Injury.objects.filter(player__organization=organization)

    # ✅ NEW: Build category_players for tooltips
    category_players = defaultdict(list)
    for player in all_players:
        category_players[player.age_category].append(player.name)
    category_players = dict(category_players)

    # Cards definitions - FIXED keys match age_category values
    CATEGORY_CARDS = OrderedDict([
        ("boys_under-15", {'gender': 'M', 'label': "B - U14"}),
        ("boys_under-16", {'gender': 'M', 'label': "B - U16"}),
        ("boys_under-19", {'gender': 'M', 'label': "B - U19"}),
        ("men_under-23",  {'gender': 'M', 'label': "B - U23"}),
        ("men_senior",    {'gender': 'M', 'label': "M - SENIOR"}),
        ("girls_under-15", {'gender': 'F', 'label': "G - U15"}),
        ("girls_under-19", {'gender': 'F', 'label': "G - U19"}),
        ("women_under-23", {'gender': 'F', 'label': "W - U23"}),
        ("women_senior",  {'gender': 'F', 'label': "W - SENIOR"}),
    ])

    category_cards = []
    for age_category, card_cat in CATEGORY_CARDS.items():
        players_qs = all_players.filter(
            gender=card_cat['gender'],
            age_category=age_category
        ).distinct()

        category_cards.append({
            'label': card_cat['label'],
            'age_category': age_category,  # ✅ CRITICAL for data-age-category
            'total': players_qs.count(),
            'full': players_qs.filter(player_status__iexact="full participation").count(),
            'limited': players_qs.filter(player_status__iexact="limited participation").count(),
            'none': players_qs.filter(player_status__iexact="no participation").count(),
            'active_injury': all_injuries.filter(player__in=players_qs, status='open').values('player_id').distinct().count(),
        })

    # Filtered data for tables
    players = all_players
    injuries = all_injuries.select_related('player', 'reported_by')

    if selected_category != 'all':
        players = players.filter(age_category=selected_category)
        injuries = injuries.filter(player__age_category=selected_category)
    if selected_gender != 'all':
        players = players.filter(gender=selected_gender)
        injuries = injuries.filter(player__gender=selected_gender)

    total_injuries_count = injuries.count()
    
    active_injuries = injuries.filter(status='open')
    active_injuries_count = active_injuries.count()

    recovered_injuries = injuries.filter(status='closed')  # or 'closed' etc.
    recovered_injuries_count = recovered_injuries.count()
    
    participation_counts = CampTournament.objects.filter(
        participants__in=players
    ).annotate(
        player_count=Count('participants', distinct=True)
    ).order_by('-player_count')

    players = players.prefetch_related('injuries__reported_by', 'camps')

    # Activity logs
    medical_logs = MedicalActivityLog.objects.filter(player__in=players).select_related('player', 'user', 'document')
    injury_logs = InjuryActivityLog.objects.filter(injury__in=injuries).select_related('injury', 'actor')
    player_logs = PlayerActivityLog.objects.filter(player__in=players).select_related('player', 'actor')
    
    for log in chain(medical_logs, injury_logs, player_logs):
        if hasattr(log, 'log_type'):
            continue
        if 'MedicalActivityLog' in str(type(log)):
            log.log_type = 'medical'
        elif 'InjuryActivityLog' in str(type(log)):
            log.log_type = 'injury'
        else:
            log.log_type = 'player'

    combined_logs = sorted(
        chain(medical_logs, injury_logs, player_logs),
        key=lambda log: getattr(log, 'timestamp', getattr(log, 'created_at', None)) or datetime.min,
        reverse=True
    )

    camps_qs = (
        CampTournament.objects.filter(participants__in=players)
        .annotate(
            year=ExtractYear('start_date'),
            player_count=Count('participants', filter=Q(participants__in=players), distinct=True),
        )
        .prefetch_related('participants')
        .order_by('-year', 'name')
    )

    

    camps_by_year_range = defaultdict(list)

    for camp in camps_qs:
        base_year = camp.year  # e.g. 2021
        if base_year is None:
            range_key = "Unknown"
        else:
            range_key = f"{base_year}-{base_year + 1}"  # e.g. "2021-2022"
        camp.player_names = ", ".join(p.name for p in camp.participants.all())
        camps_by_year_range[range_key].append(camp)

    # sort ranges by starting year desc
    camps_by_year_range = dict(
        sorted(
            camps_by_year_range.items(),
            key=lambda item: int(item[0].split('-')[0]) if item[0] != "Unknown" else -1,
            reverse=True,
        )
    )
        


    context = {
        'selected_category': selected_category,
        'selected_gender': selected_gender,
        'players': players,
        'total_injuries_count': total_injuries_count,
        'recovered_injuries_count': recovered_injuries_count,
        'active_injuries_count': active_injuries_count,
        'active_injuries': active_injuries,
        'participation_counts': participation_counts,
        'activity_logs': combined_logs,
        'category_cards': category_cards,
        'category_players': category_players, 
        'camps_by_year': camps_by_year_range,
    }
    return render(request, 'player_app/organization/organization_dashboard.html', context)


def logout_user(request):
    """
    Logs out the user and redirects to the login page.
    """
    logout(request)
    messages.success(request, "You have been logged out successfully.")
    return redirect('home')  # Adjust to your login URL name

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q, Prefetch
from django.http import QueryDict


@login_required
def players_by_category(request):
    """View to show all players filtered by age_category from dashboard card click"""
    age_category = request.GET.get('age_category')
    gender_filter = request.GET.get('gender', '')
    search_query = request.GET.get('search', '').strip()
    
    # Determine organization
    if request.user.role == "Staff":
        organization = request.user.staff.organization
    elif request.user.role == "OrganizationAdmin":
        organization = get_object_or_404(Organization, user=request.user)
    else:
        return render(request, 'error.html', {'message': 'No organization access'})

    # Base queryset with optimized prefetching
    players = Player.objects.filter(organization=organization).select_related().prefetch_related(
        Prefetch('injuries', queryset=Injury.objects.filter(status='open').only('id', 'status'), to_attr='open_injuries'),
        Prefetch('camps', queryset=CampTournament.objects.only('id', 'name'), to_attr='recent_camps')
    )

    # Filter by age_category (from dashboard card click)
    if age_category:
        players = players.filter(age_category=age_category)

    # Filter by gender
    if gender_filter:
        players = players.filter(gender=gender_filter)

    # Filter by search
    if search_query:
        players = players.filter(
            Q(name__icontains=search_query) |
            Q(injuries__name__icontains=search_query)
        ).distinct()

    # Category labels for display
    category_labels = {
        'boys_under-15': 'Boys U14',
        'boys_under-16': 'Boys U16', 
        'boys_under-19': 'Boys U19',
        'men_under-23': 'Men U23',
        'men_senior': 'Men Senior',
        'girls_under-15': 'Girls U15',
        'girls_under-19': 'Girls U19',
        'women_under-23': 'Women U23',
        'women_senior': 'Women Senior',
    }
    
    category_display = category_labels.get(age_category, 'All Players')
    if not age_category:
        category_display = f'All Players ({organization.name})'

    total_players = players.count()

    context = {
        'players': players,
        'age_category': age_category,
        'gender_filter': gender_filter,
        'search_query': search_query,
        'category_display': category_display,
        'total_players': total_players,
        'request': request,
    }
    return render(request, 'player_app/dashboard/players-by-category.html', context)





# --------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------


# View to export players to Excel

def export_players_to_excel(request):
    players = Player.objects.all()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Players'

    headers = [
        'name', 'aadhar number', 'batting style', 'bowling style', 'date of birth', 'email', 'gender',
        'guardian mobile number', 'guardian name', 'handedness', 'id card number', 'profile image',
        'medical certificates', 'primary contact number', 'relation', 'user_role', 'secondary contact number',
        'sports role', 'state', 'address', 'district', 'pincode', 'aadhar card upload', 'marksheets upload',
        'pan card upload', 'additional information', 'age category', 'allergies', 'disease', 'height',
        'nationality', 'position', 'team', 'weight'
    ]
    sheet.append(headers)

    # Make header row bold
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for player in players:
        sheet.append([
            player.name,
            player.aadhar_number,
            player.batting_style,
            player.bowling_style,
            player.date_of_birth,
            player.email,
            player.gender,
            player.guardian_mobile_number,
            player.guardian_name,
            player.handedness,
            player.id_card_number,
            player.image.url if player.image else 'N/A',
            player.medical_certificates.url if player.medical_certificates else 'N/A',
            player.primary_contact_number,
            player.relation,
            player.user_role,
            player.secondary_contact_number,
            player.sports_role,
            player.state,
            player.address,
            player.district,
            player.pincode,
            player.aadhar_card_upload.url if player.aadhar_card_upload else 'N/A',
            player.marksheets_upload.url if player.marksheets_upload else 'N/A',
            player.pan_card_upload.url if player.pan_card_upload else 'N/A',
            player.additional_information,
            player.age_category,
            player.allergies,
            player.disease,
            player.height,
            player.nationality,
            player.position,
            player.team,
            player.weight
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=players.xlsx'
    workbook.save(response)

    return response


def upload_medical_documents(request, player_id):
    player = get_object_or_404(Player, id=player_id)

    if request.method == 'POST':
        form = MultipleMedicalDocumentsForm(request.POST, request.FILES)

        if form.is_valid():
            files = request.FILES.getlist('documents')  # Get multiple files as a list

            for file in files:
                MedicalDocument.objects.create(player=player, document=file)  # Save each document

            messages.success(request, "Medical documents uploaded successfully!")
            return redirect('player_detail', pk=player.id)
        else:
            messages.error(request, "Error uploading documents.")

    form = MultipleMedicalDocumentsForm()
    return render(request, 'player_app/upload_medical_documents.html', {'form': form, 'player': player})


# View to import players to Excel

import logging

logger = logging.getLogger(__name__)

from django.db.models import Max
from accounts.models import CustomUser

from django.db import IntegrityError
from player_app.models import Player
from accounts.models import CustomUser
import pandas as pd
import logging

# Set up logging to track the process
logger = logging.getLogger(__name__)


def upload_file(request):
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES["file"]
            df = pd.read_excel(excel_file)

            # Log column names and sample data for debugging
            logger.info(f"DataFrame columns: {df.columns}")
            logger.info(f"DataFrame head: \n{df.head()}")

            # Get the highest existing user_id (Player's ID) from the database
            max_user_id = Player.objects.aggregate(Max('user_id'))['user_id__max'] or 0
            next_user_id = max_user_id + 1  # This will be the next available user_id

            for _, row in df.iterrows():
                try:
                    user_id = row.get("user_id", None)

                    if not user_id:  # If user_id is missing, auto-increment from next available user_id
                        user_id = next_user_id
                        next_user_id += 1

                    # Check if the user with this user_id already exists in the CustomUser model
                    user_exists = CustomUser.objects.filter(id=user_id).exists()

                    while user_exists:  # If user_id already exists, try the next one
                        user_id = next_user_id
                        next_user_id += 1
                        user_exists = CustomUser.objects.filter(id=user_id).exists()

                    # If user doesn't exist, create a new user
                    user = CustomUser.objects.create(
                        id=user_id,  # Explicitly setting the user_id to avoid auto-increment conflict
                        username=row.get("name", ""),
                        password=row.get("password", ""),
                    )

                    # Prepare the data for creating Player instance
                    player_data = {
                        "user": user,
                        "name": row.get("name", ""),
                        "email": row.get("email", ""),
                        "primary_contact_number": row.get("primary_contact_number", ""),
                        "secondary_contact_number": row.get("secondary_contact_number", ""),
                        "date_of_birth": row.get("date_of_birth", None),
                        "pincode": row.get("pincode", ""),
                        "address": row.get("address", ""),
                        "nationality": row.get("nationality", ""),
                        "gender": row.get("gender", ""),
                        "state": row.get("state", ""),
                        "district": row.get("district", ""),
                        "role": row.get("role", ""),
                        "batting_style": row.get("batting_style", ""),
                        "bowling_style": row.get("bowling_style", ""),
                        "handedness": row.get("handedness", ""),
                        "aadhar_number": row.get("aadhar_number", ""),
                        "sports_role": row.get("sports_role", ""),
                        "id_card_number": row.get("id_card_number", ""),
                        "weight": float(row["weight"]) if pd.notna(row.get("weight")) else None,
                        "height": float(row["height"]) if pd.notna(row.get("height")) else None,
                        "age_category": row.get("age_category", ""),
                        "team": row.get("team", ""),
                        "position": row.get("position", ""),
                        "guardian_name": row.get("guardian_name", ""),
                        "relation": row.get("relation", ""),
                        "guardian_mobile_number": row.get("guardian_mobile_number", ""),
                        "disease": row.get("disease", ""),
                        "allergies": row.get("allergies", ""),
                        "additional_information": row.get("additional_information", ""),
                        # File handling for specific fields
                        "image": row["image"] if "image" in row and pd.notna(row["image"]) else None,
                        "medical_certificates": row[
                            "medical_certificates"] if "medical_certificates" in row and pd.notna(
                            row["medical_certificates"]) else None,
                        "aadhar_card_upload": row["aadhar_card_upload"] if "aadhar_card_upload" in row and pd.notna(
                            row["aadhar_card_upload"]) else None,
                        "pan_card_upload": row["pan_card_upload"] if "pan_card_upload" in row and pd.notna(
                            row["pan_card_upload"]) else None,
                        "marksheets_upload": row["marksheets_upload"] if "marksheets_upload" in row and pd.notna(
                            row["marksheets_upload"]) else None,
                    }

                    # Create the Player instance
                    Player.objects.create(**player_data)

                except IntegrityError as e:
                    logger.error(f"IntegrityError: {e} for row: {row}")
                    # Handle row-related issues here if necessary

            return redirect("player_list")
        else:
            logger.error("Form is invalid")
    else:
        form = UploadFileForm()

    return render(request, "player_list.html", {"form": form})


def download_blank_excel(request):
    # Get all fields from the Player model
    fields = Player._meta.get_fields()

    # List to store headers (field names)
    headers = []

    for field in fields:
        # Skip related model fields (e.g., ForeignKey, OneToMany)
        if field.is_relation:
            continue
        headers.append(field.name)

    # Create a new Workbook and select the active worksheet
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Players'

    # Append headers as the first row in the sheet
    sheet.append(headers)

    # Make header row bold
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    # Set the HTTP response for downloading the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=blank_players_template.xlsx'
    workbook.save(response)

    return response


# --------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------


def get_all_group_players(request):
    group_id = request.GET.get('group_id')
    group = get_object_or_404(Player_Group, pk=group_id)
    players = group.player_set.all().values('pk', 'name', 'image')
    return JsonResponse({'players': list(players)})


def get_all_players(request):
    players = Player.objects.all().values('pk', 'name', 'image')
    return JsonResponse({'players': list(players)})


# View to manage player groups (create/delete groups)
def manage_groups(request):
    groups = Player_Group.objects.all()
    players = Player.objects.all()
    group_form = GroupForm()

    if request.method == "POST":
        if 'create_group' in request.POST:
            group_form = GroupForm(request.POST)
            if group_form.is_valid():
                group = group_form.save()
                player_ids = request.POST.getlist('group_players')
                for player_id in player_ids:
                    player = Player.objects.get(pk=player_id)
                    player.players_in_groups.add(group)
                return redirect('manage_groups')  # Ensure redirect on success
        elif 'remove_player_from_group' in request.POST:
            group_id = request.POST.get('group_id')
            player_id = request.POST.get('player_id')
            group = get_object_or_404(Player_Group, pk=group_id)
            player = get_object_or_404(Player, pk=player_id)
            player.players_in_groups.remove(group)
            return redirect('manage_groups')  # Ensure redirect on success

    # Always return a response, whether GET or POST
    context = {
        'groups': groups,
        'players': players,
        'group_form': group_form
    }
    return render(request, 'player_app/player_group_manage.html', context)  # Ensure response is returned


# Adjust this to our main player_group_manage.html as a modal
# View to manage all groups (add/remove players)
def manage_all_groups(request):
    if request.method == 'POST':
        group_id = request.POST.get('group_id')
        action = request.POST.get('action')
        player_ids = request.POST.getlist('player_ids')

        group = get_object_or_404(Player_Group, pk=group_id)
        players = Player.objects.filter(pk__in=player_ids)

        if action == 'add':
            for player in players:
                player.players_in_groups.add(group)  # Add the player to the group
            messages.success(request, 'Players added to group successfully!')
        elif action == 'remove':
            for player in players:
                player.players_in_groups.remove(group)  # Remove the player from the group
            messages.success(request, 'Players removed from group successfully!')

        return redirect('manage_groups')  # Adjust this to your actual URL name for managing groups

    groups = Player_Group.objects.all()
    players = Player.objects.all()

    context = {
        'groups': groups,
        'players': players
    }

    return render(request, 'player_app/add_player_to_group.html', context)


def rename_group(request, group_id):
    group = get_object_or_404(Player_Group, id=group_id)

    if request.method == 'POST':
        new_name = request.POST.get('name')
        if new_name:
            group.name = new_name  # Update the name
            group.save()  # Save the changes
            return redirect('group_list')  # Redirect to the list of groups or another page

    return render(request, 'rename_group.html', {'group': group})


def delete_group(request, group_id):
    print(f"Delete group function called for group id: {group_id}")  # Debug statement
    group = get_object_or_404(Player_Group, pk=group_id)
    group.delete()
    return redirect('manage_groups')


def player_home(request, pk):
    player = get_object_or_404(Player, pk=pk)
    form_assignments = FormAssignment.objects.filter(player=player)
    forms = [assignment.form for assignment in form_assignments]
    return render(request, 'player_app/player_home.html', {'player': player, 'forms': forms})


from django.shortcuts import render
from django.http import HttpResponseForbidden
from .models import CampTournament, Organization


def camps_tournaments(request):
    """
    Displays a list of camps/tournaments.
    - Super Admins can view all camps.
    - Staff can see camps based on their permissions.
    - Users in an organization can see only camps from their organization.
    """

    # Super Admin: Sees all camps
    if request.user.is_superuser:
        organizations = Organization.objects.all()  # Allow filtering by organization
        selected_org = request.GET.get('organization')

        if selected_org:
            camps = CampTournament.objects.filter(organization_id=selected_org, is_deleted=False)
        else:
            camps = CampTournament.objects.filter(is_deleted=False)  # Show all camps

    # Staff with permission: Sees all camps
    elif hasattr(request.user, 'staff') and request.user.staff.view_camps_tournaments:
        camps = CampTournament.objects.filter(is_deleted=False)

    # Regular users: See only camps in their organization
    elif hasattr(request.user, 'organization') and request.user.organization:
        camps = CampTournament.objects.filter(organization=request.user.organization, is_deleted=False)

    # No Access: Deny access if the user doesn’t meet the criteria
    else:
        return HttpResponseForbidden(
            "You must belong to an organization or have the necessary permissions to view camps and tournaments.")

    return render(request, 'player_app/camps_tournaments.html', {
        'camps': camps,
        'organizations': organizations if request.user.is_superuser else None
    })


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from player_app.models import CampTournament, Player
from accounts.models import Organization  # Ensure Organization is imported


@login_required
def create_camp(request):
    """
    Create a new Camp/Tournament.
    """
    organizations = None
    players = Player.objects.none()  # Default to no players

    # Super Admin: Get all organizations
    if request.user.is_superuser:
        organizations = Organization.objects.all()
        players = Player.objects.all()  # Super Admins can see all players

    # Organization Admins: Check if user has an organization directly
    elif hasattr(request.user, "organization") and request.user.organization:
        organization = request.user.organization
        players = Player.objects.filter(organization=organization)

    # Staff Members: Ensure they have a staff profile before accessing
    elif hasattr(request.user, "staff") and request.user.staff:
        organization = request.user.staff.organization
        players = Player.objects.filter(organization=organization)

    else:
        return HttpResponseForbidden(
            "You must be a Super Admin, Organization Admin, or a Staff member to create a camp.")

    if request.method == "POST":
        name = request.POST.get("name")
        camp_type = request.POST.get("camp_type")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")
        venue = request.POST.get("venue")

        # Super Admin: Allow selecting an organization
        if request.user.is_superuser:
            organization_id = request.POST.get("organization")
            organization = get_object_or_404(Organization, id=organization_id)
        elif hasattr(request.user, "organization") and request.user.organization:
            # Organization Admins: Auto-set organization
            organization = request.user.organization
        elif hasattr(request.user, "staff") and request.user.staff:
            # Staff: Auto-set organization from staff profile
            organization = request.user.staff.organization
        else:
            return HttpResponseForbidden("You do not have permission to create a camp.")

        # Create the camp/tournament
        camp = CampTournament.objects.create(
            name=name,
            camp_type=camp_type,
            start_date=start_date,
            end_date=end_date,
            venue=venue,
            organization=organization,
            created_by=request.user
        )

        # Add participants (Only players from the same organization)
        selected_participants = request.POST.getlist("participants")
        valid_participants = players.filter(id__in=selected_participants)  # Ensure only valid participants
        camp.participants.set(valid_participants)

        messages.success(request, "Camp/Tournament created successfully!")
        return redirect("camps_tournaments")  # Redirect after creation

    return render(request, "player_app/create_camp.html", {
        "organizations": organizations,  # Super Admin can select
        "players": players  # Filtered players for the user
    })


def delete_camp(request, camp_id):
    """
    Allows authorized users to soft-delete a camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)  # Remove organization filtering
    camp.is_deleted = True
    camp.save()
    # Log the deletion activity
    CampActivity.objects.create(
        camp=camp,
        action='deleted',
        performed_by=request.user,
        details=f"Camp/Tournament '{camp.name}' was deleted."
    )
    messages.success(request, 'Camp/Tournament deleted successfully.')
    return redirect('camps_tournaments')


def camp_detail(request, camp_id):
    """
    Displays details of a specific camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)
    return render(request, 'player_app/camp_detail.html', {'camp': camp})


from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from .models import CampTournament, Player, CampActivity


def edit_camp(request, camp_id):
    """
    Handles editing a specific camp/tournament, including participants.
    """
    camp = get_object_or_404(CampTournament, id=camp_id)

    if request.method == 'POST':
        camp.name = request.POST.get('name')
        camp.camp_type = request.POST.get('camp_type')

        # Keep existing start date (only allow editing end date)
        end_date = request.POST.get('end_date')
        if end_date:
            camp.end_date = end_date

        camp.venue = request.POST.get('venue')

        # Update participants (only from the same organization)
        participant_ids = request.POST.getlist('participants')
        camp.participants.set(participant_ids)

        camp.save()

        # Log the update activity
        CampActivity.objects.create(
            camp=camp,
            action='updated',
            performed_by=request.user,
            details=f"Camp/Tournament '{camp.name}' was updated."
        )

        messages.success(request, "Camp/Tournament updated successfully!")
        return redirect('camp_detail', camp_id=camp.id)

    # Get only players from the same organization
    participants = Player.objects.filter(organization=camp.organization)

    return render(request, 'player_app/edit_camp.html', {
        'camp': camp,
        'participants': participants
    })


def trash_camps(request):
    """
    Lists all deleted camps/tournaments for management (restore or permanent delete).
    """
    deleted_camps = CampTournament.objects.filter(is_deleted=True)
    return render(request, 'player_app/trash_camps.html', {'deleted_camps': deleted_camps})


def restore_camp(request, camp_id):
    """
    Restores a soft-deleted camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id, is_deleted=True)
    camp.is_deleted = False
    camp.save()
    messages.success(request, 'Camp/Tournament restored successfully.')
    return redirect('trash_camps')


def permanently_delete_camp(request, camp_id):
    """
    Permanently deletes a camp/tournament.
    """
    camp = get_object_or_404(CampTournament, id=camp_id, is_deleted=True)
    camp.delete()
    messages.success(request, 'Camp/Tournament permanently deleted.')
    return redirect('trash_camps')


def add_participant(request, camp_id, participant_id):
    camp = get_object_or_404(CampTournament, id=camp_id)
    participant = get_object_or_404(Player, id=participant_id)

    camp.participants.add(participant)

    # Log the participant addition activity
    CampActivity.objects.create(
        camp=camp,
        action='player_added',
        performed_by=request.user,
        details=f"Player '{participant.name}' was added to the camp/tournament."
    )

    messages.success(request, f"Player '{participant.name}' added successfully.")
    return redirect('camp_detail', camp_id=camp.id)


def download_activity_history(request, camp_id):
    camp = get_object_or_404(CampTournament, id=camp_id)
    activities = camp.activities.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{camp.name}_activity_history.csv"'

    writer = csv.writer(response)
    writer.writerow(['Action', 'Performed By', 'Timestamp', 'Details'])

    for activity in activities:
        writer.writerow([
            activity.get_action_display(),
            activity.performed_by.username if activity.performed_by else "System",
            activity.timestamp,
            activity.details
        ])

    return response


def create_program(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        program_type = request.POST.get('program_type')
        template = 'template' in request.POST

        Program.objects.create(
            name=name,
            description=description,
            program_type=program_type,
            template=template,
            created_by=request.user
        )
        messages.success(request, "Program created successfully!")
        return redirect('program_list')

    return render(request, 'player_app/create_program.html')


def assign_program(request):
    """
    Assign a program to a player.
    """
    if request.method == 'POST':
        player_id = request.POST.get('player_id')
        program_id = request.POST.get('program_id')  # Use `program_id` instead of `id`
        injury_id = request.POST.get('injury_id', None)  # Optional for rehab programs

        # Validate the program_id
        if not program_id:
            messages.error(request, "Please select a valid program.")
            return redirect('assign_program')

        # Get the player and program objects
        player = get_object_or_404(Player, id=player_id)
        program = get_object_or_404(Program, program_id=program_id)  # Use `program_id`

        # Validate rehab programs require an injury ID
        if program.program_type == 'rehab' and not injury_id:
            messages.error(request, "Rehab programs must be assigned with an injury ID.")
            return redirect('assign_program')

        # Assign the program to the player
        AssignedProgram.objects.create(
            player=player,
            program=program,
            injury_id=injury_id,
            assigned_by=request.user
        )

        messages.success(request, "Program assigned successfully!")
        return redirect('program_list')

    # Fetch all players and programs for the dropdowns
    players = Player.objects.all()
    programs = Program.objects.all()
    return render(request, 'player_app/assign_program.html', {'players': players, 'programs': programs})


def save_workout_data(request, program_id):
    """
    Save workout data for a specific assigned program.
    Super Admins and Players can save data.
    """
    assigned_program = get_object_or_404(AssignedProgram, id=program_id)

    # Check if the user is a Player, Super Admin, or Staff with appropriate permission
    if not (
            request.user.is_superuser or
            hasattr(request.user, 'player') or
            (hasattr(request.user, 'staff') and request.user.staff.assign_program)
    ):
        return HttpResponse(
            "<h1 style='color: red; text-align: center;'>Permission Denied</h1>"
            "<p style='text-align: center;'>You do not have permission to save workout data.</p>",
            status=403
        )

    if request.method == 'POST':
        workout_details = request.POST.get('workout_details')

        # If the user is a Super Admin, use a placeholder or assign the workout to the first player
        player = getattr(request.user, 'player', None)
        if request.user.is_superuser and not player:
            player = assigned_program.player  # Assign to the player in the AssignedProgram

        # Save workout data
        WorkoutData.objects.create(
            assigned_program=assigned_program,
            player=player,
            workout_details=workout_details
        )

        messages.success(request, "Workout data saved successfully!")
        return redirect('program_list')  # Redirect to the program list or another appropriate page

    return render(request, 'player_app/save_workout_data.html', {'assigned_program': assigned_program})


def program_list(request):
    """
    Displays a list of training programs based on user roles.
    """
    user = request.user

    # Super Admin can filter programs by organization
    if user.is_superuser:
        organizations = Organization.objects.all()
        selected_org_id = request.GET.get('organization')
        if selected_org_id:
            programs = Program.objects.filter(created_by__organization=selected_org_id)
        else:
            programs = Program.objects.all()
    else:
        # Regular users see only their organization's programs
        programs = Program.objects.filter(created_by__organization=user.organization)
        organizations = None  # No need for dropdown

    return render(request, 'player_app/program_list.html', {
        'programs': programs,
        'organizations': organizations
    })


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from .forms import InjuryForm
from .models import Player, Organization

@login_required
def create_injury(request):
    """
    Create an injury record with organization-based player filtering.
    """
    organizations = None  # Initialize organizations variable

    # Super Admin: Can select an organization
    if request.user.is_superuser:
        organizations = Organization.objects.all()  # Fetch all organizations
        players = Player.objects.all()  # Show all players
    else:
        # Organization users can only select players from their own organization
        if hasattr(request.user, 'organization') and request.user.organization:
            players = Player.objects.filter(organization=request.user.organization)
        else:
            players = Player.objects.none()  # No players if no organization found

    if request.method == 'POST':
        form = InjuryForm(request.POST, request.FILES)

        if form.is_valid():
            injury = form.save(commit=False)

            # Auto-assign organization if the user is not a super admin
            if not request.user.is_superuser:
                injury.player.organization = request.user.organization

            injury.save()
            return redirect('injury_list')

    else:
        form = InjuryForm()

    return render(request, 'player_app/create_injury.html', {
        'form': form,
        'players': players,
        'organizations': organizations
    })



from django.shortcuts import render
from .models import Injury, Organization

def injury_list(request):
    injuries = Injury.objects.all()
    organizations = Organization.objects.all()  # Fetch all organizations
    selected_org_id = request.GET.get("organization", None)

    # Filter injuries by selected organization (if applicable)
    if selected_org_id:
        injuries = injuries.filter(player__organization_id=selected_org_id)

    return render(request, "player_app/injury_list.html", {
        "injuries": injuries,
        "organizations": organizations,
        "selected_org_id": int(selected_org_id) if selected_org_id else None
    })



from django.shortcuts import render, get_object_or_404
from .models import Player, Injury, AssignedProgram, WorkoutData, CampTournament


def player_injury_details(request, player_id):
    player = get_object_or_404(Player, id=player_id)
    injuries = Injury.objects.filter(player=player)
    assigned_programs = AssignedProgram.objects.filter(player=player)
    workout_data = WorkoutData.objects.filter(assigned_program__in=assigned_programs)

    # Fetch all camps/tournaments the player has participated in
    camps = CampTournament.objects.filter(participants=player)

    return render(request, 'player_app/player_injury_details.html', {
        'player': player,
        'injuries': injuries,
        'assigned_programs': assigned_programs,
        'workout_data': workout_data,
        'camps': camps,  # Send camp data to template
    })


def update_injury(request, pk):
    injury = get_object_or_404(Injury, pk=pk)
    if request.method == 'POST':
        form = InjuryForm(request.POST, request.FILES, instance=injury)
        if form.is_valid():
            if form.cleaned_data['status'] == 'closed' and not request.FILES.get('medical_documents'):
                return HttpResponse("Medical documentation is required to close the injury.")
            form.save()
            return redirect('injury_list')
    else:
        form = InjuryForm(instance=injury)
    return render(request, 'player_app/update_injury.html', {'form': form})


@login_required
def update_injury_status(request, pk):
    injury = get_object_or_404(Injury, pk=pk)

    # Permission check
    if not (request.user.is_superuser or
            (hasattr(request.user, 'staff') and request.user.staff.injury_tracking)):
        return HttpResponseForbidden("You don't have permission to update the injury status.")

    if request.method == 'POST':
        status = request.POST.get('status')

        # If closing, redirect to the document upload page
        if status == 'closed':
            return render(request, 'player_app/confirm_close.html', {'injury': injury})

        # Otherwise, just update the status
        injury.status = status
        injury.save()
        return redirect('injury_list')


@login_required
def confirm_close(request, pk):
    injury = get_object_or_404(Injury, pk=pk)

    # Permission check
    if not (request.user.is_superuser or
            (hasattr(request.user, 'staff') and request.user.staff.injury_tracking)):
        return HttpResponseForbidden("You don't have permission to update the injury status.")

    if request.method == 'POST':
        medical_documents = request.FILES.get('medical_documents')

        if medical_documents:
            injury.medical_documents = medical_documents
            injury.status = 'closed'
            injury.save()
            return redirect('injury_list')
        else:
            return HttpResponse("Medical documentation is required to close the injury.")
    return redirect('injury_list')


def add_treatment_recommendation(request, injury_id):
    # Check if the user is a Physio or Super Admin
    physio = Staff.objects.filter(user=request.user, role="physio").first()
    is_super_admin = request.user.is_superuser  # Check if the user is a Super Admin

    if not physio and not is_super_admin:
        messages.error(request, "You must be a Physio or Super Admin to recommend treatments.")
        return redirect("injury_list")

    injury = get_object_or_404(Injury, id=injury_id)
    player = injury.player  # Ensure the correct player is linked to the injury

    # Get existing treatment recommendations for the injury
    treatment_recommendations = TreatmentRecommendation.objects.filter(injury=injury)

    if request.method == "POST":
        form = TreatmentRecommendationForm(request.POST)
        if form.is_valid():
            recommendation = form.save(commit=False)
            recommendation.injury = injury  # Link treatment to the injury
            recommendation.player = player  # Ensure the correct player is assigned

            if is_super_admin:
                # Allow Super Admin to select Physio manually
                physio_id = request.POST.get('physio')  # Get selected physio ID from the form data
                if physio_id:
                    recommendation.physio = Staff.objects.get(id=physio_id)
                else:
                    messages.error(request, "Please select a Physio.")
                    return redirect("add_treatment", injury_id=injury.id)
            else:
                # Automatically assign the logged-in Physio
                recommendation.physio = physio

            recommendation.save()

            messages.success(request, "Treatment recommendation added successfully!")
            return redirect("player_injury_details", player_id=player.id)

    else:
        form = TreatmentRecommendationForm(initial={'player': player})  # Pre-fill the correct player

    return render(request, "player_app/add_treatment.html", {
        "form": form,
        "injury": injury,
        "treatment_recommendations": treatment_recommendations,
        "physios": Staff.objects.filter(role="physio"),  # Pass available Physios for selection
    })

# --------------------------------------------------------------------------------------------------------------------------------
# Team Management Views
# --------------------------------------------------------------------------------------------------------------------------------
from datetime import datetime, date, time
def teams_dashboard(request):
    category_keys = [
        ('boys_under-15', 'Boys under 15'),
        ('boys_under-19', 'Boys under 19'),
        ('men_under-23', 'Men Under 23'),
        ('men_senior', 'Men Senior'),
    ]

    categories = []
    for key, display in category_keys:
        teams = Team.objects.filter(category=key, is_active=True).prefetch_related('players', 'staff')
        categories.append((key, display, teams))

    today = date.today()

    end = CampTournament.objects.filter(
        organization=request.user.organization,
        is_deleted=False,
    )
   

    # Ongoing: started on or before today, and end_date after today or null (no end date means ongoing)
    ongoing_camps = CampTournament.objects.filter(
        organization=request.user.organization,
        is_deleted=False,
        start_date__lte=today,
    ).filter(
        Q(end_date__gt=today) | Q(end_date__isnull=True)
    ).order_by('name')

    # Closed: started on or before today and ended on or before today
    closed_camps = CampTournament.objects.filter(
        organization=request.user.organization,
        is_deleted=False,
        start_date__lte=today,
        end_date__lte=today,
    ).order_by('name')


    camps = {
        'ongoing': ongoing_camps,
        'closed': closed_camps,
    }

    context = {
        'categories': categories,
        'camps': camps,
    }
    return render(request, 'player_app/organization/teams_dashboard.html', context)
    
@login_required
def player_record(request):
    user_organization = getattr(request.user, 'organization', None)
    players_in_org = Player.objects.filter(organization=user_organization)
    session = None
    if 'report_settings' in request.session:
        session = True
    return render(request, 'player_app/organization/player_record.html', {
        'players': players_in_org,'session': session,
    })


@login_required
def player_data(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        player_id = request.POST.get('playerSelect')
        test = request.POST.get('testSelect')
        date_option = request.POST.get('date_option')
        num_tests = request.POST.get('numTestsToShow')  # Get the selected number of tests

        # Validate date range
        if date_option == 'range':
            if not start_date or not end_date:
                players_in_org = Player.objects.filter(organization=getattr(request.user, 'organization', None))
                error_message = "Please select both start and end dates for the date range."
                return render(request, 'player_app/organization/player_record.html', {
                    'players': players_in_org,
                    'error_message': error_message,
                })

        # Defaults for start and end dates
        if not start_date:
            start_date = '2000-01-01'
        if not end_date:
            from datetime import date
            end_date = date.today().strftime('%Y-%m-%d')

        # Pass num_tests parameter within redirect URL query string
        return redirect(f"{reverse('player_info', kwargs={'player_id': player_id, 'test': test, 'start': start_date, 'end': end_date})}?num_tests={num_tests}")
    else:
        players_in_org = Player.objects.filter(organization=getattr(request.user, 'organization', None))
        return render(request, 'player_app/organization/player_record.html', {
            'players': players_in_org,
        })
    
@login_required
def player_info(request, player_id, test, start, end):
    player = get_object_or_404(Player, id=player_id)
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)

    num_tests = int(request.GET.get('num_tests', 5))

    # Filter tests by player, test, date range
    if start and end:
        tests_qs = TestAndResult.objects.filter(player_id=player_id, test=test, date__range=[start, end])
    else:
        tests_qs = TestAndResult.objects.filter(player_id=player_id, test=test)

    tests = tests_qs.order_by('-date')[:num_tests]

    player_cat = player.age_category
    category = Category.objects.filter(name=player_cat).first() if player_cat else None

    target_value = None
    if category:
        target_obj = CategoryTarget.objects.filter(category=category).order_by('-settings__created_at').first()
        if target_obj:
            target_value = target_obj.target_value

    # Read min/max formula and toggles from session with defaults
    session_settings = request.session.get('report_settings', {})
    min_max_formula = session_settings.get('min_max_formula', 'all_players')
    min_is_better = session_settings.get('min_is_better', False)
    grp_avg_option = session_settings.get('grp_avg_option', None)
    
    # Calculate min and max based on formula
    if min_max_formula == "all_players":
        min_value = TestAndResult.objects.filter(test=test).aggregate(Min('best'))['best__min']
        max_value = TestAndResult.objects.filter(test=test).aggregate(Max('best'))['best__max']
    
    elif min_max_formula == "all_players_by_gender":
        min_value = TestAndResult.objects.filter(test=test, player__gender=player.gender).aggregate(Min('best'))['best__min']
        max_value = TestAndResult.objects.filter(test=test, player__gender=player.gender).aggregate(Max('best'))['best__max']
        
    elif min_max_formula == "category_based":
        min_value = TestAndResult.objects.filter(test=test, player__age_category=player.age_category).aggregate(Min('best'))['best__min']
        max_value = TestAndResult.objects.filter(test=test, player__age_category=player.age_category).aggregate(Max('best'))['best__max']
          
    elif min_max_formula == "date_based":
        group_filter = TestAndResult.objects.filter(test=test)
        if start and end:
            group_filter = group_filter.filter(date__range=[start, end])
        min_value = group_filter.aggregate(Min('best'))['best__min']
        max_value = group_filter.aggregate(Max('best'))['best__max']
    elif min_max_formula == "manual_entry":
        min_value = session_settings.get('manual_min', None)
        max_value = session_settings.get('manual_max', None)
    else:
        min_value = tests_qs.aggregate(Min('best'))['best__min']
        max_value = tests_qs.aggregate(Max('best'))['best__max']

    # Swap min and max if "min is better"
    if min_is_better and (min_value is not None and max_value is not None):
        min_value, max_value = max_value, min_value

    individual_avg = tests.aggregate(avg_best=Avg('best'))['avg_best']

    # Group average calculation based on grp_avg_option
    group_tests = TestAndResult.objects.filter(test=test)

    if grp_avg_option == "all_players_date":
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end])
    elif grp_avg_option == "all_players_gender_date":
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end], player__gender=player.gender)
        else:
            group_tests = group_tests.filter(player__gender=player.gender)
    elif grp_avg_option == "category_based_date":
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end], player__age_category=player.age_category)
        else:
            group_tests = group_tests.filter(player__age_category=player.age_category)
    
    elif grp_avg_option == "camp_or_tournament":
        if tests.exists():
            phase = tests.first().phase
            group_tests = TestAndResult.objects.filter(test=test, phase=phase)
            group_avg = group_tests.aggregate(avg_best=Avg('best'))['avg_best']
        else:
            group_avg = None

    else:
        if start and end:
            group_tests = group_tests.filter(date__range=[start, end])

    group_avg = group_tests.aggregate(avg_best=Avg('best'))['avg_best']

    individual_normalized = (
        (individual_avg - min_value) / (max_value - min_value) * 100
        if max_value is not None and min_value is not None and max_value != min_value else 0
    )

    session = None
    if 'report_settings' in request.session:
        session = True

    abc = "start"

    return render(request, 'player_app/organization/player_record.html', {
        'player': player,
        'players': players,
        'tests': tests,
        'individual_avg': individual_avg,
        'group_avg': group_avg,
        'abc': abc,
        'max_value': max_value,
        'min_value': min_value,
        'test': test,
        'individual_normalized': individual_normalized,
        'target_value': target_value,
        'session': session,
    })
 


def delete_session(request):
    if 'report_settings' in request.session:
        del request.session['report_settings']
    return redirect('player_report')

@login_required
def get_players_by_test(request):
    test = request.GET.get('test')
    user_organization = getattr(request.user, 'organization', None)
    if not user_organization or not test:
        return JsonResponse({'players': []})

    # Get player IDs who have done the selected test and belong to the organization
    player_ids = TestAndResult.objects.filter(
        test=test, player__organization=user_organization
    ).values_list('player_id', flat=True).distinct()

    players = Player.objects.filter(id__in=player_ids).values('id', 'name')
    players_list = list(players)

    return JsonResponse({'players': players_list})

from django.utils.dateparse import parse_date
@login_required
def get_player_test_results(request):
    player_id = request.GET.get('player_id')
    test = request.GET.get('test')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    limit = int(request.GET.get('limit', 5))

    start_date = parse_date(start_date_str) if start_date_str else None
    end_date = parse_date(end_date_str) if end_date_str else None

    filters = {
        'player_id': player_id,
        'test': test
    }

    if start_date and end_date:
        filters['date__range'] = [start_date, end_date]

    qs = TestAndResult.objects.filter(**filters).order_by('-date')[:limit]

    results = list(qs.values('date', 'best', 'phase'))
    return JsonResponse({'results': results, 'count': qs.count()})

def new_test_details(request):
    results = TestAndResult.objects.select_related('player').all()
    return render(request, 'player_app/organization/new_test_details.html', {'results': results})



from django.forms.models import model_to_dict
def nomative_data(request):
    data = NomativeData.objects.all()
    return render(request, 'player_app/organization/nomative_data.html', {'datas': data})
# @login_required
# def new_test_dashboard(request):
#     user_organization = getattr(request.user, 'organization', None)
#     if not user_organization:
#         return render(request, 'player_app/organization/test_results.html', {
#             'error_message': "Your account is not linked to any organization.",
#         })
#     players_in_org = Player.objects.filter(organization=user_organization)

#     # Base queryset ordered by player, test and date
#     qs = TestAndResult.objects.select_related('player').filter(player__in=players_in_org).order_by(
#         'player__name', 'test', 'date'
#     )

#     # Calculate group averages per test using "best" field (one average per test)
#     group_averages = TestAndResult.objects.filter(player__in=players_in_org) \
#         .values('test').annotate(group_avg=Avg('best')).order_by('test')
#     group_avg_dict = {item['test']: item['group_avg'] for item in group_averages}

#     # Calculate individual averages per player per test using "best" field
#     indv_averages = TestAndResult.objects.filter(player__in=players_in_org).values('player', 'test').annotate(indv_avg=Avg('best'))
#     indv_avg_dict = {(item['player'], item['test']): item['indv_avg'] for item in indv_averages}

#     # Annotate each trial with averages for easy display in template
#     trials_with_avg = []
#     for trial in qs:
#         trial.individual_average = indv_avg_dict.get((trial.player.id, trial.test), None)
#         trial.group_average = group_avg_dict.get(trial.test, None)
#         trials_with_avg.append(trial)

#     context = {
#         'trials': trials_with_avg,
#     }

#     return render(request, 'player_app/organization/new_test_dashboard.html', context)

def report_settings_view(request):
    categories = Category.objects.all()

    # Get latest targets from DB for this user (latest ReportSettings)
    latest_settings = ReportSettings.objects.filter(user=request.user).order_by('-created_at').first()
    targets_dict = {}
    if latest_settings:
        targets_dict = {ct.category_id: ct.target_value for ct in latest_settings.category_targets.all()}

    # Get non-target settings from session
    session_settings = request.session.get('report_settings', {})

    # Build dummy settings for template with session data
    settings_data = type('Settings', (), {})()
    settings_data.min_max_formula = session_settings.get('min_max_formula', 'all_players')
    settings_data.min_is_better = session_settings.get('min_is_better', False)
    settings_data.indv_avg_option = session_settings.get('indv_avg_option', 'total_result')
    settings_data.grp_avg_option = session_settings.get('grp_avg_option', 'all_players_date')

    category_targets = []
    for category in categories:
        target_value = targets_dict.get(category.id, '')
        category_targets.append((category, target_value))

    context = {
        'category_targets': category_targets,
        'settings': settings_data,
    }
    return render(request, 'player_app/organization/report_settings.html', context)



def save_report_settings_view(request):
    if request.method != 'POST':
        return redirect(reverse('report_settings'))

    categories = Category.objects.all()

    # Capture form data from POST
    min_max_formula = request.POST.get('min_max_formula', 'all_players')
    min_is_better = 'min_is_better' in request.POST
    indv_avg_option = request.POST.get('indv_avg_option', 'total_result')
    grp_avg_option = request.POST.get('grp_avg_option', 'all_players_date')

    targets = {}
    for category in categories:
        val = request.POST.get(f'target_{category.id}', '').strip()
        if val:
            try:
                targets[str(category.id)] = float(val)
            except ValueError:
                # Skip invalid float entries or handle as needed
                pass

    # Save all to session
    request.session['report_settings'] = {
        'min_max_formula': min_max_formula,
        'min_is_better': min_is_better,
        'indv_avg_option': indv_avg_option,
        'grp_avg_option': grp_avg_option,
        'targets': targets,
    }

    # Mark session as modified to ensure Django saves changes
    request.session.modified = True

    messages.success(request, 'Report settings saved to session successfully!')
    return redirect(reverse('player_report'))

@login_required
def test_results_main(request):
    user = request.user
    # User's organization (adjust based on your user model)
    user_organization = getattr(user, 'organization', None)
    if not user_organization:
        return render(request, 'player_app/organization/test_dashboard.html', {
            'error_message': "Your account is not linked to any organization.",
        })

    # Players in user's organization
    players_in_org = Player.objects.filter(organization=user_organization)

    # Forms initialization with restricted player queryset
    if request.method == 'POST':
        add_form = TestAndResultForm(request.POST)
        add_form.fields['player'].queryset = players_in_org
        if add_form.is_valid():
            add_form.save()
            return redirect('test_dashboard')
    else:
        add_form = TestAndResultForm()
        add_form.fields['player'].queryset = players_in_org

    filter_form = TestSummaryFilterForm(request.GET or None)
    filter_form.fields['player'].queryset = players_in_org

    # Base queryset with related player, user; filtered by org players
    qs = (
        TestAndResult.objects
        .select_related('player', 'reported_by')
        .filter(player__in=players_in_org)
        .order_by('player__name', 'test', 'date', 'id')
    )

    # Apply filtering
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('player'):
            qs = qs.filter(player=filter_form.cleaned_data['player'])
        if filter_form.cleaned_data.get('test'):
            qs = qs.filter(test=filter_form.cleaned_data['test'])

    # Group test results by (player_id, test)
    grouped_results = defaultdict(list)
    for tr in qs:
        key = (tr.player.id, tr.test)
        grouped_results[key].append(tr)

    summary_rows = []
    for (player_id, test), trials in grouped_results.items():
        if not trials:
            continue

        # Sort trials by date and id to keep chronological order
        trials = sorted(trials, key=lambda x: (x.date or x.created_at, x.id))
        last_two_trials = trials[-2:] if len(trials) >= 2 else trials[-1:]

        # Trials data for columns
        trial_1_best = last_two_trials[0].best if len(last_two_trials) == 2 else None
        trial_2_best = last_two_trials[-1].best
        best_trial = min(t.best for t in trials if t.best is not None)

        last_trial_obj = last_two_trials[-1]

        # Individual average stored in PlayerAggregate or computed from TestAndResult
        player_agg = PlayerAggregate.objects.filter(player__id=player_id, test=test).first()
        indv_average = player_agg.individual_average if player_agg else None

        # Group average from GenderAggregate or CategoryAggregate fallback
        gender = getattr(last_trial_obj.player, 'gender', None)
        category = getattr(last_trial_obj.player, 'category', None)

        group_average = None
        if gender:
            gender_agg = GenderAggregate.objects.filter(gender=gender, test=test).first()
            group_average = gender_agg.average if gender_agg else None
        if group_average is None and category:
            cat_agg = CategoryAggregate.objects.filter(category=category, test=test).first()
            group_average = cat_agg.average if cat_agg else None

        summary_rows.append({
            'serial': None,  # Will assign later per test
            'player_name': last_trial_obj.player.name,
            'test': test,
            'last_date': last_trial_obj.date,
            'last_phase': last_trial_obj.phase,
            'trial_1': trial_1_best,
            'trial_2': trial_2_best,
            'best_trial': best_trial,
            'indv_average': indv_average,
            'group_average': group_average,
        })

    # Group rows by test for separate tables
    summary_by_test = defaultdict(list)
    for row in summary_rows:
        summary_by_test[row['test']].append(row)

    # Assign serial numbers per test group
    for test_name, rows in summary_by_test.items():
        for idx, row in enumerate(rows, start=1):
            row['serial'] = idx

    context = {
        'add_form': add_form,
        'form': filter_form,
        'summary_by_test': dict(summary_by_test),
    }

    return render(request, 'player_app/organization/test_result_main.html', context)


# New test Views
def test_dashboard_new(request):
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']
    return render(request, 'player_app/organization/organization_main_test_dash.html')

from django.core.paginator import Paginator
from django.db.models import Q

def test_results_view(request, test_name):
    search_query = request.GET.get('search', '').strip()

    results = TestAndResult.objects.filter(test=test_name).select_related('player', 'reported_by').order_by('-date')

    # Filter by player name if search query is present
    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    from django.db.models import Max

    latest_dates = TestAndResult.objects.filter(
        player_id__in=player_ids,
        indv_average__isnull=False
    ).values('player_id').annotate(latest_date=Max('date'))

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = TestAndResult.objects.filter(
            player_id=entry['player_id'],
            date=entry['latest_date'],
            indv_average__isnull=False
        ).first()
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.indv_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,  # pass to template for form value retention
    }
    return render(request, 'player_app/organization/organization_test_data.html', context)


from django.shortcuts import render, redirect
from .models import TestAndResult, Player, User  # Make sure these models are imported

def add_test_results(request, test_name=None):
    user_organization = getattr(request.user, 'organization', None)
    
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    
     # Handle form submission
    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_id = request.POST.get('reported_by')

        # Basic field validation (add your own as needed)
        errors = []
        if not player_id: errors.append("Player is required.")
        if not test: errors.append("Test is required.")
        if not date: errors.append("Date is required.")
        if not phase: errors.append("Phase is required.")
        if not best: errors.append("Best is required.")
        if not reported_by_id: errors.append("Reported by is required.")


        # If no errors, save the result
        if not errors:
            phase_data = CampTournament.objects.get(id=int(phase))
            nomative_data = NomativeData.objects.get(final_level=float(best))
            total_distance = nomative_data.total_distance
            approximately_vo2max = nomative_data.approximately_vo2max

            player = Player.objects.get(pk=player_id)
            reported_by = User.objects.get(pk=reported_by_id)
            TestAndResult.objects.create(
                player=player,
                test=test,
                date=date,
                phase=phase_data,
                best=float(best),
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by
            )
            return redirect('test_results_by_name', test_name=test)
        # Pass errors back to template if any
        else:
            return render(request, 'player_app/organization/organization_test_add.html', {
                'test_name': test_name,
                'errors': errors,
                'players': players,
                'events': events,
                'staff':staff, 
                # you may need players/staff for dropdowns
            })
    else:
        return render(request, 'player_app/organization/organization_test_add.html', {
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# Run A 3x6 test views
def add_run_3x6_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = 'Run A 3x6'
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase = request.POST.get('phase')
        run_a_3x6_attempt1 = request.POST.get('trial1')
        run_a_3x6_attempt2 = request.POST.get('trial2')
        run_a_3x6_attempt3 = request.POST.get('trial3')
        run_a_3x6_attempt4 = request.POST.get('trial4')
        run_a_3x6_attempt5 = request.POST.get('trial5')
        run_a_3x6_attempt6 = request.POST.get('trial6')
        run_a_3x6_average = request.POST.get('run_a_average')

        notes = request.POST.get('notes')
        reported_by_id = request.POST.get('reported_by')

        # Basic field validation (add your own as needed)
        errors = []
        if not player_id: errors.append("Player is required.")
        if not test: errors.append("Test is required.")
        if not date: errors.append("Date is required.")
        if not phase: errors.append("Phase is required.")
        if not run_a_3x6_attempt1: errors.append("Trial 1 is required.")
        if not run_a_3x6_attempt2: errors.append("Trial 2 is required.")
        if not run_a_3x6_attempt3: errors.append("Trial 3 is required.")
        if not run_a_3x6_attempt4: errors.append("Trial 4 is required.")
        if not run_a_3x6_attempt5: errors.append("Trial 5 is required.")
        if not run_a_3x6_attempt6: errors.append("Trial 6 is required.")
        if not run_a_3x6_average: errors.append("Average is required.")
        if not reported_by_id: errors.append("Reported by is required.")


        # If no errors, save the result
        if not errors:
            phase_data = CampTournament.objects.get(id=int(phase))
            
             # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_id)
            reported_by = staff_obj.user  # adjust if field name differs

            player = Player.objects.get(pk=player_id)

            RunA3x6Test.objects.create(
                player=player,
                date=date,
                phase=phase_data,
                trial1=int(run_a_3x6_attempt1),
                trial2=int(run_a_3x6_attempt2),
                trial3=int(run_a_3x6_attempt3),
                trial4=int(run_a_3x6_attempt4),
                trial5=int(run_a_3x6_attempt5),
                trial6=int(run_a_3x6_attempt6),
                average=float(run_a_3x6_average),
                notes=notes,
                reported_by=reported_by,
            
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('test_dashboard_new')
        # Pass errors back to template if any
        else:
            return render(request, 'player_app/tests/runa3x6.html', {
                'test_name': test_name,
                'errors': errors,
                'players': players,
                'events': events,
                'staff':staff, 
                # you may need players/staff for dropdowns
            })


    return render(request, 'player_app/tests/runa3x6.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# Run A 3x6 test views
def run_3x6_test_view(request):
    test_name = 'Run A 3x6'
    search_query = request.GET.get('search', '').strip()

    # Use the dedicated model instead of TestAndResult
    results = (
        RunA3x6Test.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/runa3x6_data.html', context)

# Glute Bridges Test views
def add_glute_bridges_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "S/L Glute Bridges"
    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase = request.POST.get('phase')
        sl_right = request.POST.get('sl_glute_right')
        sl_left = request.POST.get('sl_glute_left')
        notes = request.POST.get('notes')
        reported_by_id = request.POST.get('reported_by')

        # Basic field validation (add your own as needed)
        errors = []
        if not player_id: errors.append("Player is required.")
        if not test: errors.append("Test is required.")
        if not date: errors.append("Date is required.")
        if not phase: errors.append("Phase is required.")
        if not sl_right: errors.append("S/L Glute Right is required.")
        if not sl_left: errors.append("S/L Glute Left is required.")
    
        if not reported_by_id: errors.append("Reported by is required.")


        # If no errors, save the result
        if not errors:
            phase_data = CampTournament.objects.get(id=int(phase))
            
            player = Player.objects.get(pk=player_id)
            staff_obj = get_object_or_404(Staff, pk=reported_by_id)
            reported_by = staff_obj.user 
            SLGluteBridges.objects.create(
                player=player,
                date=date,
                phase=phase_data,
                right=int(sl_right),
                left=int(sl_left),
                notes=notes,
                reported_by=reported_by
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
 
            return redirect('test_dashboard_new')
        # Pass errors back to template if any
        else:
            return render(request, 'player_app/tests/glute_bridges.html', {
                'test_name': test_name,
                'errors': errors,
                'players': players,
                'events': events,
                'staff':staff, 
                # you may need players/staff for dropdowns
            })
    return render(request, 'player_app/tests/glute_bridges.html',{
            'test_name': test_name,
            'players': players,
            'events': events,
            'staff':staff,
            # you may need players/staff for dropdowns
        })

# Glute Bridges Test Data views
def glute_bridges_test_view(request):
    test_name = "S/L Glute Bridges"

    search_query = request.GET.get('search', '').strip()

    # Base queryset
    results = (
        SLGluteBridges.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Players actually present in the (possibly filtered) result set
    player_ids = results.values_list('player_id', flat=True).distinct()

    # Get latest date per player where averages are present
    latest_dates = (
        SLGluteBridges.objects
        .filter(
            player_id__in=player_ids,
            individual_average_left__isnull=False,
            individual_average_right__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    # Map: player_id -> (avg_left, avg_right)
    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            SLGluteBridges.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average_left__isnull=False,
                individual_average_right__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = {
                'left': latest_result.individual_average_left,
                'right': latest_result.individual_average_right,
            }

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/glute_bridges_data.html', context)


# Lunge Calf Raises Test views
def add_lunge_calf_raises_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "S/L Lunge Calf Raises"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')          # only for display/consistency
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        sl_cr_right = request.POST.get('sl_lunge_calf_right')
        sl_cr_left = request.POST.get('sl_lunge_calf_left')
        # diff/ratio are computed in model.save(), no need to trust POST
        notes = request.POST.get('notes')
        reported_by_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not sl_cr_right:
            errors.append("S/L Lunge Calf Right is required.")
        if not sl_cr_left:
            errors.append("S/L Lunge Calf Left is required.")
        if not reported_by_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            right_val = float(sl_cr_right) if sl_cr_right not in (None, "") else None
            left_val = float(sl_cr_left) if sl_cr_left not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Right and Left values must be valid numbers.")
            right_val = left_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)
            staff_obj = get_object_or_404(Staff, pk=reported_by_id)
            reported_by_user = staff_obj.user  #

            # Model will compute sl_cr_difference, sl_cr_ratio, min/max, averages
            SLLungeCalfRaises.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                right=right_val,
                left=left_val,
                notes=notes,
                reported_by=reported_by_user,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('test_dashboard_new')

        # there are errors
        return render(request, 'player_app/tests/lunge_calf_raises.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    # GET
    return render(request, 'player_app/tests/lunge_calf_raises.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


# List / data view
def lunge_calf_raises_test_view(request):
    test_name = 'S/L Lunge Calf Raises'

    search_query = request.GET.get('search', '').strip()

    # Use the dedicated model instead of TestAndResult
    results = (
        SLLungeCalfRaises.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/lunge_calf_raises_data.html', context)



User = get_user_model()

def add_mb_rotational_throw_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "MB Rotational Throws"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')          # for display, not stored in model
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        mb_right = request.POST.get('mb_abs_right')
        mb_left = request.POST.get('mb_abs_left')
        # mb_difference / mb_ratio are computed in model.save()
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not mb_right:
            errors.append("MB Rotational Throw Right is required.")
        if not mb_left:
            errors.append("MB Rotational Throw Left is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            right_val = float(mb_right) if mb_right not in (None, "") else None
            left_val = float(mb_left) if mb_left not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Right and Left values must be valid numbers.")
            right_val = left_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            # MBRotationalThrows model computes mb_difference, mb_ratio, min/max, averages
            MBRotationalThrows.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                right=right_val,
                left=left_val,
                notes=notes,
                reported_by=reported_by_user,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('test_dashboard_new')

        # Errors: re-render form
        return render(request, 'player_app/tests/rotational_throws.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    # GET
    return render(request, 'player_app/tests/rotational_throws.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


# MB Rotational Throw Test Data views
def mb_rotational_throw_test_view(request):
    test_name = "MB Rotational Throws"
    search_query = request.GET.get('search', '').strip()

    # Use MBRotationalThrows model, not TestAndResult
    results = (
        MBRotationalThrows.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/rotational_throws_data.html', context)



# Copen Hagen Test views
def add_copen_hagen_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "Copenhagen"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        cp_hagen_right = request.POST.get('copen_hagen_right')
        cp_hagen_left = request.POST.get('copen_hagen_left')
        # difference / ratio are computed in model.save()
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not cp_hagen_right:
            errors.append("Copenhagen Right is required.")
        if not cp_hagen_left:
            errors.append("Copenhagen Left is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            right_val = float(cp_hagen_right) if cp_hagen_right not in (None, "") else None
            left_val = float(cp_hagen_left) if cp_hagen_left not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Right and Left values must be valid numbers.")
            right_val = left_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust field name if needed

            # CopenhagenTest model computes difference, ratio, min/max, averages
            CopenhagenTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                right=right_val,
                left=left_val,
                notes=notes,
                reported_by=reported_by_user,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('test_dashboard_new')

        return render(request, 'player_app/tests/copen_hagen.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    # GET
    return render(request, 'player_app/tests/copen_hagen.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })

def copen_hagen_test_view(request):
    test_name = "Copenhagen"
    search_query = request.GET.get('search', '').strip()

    results = (
        CopenhagenTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/copen_hagen_data.html', context)


# S/L Hop Test views

def add_sl_hop_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "S/L Hop Test"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')           # label only
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        sl_hop_right = request.POST.get('sl_hop_right')
        sl_hop_left = request.POST.get('sl_hop_left')
        # difference / ratio are computed in model.save()
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not sl_hop_right:
            errors.append("S/L Hop Right is required.")
        if not sl_hop_left:
            errors.append("S/L Hop Left is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            right_val = float(sl_hop_right) if sl_hop_right not in (None, "") else None
            left_val = float(sl_hop_left) if sl_hop_left not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Right and Left values must be valid numbers.")
            right_val = left_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust field name if different

            # SLHopTest model computes difference, ratio, min/max, averages
            SLHopTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                right=right_val,
                left=left_val,
                notes=notes,
                reported_by=reported_by_user,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('test_dashboard_new')

        # Errors: re-render form
        return render(request, 'player_app/tests/sl_hop.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    # GET
    return render(request, 'player_app/tests/sl_hop.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


# S/L Hop Test Data views
def sl_hop_test_view(request):
    test_name = "S/L Hop Test"
    search_query = request.GET.get('search', '').strip()

    # Use SLHopTest model instead of TestAndResult
    results = (
        SLHopTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/sl_hop_data.html', context)


# CMJ Test views
def add_cmj_scores_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "CMJ Scores"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        cmj_body_weight = request.POST.get('body_weight')
        cmj_push_off_distance = request.POST.get('push_off_distance')
        cmj_box_height = request.POST.get('box_height')
        cmj_load = request.POST.get('load')
        cmj_jump_height = request.POST.get('jump_height')
        cmj_flight_time = request.POST.get('flight_time')
        cmj_contact_time = request.POST.get('contact_time')
        cmj_force = request.POST.get('force')
        cmj_velocity = request.POST.get('velocity')
        cmj_power = request.POST.get('power')
        cmj_reactive_strength_index = request.POST.get('rsi')
        cmj_stiffness = request.POST.get('stiffness')
        cmj_readliness_color = request.POST.get('readiness_colour')
        cmj_jump_type = request.POST.get('jump_type')

        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not cmj_body_weight:
            errors.append("CMJ Body Weight is required.")
        if not cmj_push_off_distance:
            errors.append("CMJ Push Off Distance is required.")
        if not cmj_box_height:
            errors.append("CMJ Box Height is required.")
        if not cmj_load:
            errors.append("CMJ Load is required.")
        if not cmj_jump_height:
            errors.append("CMJ Jump Height is required.")
        if not cmj_flight_time:
            errors.append("CMJ Flight Time is required.")
        if not cmj_contact_time:
            errors.append("CMJ Contact Time is required.")
        if not cmj_force:
            errors.append("CMJ Force is required.")
        if not cmj_velocity:
            errors.append("CMJ Velocity is required.")
        if not cmj_power:
            errors.append("CMJ Power is required.")
        if not cmj_reactive_strength_index:
            errors.append("CMJ Reactive Strength Index is required.")
        if not cmj_stiffness:
            errors.append("CMJ Stiffness is required.")
        if not cmj_readliness_color:
            errors.append("CMJ Readiness Color is required.")
        if not cmj_jump_type:
            errors.append("CMJ Jump Type is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # Numeric casting for float fields
        def to_float(value, field_name):
            nonlocal errors
            if value in (None, ""):
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                errors.append(f"{field_name} must be a valid number.")
                return None

        bw_val = to_float(cmj_body_weight, "CMJ Body Weight")
        pod_val = to_float(cmj_push_off_distance, "CMJ Push Off Distance")
        box_val = to_float(cmj_box_height, "CMJ Box Height")
        load_val = to_float(cmj_load, "CMJ Load")
        jh_val = to_float(cmj_jump_height, "CMJ Jump Height")
        ft_val = to_float(cmj_flight_time, "CMJ Flight Time")
        ct_val = to_float(cmj_contact_time, "CMJ Contact Time")
        force_val = to_float(cmj_force, "CMJ Force")
        vel_val = to_float(cmj_velocity, "CMJ Velocity")
        power_val = to_float(cmj_power, "CMJ Power")
        rsi_val = to_float(cmj_reactive_strength_index, "CMJ Reactive Strength Index")
        stiff_val = to_float(cmj_stiffness, "CMJ Stiffness")

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            # Save to CMJTest model (replace name if different)
            CMJTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                cmj_body_weight=bw_val,
                cmj_push_off_distance=pod_val,
                cmj_box_height=box_val,
                cmj_load=load_val,
                cmj_jump_height=jh_val,
                cmj_flight_time=ft_val,
                cmj_contact_time=ct_val,
                cmj_force=force_val,
                cmj_velocity=vel_val,
                cmj_power=power_val,
                cmj_reactive_strength_index=rsi_val,
                cmj_stiffness=stiff_val,
                cmj_readiness_color=cmj_readliness_color,
                cmj_jump_type=cmj_jump_type,
                notes=notes,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('test_dashboard_new')

        # errors → re-render form
        return render(request, 'player_app/tests/cmj_scores.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    # GET
    return render(request, 'player_app/tests/cmj_scores.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


# CMJ Test Data views
def cmj_scores_test_view(request):
    test_name = "CMJ Scores"
    search_query = request.GET.get('search', '').strip()

    # Use CMJTest model instead of TestAndResult
    results = (
        CMJTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/cmj_scores_data.html', context)


# Anthropometry Test views
def add_anthropometry_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "Anthropometry Test"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        anthropometry_height = request.POST.get('height')
        anthropometry_weight = request.POST.get('weight')
        anthropometry_age = request.POST.get('age')
        anthropometry_chest = request.POST.get('chest')
        anthropometry_mid_axillary = request.POST.get('mid_axillary')
        anthropometry_subscapular = request.POST.get('subscapular')
        anthropometry_triceps = request.POST.get('triceps')
        anthropometry_abdomen = request.POST.get('abdomen')
        anthropometry_suprailiac = request.POST.get('suprailiac')
        anthropometry_mid_thigh = request.POST.get('mid_thigh')
        anthropometry_total_skinfold = request.POST.get('total_skinfold')
        anthropometry_body_density = request.POST.get('body_density')
        anthropometry_fat_percentage = request.POST.get('fat_percent')
        anthropometry_error_corrected = request.POST.get('error_corrected')

        anthropometry_chest_n = request.POST.get('chest_n')
        anthropometry_chest_e = request.POST.get('chest_e')
        anthropometry_upper_arm = request.POST.get('upper_arm')
        anthropometry_waist = request.POST.get('waist')
        anthropometry_abdomen_cm = request.POST.get('abdomen_n')
        anthropometry_hip = request.POST.get('hip')
        anthropometry_thigh = request.POST.get('thigh')
        anthropometry_calf = request.POST.get('calf')

        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")

        # required checks (you can relax some if optional)
        if not anthropometry_height:
            errors.append("Anthropometry Height is required.")
        if not anthropometry_weight:
            errors.append("Anthropometry Weight is required.")
        if not anthropometry_age:
            errors.append("Anthropometry Age is required.")
        if not anthropometry_chest:
            errors.append("Anthropometry Chest is required.")
        if not anthropometry_mid_axillary:
            errors.append("Anthropometry Mid Axillary is required.")
        if not anthropometry_subscapular:
            errors.append("Anthropometry Subscapular is required.")
        if not anthropometry_triceps:
            errors.append("Anthropometry Triceps is required.")
        if not anthropometry_abdomen:
            errors.append("Anthropometry Abdomen is required.")
        if not anthropometry_suprailiac:
            errors.append("Anthropometry Suprailiac is required.")
        if not anthropometry_mid_thigh:
            errors.append("Anthropometry Mid Thigh is required.")
        if not anthropometry_total_skinfold:
            errors.append("Anthropometry Total Skinfold is required.")
        if not anthropometry_body_density:
            errors.append("Anthropometry Body Density is required.")
        if not anthropometry_fat_percentage:
            errors.append("Anthropometry Fat Percentage is required.")
        if not anthropometry_error_corrected:
            errors.append("Anthropometry Error Corrected is required.")
        if not anthropometry_chest_n:
            errors.append("Anthropometry Chest N is required.")
        if not anthropometry_chest_e:
            errors.append("Anthropometry Chest E is required.")
        if not anthropometry_upper_arm:
            errors.append("Anthropometry Upper Arm is required.")
        if not anthropometry_waist:
            errors.append("Anthropometry Waist is required.")
        if not anthropometry_abdomen_cm:
            errors.append("Anthropometry Abdomen N is required.")
        if not anthropometry_hip:
            errors.append("Anthropometry Hip is required.")
        if not anthropometry_thigh:
            errors.append("Anthropometry Thigh is required.")
        if not anthropometry_calf:
            errors.append("Anthropometry Calf is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # Helper to cast numeric fields
        def to_float(value, field_name):
            nonlocal errors
            if value in (None, ""):
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                errors.append(f"{field_name} must be a valid number.")
                return None

        # Example: cast main numeric metrics (adjust types to your model)
        height_val = to_float(anthropometry_height, "Anthropometry Height")
        weight_val = to_float(anthropometry_weight, "Anthropometry Weight")
        age_val = to_float(anthropometry_age, "Anthropometry Age")
        chest_val = to_float(anthropometry_chest, "Anthropometry Chest")
        mid_axillary_val = to_float(anthropometry_mid_axillary, "Anthropometry Mid Axillary")
        subscapular_val = to_float(anthropometry_subscapular, "Anthropometry Subscapular")
        triceps_val = to_float(anthropometry_triceps, "Anthropometry Triceps")
        abdomen_val = to_float(anthropometry_abdomen, "Anthropometry Abdomen")
        suprailiac_val = to_float(anthropometry_suprailiac, "Anthropometry Suprailiac")
        mid_thigh_val = to_float(anthropometry_mid_thigh, "Anthropometry Mid Thigh")
        total_skinfold_val = to_float(anthropometry_total_skinfold, "Anthropometry Total Skinfold")
        body_density_val = to_float(anthropometry_body_density, "Anthropometry Body Density")
        fat_percent_val = to_float(anthropometry_fat_percentage, "Anthropometry Fat Percentage")
        error_corrected_val = to_float(anthropometry_error_corrected, "Anthropometry Error Corrected")
        chest_n_val = to_float(anthropometry_chest_n, "Anthropometry Chest N")
        chest_e_val = to_float(anthropometry_chest_e, "Anthropometry Chest E")
        upper_arm_val = to_float(anthropometry_upper_arm, "Anthropometry Upper Arm")
        waist_val = to_float(anthropometry_waist, "Anthropometry Waist")
        abdomen_cm_val = to_float(anthropometry_abdomen_cm, "Anthropometry Abdomen N")
        hip_val = to_float(anthropometry_hip, "Anthropometry Hip")
        thigh_val = to_float(anthropometry_thigh, "Anthropometry Thigh")
        calf_val = to_float(anthropometry_calf, "Anthropometry Calf")

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            # Save into dedicated Anthropometry model
            AnthropometryTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,

                anthropometry_height=height_val,
                anthropometry_weight=weight_val,
                anthropometry_age=age_val,
                anthropometry_chest=chest_val,
                anthropometry_mid_axillary=mid_axillary_val,
                anthropometry_subscapular=subscapular_val,
                anthropometry_triceps=triceps_val,
                anthropometry_abdomen=abdomen_val,
                anthropometry_suprailiac=suprailiac_val,
                anthropometry_mid_thigh=mid_thigh_val,
                anthropometry_total_skinfold=total_skinfold_val,
                anthropometry_body_density=body_density_val,
                anthropometry_fat_percentage=fat_percent_val,
                anthropometry_error_corrected=error_corrected_val,
                anthropometry_chest_n=chest_n_val,
                anthropometry_chest_e=chest_e_val,
                anthropometry_upper_arm=upper_arm_val,
                anthropometry_waist=waist_val,
                anthropometry_abdomen_cm=abdomen_cm_val,
                anthropometry_hip=hip_val,
                anthropometry_thigh=thigh_val,
                anthropometry_calf=calf_val,
                gender=player.gender,
                category=player.age_category,
                notes=notes,
                reported_by=reported_by_user,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('test_dashboard_new')

        return render(request, 'player_app/tests/anthropometry.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/anthropometry.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


# Anthropometry Test Data views
def anthropometry_test_view(request):
    test_name = "Anthropometry Test"
    search_query = request.GET.get('search', '').strip()

    # Use AnthropometryTest model, not TestAndResult
    results = (
        AnthropometryTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/anthropometry_data.html', context)


# Dexa Scan Test views
def add_dexa_scan_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "DEXA Scan Test"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        dexa_height = request.POST.get('dexa_height')
        dexa_weight = request.POST.get('dexa_weight')
        dexa_bmi = request.POST.get('dexa_bmi')
        dexa_rmr = request.POST.get('dexa_rmr')
        dexa_bmd = request.POST.get('dexa_bmd')
        dexa_tscore = request.POST.get('dexa_tscore')
        dexa_total_fat = request.POST.get('dexa_total_fat')
        dexa_lean = request.POST.get('dexa_lean')
        dexa_lean_mass = request.POST.get('dexa_lean_mass')
        dexa_testosterone = request.POST.get('dexa_testosterone')

        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not dexa_height:
            errors.append("Dexa Height is required.")
        if not dexa_weight:
            errors.append("Dexa Weight is required.")
        if not dexa_bmi:
            errors.append("Dexa BMI is required.")
        if not dexa_rmr:
            errors.append("Dexa RMR is required.")
        if not dexa_bmd:
            errors.append("Dexa BMD is required.")
        if not dexa_tscore:
            errors.append("Dexa T Score is required.")
        if not dexa_total_fat:
            errors.append("Dexa Total Fat is required.")
        if not dexa_lean:
            errors.append("Dexa Lean is required.")
        if not dexa_lean_mass:
            errors.append("Dexa Lean Mass is required.")
        if not dexa_testosterone:
            errors.append("Dexa Testosterone is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # Helper for numeric fields
        def to_float(value, field_name):
            nonlocal errors
            if value in (None, ""):
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                errors.append(f"{field_name} must be a valid number.")
                return None

        height_val = to_float(dexa_height, "Dexa Height")
        weight_val = to_float(dexa_weight, "Dexa Weight")
        bmi_val = to_float(dexa_bmi, "Dexa BMI")
        rmr_val = to_float(dexa_rmr, "Dexa RMR")
        bmd_val = to_float(dexa_bmd, "Dexa BMD")
        tscore_val = to_float(dexa_tscore, "Dexa T Score")
        total_fat_val = to_float(dexa_total_fat, "Dexa Total Fat")
        lean_val = to_float(dexa_lean, "Dexa Lean")
        lean_mass_val = to_float(dexa_lean_mass, "Dexa Lean Mass")
        testosterone_val = to_float(dexa_testosterone, "Dexa Testosterone")

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            # Save to DexaScanTest model
            DexaScanTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,

                dexa_height=height_val,
                dexa_weight=weight_val,
                dexa_bmi=bmi_val,
                dexa_rmr=rmr_val,
                dexa_bmd=bmd_val,
                dexa_tscore=tscore_val,
                dexa_total_fat=total_fat_val,
                dexa_lean=lean_val,
                dexa_lean_mass=lean_mass_val,
                dexa_testosterone=testosterone_val,
                gender=player.gender,
                category=player.age_category,
                notes=notes,
                reported_by=reported_by_user,
            )
            
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('test_dashboard_new')

        return render(request, 'player_app/tests/dexa_scan.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/dexa_scan.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })

# Dexa Scan Test Data views
def dexa_scan_test_view(request):
    test_name = "DEXA Scan Test"
    search_query = request.GET.get('search', '').strip()

    # Use DexaScanTest model instead of TestAndResult
    results = (
        DexaScanTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/dexa_scan_data.html', context)

# Blood Work Test views
def add_blood_test(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "Blood Work"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')

        blood_hemoglobin = request.POST.get('blood_hemoglobin')
        blood_rbc = request.POST.get('blood_rbc')
        blood_platelets = request.POST.get('blood_platelets')
        blood_albumin = request.POST.get('blood_albumin')
        blood_globulin = request.POST.get('blood_globulin')
        blood_uric_acid = request.POST.get('blood_uric_acid')
        blood_creatinine = request.POST.get('blood_creatinine')
        blood_testosterone = request.POST.get('blood_testosterone')
        blood_iron = request.POST.get('blood_iron')
        blood_vitamin_d3 = request.POST.get('blood_vitamin_d3')
        blood_cholesterol = request.POST.get('blood_cholesterol')
        blood_hdl = request.POST.get('blood_hdl')
        blood_ldl = request.POST.get('blood_ldl')
        blood_ldl_hdl_ratio = request.POST.get('blood_ldl_hdl_ratio')
        blood_vitamin_b12 = request.POST.get('blood_vitamin_b12')
        blood_lipoprotein = request.POST.get('blood_lipoprotein')
        blood_homocysteine = request.POST.get('blood_homocysteine')
        blood_protein = request.POST.get('blood_protein')
        blood_t3 = request.POST.get('blood_t3')
        blood_t4 = request.POST.get('blood_t4')
        blood_tsh = request.POST.get('blood_tsh')

        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []

        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not blood_hemoglobin:
            errors.append("Blood Hemoglobin is required.")
        if not blood_rbc:
            errors.append("Blood RBC is required.")
        if not blood_platelets:
            errors.append("Blood Platelets is required.")
        if not blood_albumin:
            errors.append("Blood Albumin is required.")
        if not blood_globulin:
            errors.append("Blood Globulin is required.")
        if not blood_uric_acid:
            errors.append("Blood Uric Acid is required.")
        if not blood_creatinine:
            errors.append("Blood Creatinine is required.")
        if not blood_testosterone:
            errors.append("Blood Testosterone is required.")
        if not blood_iron:
            errors.append("Blood Iron is required.")
        if not blood_vitamin_d3:
            errors.append("Blood Vitamin D3 is required.")
        if not blood_cholesterol:
            errors.append("Blood Cholesterol is required.")
        if not blood_hdl:
            errors.append("Blood HDL is required.")
        if not blood_ldl:
            errors.append("Blood LDL is required.")
        if not blood_ldl_hdl_ratio:
            errors.append("Blood LDL/HDL Ratio is required.")
        if not blood_vitamin_b12:
            errors.append("Blood Vitamin B12 is required.")
        if not blood_lipoprotein:
            errors.append("Blood Lipoprotein is required.")
        if not blood_homocysteine:
            errors.append("Blood Homocysteine is required.")
        if not blood_protein:
            errors.append("Blood Protein is required.")
        if not blood_t3:
            errors.append("Blood T3 is required.")
        if not blood_t4:
            errors.append("Blood T4 is required.")
        if not blood_tsh:
            errors.append("Blood TSH is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # helper for numeric fields (if model uses FloatField/DecimalField)
        def to_float(value, field_name):
            nonlocal errors
            if value in (None, ""):
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                errors.append(f"{field_name} must be a valid number.")
                return None

        # cast numerics (adjust if some are CharField)
        hb_val = to_float(blood_hemoglobin, "Blood Hemoglobin")
        rbc_val = to_float(blood_rbc, "Blood RBC")
        platelets_val = to_float(blood_platelets, "Blood Platelets")
        albumin_val = to_float(blood_albumin, "Blood Albumin")
        globulin_val = to_float(blood_globulin, "Blood Globulin")
        uric_acid_val = to_float(blood_uric_acid, "Blood Uric Acid")
        creatinine_val = to_float(blood_creatinine, "Blood Creatinine")
        testosterone_val = to_float(blood_testosterone, "Blood Testosterone")
        iron_val = to_float(blood_iron, "Blood Iron")
        vit_d3_val = to_float(blood_vitamin_d3, "Blood Vitamin D3")
        chol_val = to_float(blood_cholesterol, "Blood Cholesterol")
        hdl_val = to_float(blood_hdl, "Blood HDL")
        ldl_val = to_float(blood_ldl, "Blood LDL")
        ldl_hdl_ratio_val = to_float(blood_ldl_hdl_ratio, "Blood LDL/HDL Ratio")
        vit_b12_val = to_float(blood_vitamin_b12, "Blood Vitamin B12")
        lipoprotein_val = to_float(blood_lipoprotein, "Blood Lipoprotein")
        homocysteine_val = to_float(blood_homocysteine, "Blood Homocysteine")
        protein_val = to_float(blood_protein, "Blood Protein")
        t3_val = to_float(blood_t3, "Blood T3")
        t4_val = to_float(blood_t4, "Blood T4")
        tsh_val = to_float(blood_tsh, "Blood TSH")

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User for reported_by
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # change if relation name differs

            BloodTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,

                blood_hemoglobin=hb_val,
                blood_rbc=rbc_val,
                blood_platelets=platelets_val,
                blood_albumin=albumin_val,
                blood_globulin=globulin_val,
                blood_uric_acid=uric_acid_val,
                blood_creatinine=creatinine_val,
                blood_testosterone=testosterone_val,
                blood_iron=iron_val,
                blood_vitamin_d3=vit_d3_val,
                blood_cholesterol=chol_val,
                blood_hdl=hdl_val,
                blood_ldl=ldl_val,
                blood_ldl_hdl_ratio=ldl_hdl_ratio_val,
                blood_vitamin_b12=vit_b12_val,
                blood_lipoprotein=lipoprotein_val,
                blood_homocysteine=homocysteine_val,
                blood_protein=protein_val,
                blood_t3=t3_val,
                blood_t4=t4_val,
                blood_tsh=tsh_val,
                gender=player.gender,
                category=player.age_category,
                notes=notes,
                reported_by=reported_by_user,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('test_dashboard_new')

        return render(request, 'player_app/tests/blood_work.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/blood_work.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


# Blood Work Test Data views
def blood_test_view(request,test_name=None):
    search_query = request.GET.get('search', '').strip()

    # Use BloodTest model instead of TestAndResult
    results = (
        BloodTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/blood_work_data.html', context)


# Run A 3 test view
def add_runa3_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = 'Run A 3'

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation for best
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # NomativeData lookup
            nomative_data = get_object_or_404(NomativeData, final_level=best_val)
            total_distance = nomative_data.total_distance
            approximately_vo2max = nomative_data.approximately_vo2max

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            # Save into RunA3Test; model will update aggregates/min/max/avg
            RunA3Test.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('runa3_test_view')

        return render(request, 'player_app/tests/runa3.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/runa3.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


def runa3_test_view(request):
    test_name = 'Run A 3'
    search_query = request.GET.get('search', '').strip()

    results = (
        RunA3Test.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (from RunA3Test)
    latest_dates = (
        RunA3Test.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            RunA3Test.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/runa3_data.html', context)


# 40 Meter Test views
def add_forty_meter_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = '40m'   # or exactly what you use in UI

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # NomativeData lookup
            nomative_data = get_object_or_404(NomativeData, final_level=best_val)
            total_distance = nomative_data.total_distance
            approximately_vo2max = nomative_data.approximately_vo2max

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            FortyMeterTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))           
            return redirect('forty_meter_test_view')

        return render(request, 'player_app/tests/forty_meter.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/forty_meter.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })

def forty_meter_test_view(request):
    test_name = '40m'
    search_query = request.GET.get('search', '').strip()

    results = (
        FortyMeterTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in model.save())
    latest_dates = (
        FortyMeterTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            FortyMeterTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/forty_meter_data.html', context)


# 20 Meter Test views
def add_twenty_meter_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = '20m'   # must match what you use in UI/TestAndResult choices

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # NomativeData lookup (same logic you used for RunA3/40m)
            nomative_data = get_object_or_404(NomativeData, final_level=best_val)
            total_distance = nomative_data.total_distance
            approximately_vo2max = nomative_data.approximately_vo2max

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            TwentyMeterTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))

            return redirect('twenty_meter_test_view')

        return render(request, 'player_app/tests/twenty_meter.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/twenty_meter.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


def twenty_meter_test_view(request):
    test_name = '20m'
    search_query = request.GET.get('search', '').strip()

    results = (
        TwentyMeterTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in model.save())
    latest_dates = (
        TwentyMeterTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            TwentyMeterTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/twenty_meter_data.html', context)


# 10 Meter Test views
def add_ten_meter_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = '10m'   # keep consistent with UI / choices

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        
        
        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # NomativeData lookup (same table as other sprint tests)
            nomative_data = get_object_or_404(NomativeData, final_level=best_val)
            total_distance = nomative_data.total_distance
            approximately_vo2max = nomative_data.approximately_vo2max

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            TenMeterTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('ten_meter_test_view')

        return render(request, 'player_app/tests/ten_meter.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/ten_meter.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


def ten_meter_test_view(request):
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']

    test_name = '10m'
    search_query = request.GET.get('search', '').strip()

    results = (
        TenMeterTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in model.save())
    latest_dates = (
        TenMeterTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            TenMeterTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/ten_meter_data.html', context)




def add_sbj_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = 'SBJ'   # label used in the UI

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            SBJTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('sbj_test_view')

        return render(request, 'player_app/tests/sbj.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/sbj.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })

def sbj_test_view(request):
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']
    test_name = 'SBJ'
    search_query = request.GET.get('search', '').strip()

    results = (
        SBJTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in SBJTest.save())
    latest_dates = (
        SBJTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            SBJTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/sbj_data.html', context)



def add_yoyo_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = 'YoYo Test'   # label in UI

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Optional: NomativeData lookup if YoYo uses same table
            try:
                nomative_data = NomativeData.objects.get(final_level=best_val)
                total_distance = nomative_data.total_distance
                approximately_vo2max = nomative_data.approximately_vo2max
            except NomativeData.DoesNotExist:
                total_distance = None
                approximately_vo2max = None

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            YoYoTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('yoyo_test_view')

        return render(request, 'player_app/tests/yoyo.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/yoyo.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


def yoyo_test_view(request):
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']
    test_name = 'YoYo Test'
    search_query = request.GET.get('search', '').strip()

    results = (
        YoYoTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in YoYoTest.save())
    latest_dates = (
        YoYoTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            YoYoTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/yoyo_data.html', context)


def add_one_mile_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = '1 Mile'   # must match TestAndResult / TEST_CHOICES label

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Optional: if you have a NomativeData mapping for 1 Mile
            try:
                nomative_data = NomativeData.objects.get(final_level=best_val)
                total_distance = nomative_data.total_distance
                approximately_vo2max = nomative_data.approximately_vo2max
            except NomativeData.DoesNotExist:
                total_distance = None
                approximately_vo2max = None

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            OneMileTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('one_mile_test_view')

        return render(request, 'player_app/tests/one_mile.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/one_mile.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


def one_mile_test_view(request):
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']
    test_name = '1 Mile'
    search_query = request.GET.get('search', '').strip()

    results = (
        OneMileTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in OneMileTest.save())
    latest_dates = (
        OneMileTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            OneMileTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/one_mile_data.html', context)


def add_two_km_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = '2 KM'   # must match TestAndResult.TEST_CHOICES label

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Optional: NomativeData mapping for 2 KM
            try:
                nomative_data = NomativeData.objects.get(final_level=best_val)
                total_distance = nomative_data.total_distance
                approximately_vo2max = nomative_data.approximately_vo2max
            except NomativeData.DoesNotExist:
                total_distance = None
                approximately_vo2max = None

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            TwoKmTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                distance_covered=total_distance,
                predicted_vo2max=approximately_vo2max,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )

            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('two_km_test_view')

        return render(request, 'player_app/tests/two_km.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/two_km.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })

def two_km_test_view(request):
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']
    test_name = '2 KM'
    search_query = request.GET.get('search', '').strip()

    results = (
        TwoKmTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in TwoKmTest.save())
    latest_dates = (
        TwoKmTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            TwoKmTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/two_km_data.html', context)

# Push-ups Test views
def add_pushups_test(request):
    user_organization = getattr(request.user, 'organization', None)

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)
    staff = Staff.objects.filter(organization=user_organization)
    test_name = 'Push-ups'   # must match TEST_NAME in model

    if request.method == "POST":
        player_id = request.POST.get('player')
        test = request.POST.get('test')
        date = request.POST.get('date')
        phase_id = request.POST.get('phase')
        best = request.POST.get('best')   # max reps
        notes = request.POST.get('notes')
        reported_by_staff_id = request.POST.get('reported_by')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not test:
            errors.append("Test is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")
        if not best:
            errors.append("Best is required.")
        if not reported_by_staff_id:
            errors.append("Reported by is required.")

        # numeric validation for reps
        try:
            best_val = float(best) if best not in (None, "") else None
        except (TypeError, ValueError):
            errors.append("Best must be a valid number.")
            best_val = None

        if not errors:
            phase_obj = get_object_or_404(CampTournament, id=int(phase_id))
            player = get_object_or_404(Player, pk=player_id)

            # Staff → User
            staff_obj = get_object_or_404(Staff, pk=reported_by_staff_id)
            reported_by_user = staff_obj.user  # adjust if field name differs

            PushUpsTest.objects.create(
                player=player,
                date=date,
                phase=phase_obj,
                best=best_val,
                notes=notes,
                reported_by=reported_by_user,
                gender=player.gender,
                category=player.age_category,
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('pushups_test_view')

        return render(request, 'player_app/tests/pushups.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
            'staff': staff,
        })

    return render(request, 'player_app/tests/pushups.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
        'staff': staff,
    })


def pushups_test_view(request):
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']
    test_name = 'Push-ups'
    search_query = request.GET.get('search', '').strip()

    results = (
        PushUpsTest.objects
        .select_related('player', 'phase', 'reported_by')
        .order_by('-date')
    )

    if search_query:
        results = results.filter(player__name__icontains=search_query)

    paginator = Paginator(results, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    player_ids = results.values_list('player_id', flat=True).distinct()

    # latest per-player individual_average (computed in PushUpsTest.save())
    latest_dates = (
        PushUpsTest.objects
        .filter(
            player_id__in=player_ids,
            individual_average__isnull=False,
        )
        .values('player_id')
        .annotate(latest_date=Max('date'))
    )

    latest_indv_avgs = {}
    for entry in latest_dates:
        latest_result = (
            PushUpsTest.objects
            .filter(
                player_id=entry['player_id'],
                date=entry['latest_date'],
                individual_average__isnull=False,
            )
            .first()
        )
        if latest_result:
            latest_indv_avgs[entry['player_id']] = latest_result.individual_average

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'latest_indv_avgs': latest_indv_avgs,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/pushups_data.html', context)



def add_msk_injury_assessment(request):
    user_organization = getattr(request.user, 'organization', None)
    test_name = "MSK Injury Assessment"

    players = Player.objects.filter(organization=user_organization)
    events = CampTournament.objects.filter(is_deleted=False)

    if request.method == "POST":
        player_id = request.POST.get('player')
        physiotherapist_name = request.POST.get('physiotherapist_name')
        date = request.POST.get('date')

        height = request.POST.get('height')
        weight = request.POST.get('weight')

        skills = request.POST.get('skills')

        participation_category = request.POST.get('participation_category')
        dominant_side = request.POST.get('dominant_side')
        lead_leg = request.POST.get('lead_leg')

        rear_foot_right = request.POST.get('rear_foot_right')
        rear_foot_left = request.POST.get('rear_foot_left')
        mid_foot_right = request.POST.get('mid_foot_right')
        mid_foot_left = request.POST.get('mid_foot_left')

        lld = request.POST.get('lld')

        df_lunge_right = request.POST.get('df_lunge_right')
        df_lunge_left = request.POST.get('df_lunge_left')

        tibial_dial_right = request.POST.get('tibial_dial_right')
        tibial_dial_left = request.POST.get('tibial_dial_left')

        hip_ir_90_supine_right = request.POST.get('hip_ir_90_supine_right')
        hip_ir_90_supine_left = request.POST.get('hip_ir_90_supine_left')

        fabers_right = request.POST.get('fabers_right')
        fabers_left = request.POST.get('fabers_left')

        shoulder_ir_90_right = request.POST.get('shoulder_ir_90_right')
        shoulder_ir_90_left = request.POST.get('shoulder_ir_90_left')

        shoulder_er_90_right = request.POST.get('shoulder_er_90_right')
        shoulder_er_90_left = request.POST.get('shoulder_er_90_left')

        pec_minor_length_right = request.POST.get('pec_minor_length_right')
        pec_minor_length_left = request.POST.get('pec_minor_length_left')

        thoracic_rotation_right = request.POST.get('thoracic_rotation_right')
        thoracic_rotation_left = request.POST.get('thoracic_rotation_left')

        thomas_pos1_right = request.POST.get('thomas_pos1_right')
        thomas_pos1_left = request.POST.get('thomas_pos1_left')

        thomas_pos2_right = request.POST.get('thomas_pos2_right')
        thomas_pos2_left = request.POST.get('thomas_pos2_left')

        thomas_pos3_right = request.POST.get('thomas_pos3_right')
        thomas_pos3_left = request.POST.get('thomas_pos3_left')

        thomas_pos4_right = request.POST.get('thomas_pos4_right')
        thomas_pos4_left = request.POST.get('thomas_pos4_left')

        anterior_apprehension_right = request.POST.get('anterior_apprehension_right')
        anterior_apprehension_left = request.POST.get('anterior_apprehension_left')

        relocation_right = request.POST.get('relocation_right')
        relocation_left = request.POST.get('relocation_left')

        gird_erg_right = request.POST.get('gird_erg_right')
        gird_erg_left = request.POST.get('gird_erg_left')

        hk_test_right = request.POST.get('hk_test_right')
        hk_test_left = request.POST.get('hk_test_left')

        obriens_right = request.POST.get('obriens_right')
        obriens_left = request.POST.get('obriens_left')

        phase_id = request.POST.get('phase')
        comments = request.POST.get('comments')

        errors = []
        if not player_id:
            errors.append("Player is required.")
        if not physiotherapist_name:
            errors.append("Physiotherapist name is required.")
        if not date:
            errors.append("Date is required.")
        if not phase_id:
            errors.append("Phase is required.")

        # helper to cast optional floats
        def to_float(value):
            if value in (None, ""):
                return None
            try:
                return float(value)
            except (TypeError, ValueError):
                return None

        if not errors:
            player = get_object_or_404(Player, pk=player_id)
            phase_obj = get_object_or_404(CampTournament, pk=phase_id)

            MSKInjuryAssessment.objects.create(
                physiotherapist_name=physiotherapist_name,
                player=player,
                date=date,
                height=to_float(height),
                weight=to_float(weight),
                skills=skills,
                participation_category=participation_category,
                dominant_side=dominant_side,
                lead_leg=lead_leg,
                rear_foot_right=rear_foot_right,
                rear_foot_left=rear_foot_left,
                mid_foot_right=mid_foot_right,
                mid_foot_left=mid_foot_left,
                lld=lld,
                df_lunge_right=to_float(df_lunge_right),
                df_lunge_left=to_float(df_lunge_left),
                tibial_dial_right=to_float(tibial_dial_right),
                tibial_dial_left=to_float(tibial_dial_left),
                hip_ir_90_supine_right=to_float(hip_ir_90_supine_right),
                hip_ir_90_supine_left=to_float(hip_ir_90_supine_left),
                fabers_right=to_float(fabers_right),
                fabers_left=to_float(fabers_left),
                shoulder_ir_90_right=to_float(shoulder_ir_90_right),
                shoulder_ir_90_left=to_float(shoulder_ir_90_left),
                shoulder_er_90_right=to_float(shoulder_er_90_right),
                shoulder_er_90_left=to_float(shoulder_er_90_left),
                pec_minor_length_right=to_float(pec_minor_length_right),
                pec_minor_length_left=to_float(pec_minor_length_left),
                thoracic_rotation_right=to_float(thoracic_rotation_right),
                thoracic_rotation_left=to_float(thoracic_rotation_left),
                thomas_pos1_right=to_float(thomas_pos1_right),
                thomas_pos1_left=to_float(thomas_pos1_left),
                thomas_pos2_right=to_float(thomas_pos2_right),
                thomas_pos2_left=to_float(thomas_pos2_left),
                thomas_pos3_right=thomas_pos3_right,
                thomas_pos3_left=thomas_pos3_left,
                thomas_pos4_right=to_float(thomas_pos4_right),
                thomas_pos4_left=to_float(thomas_pos4_left),
                anterior_apprehension_right=anterior_apprehension_right,
                anterior_apprehension_left=anterior_apprehension_left,
                relocation_right=relocation_right,
                relocation_left=relocation_left,
                gird_erg_right=gird_erg_right,
                gird_erg_left=gird_erg_left,
                hk_test_right=hk_test_right,
                hk_test_left=hk_test_left,
                obriens_right=obriens_right,
                obriens_left=obriens_left,
                phase=phase_obj,
                comments=comments,
                
            )
            if 'phase_id_test' in request.session:
                return redirect('phase_tests_view',id=request.session.get('phase_id_test'))
            return redirect('msk_injury_assessment_list')

        return render(request, 'player_app/tests/msk_injury_assessment_form.html', {
            'test_name': test_name,
            'errors': errors,
            'players': players,
            'events': events,
        })

    return render(request, 'player_app/tests/msk_injury_assessment_form.html', {
        'test_name': test_name,
        'players': players,
        'events': events,
    })

def msk_injury_assessment_list(request):
    
    if 'phase_id_test' in request.session:
        del request.session['phase_id_test']
    test_name = "MSK Injury Assessment"
    search_query = request.GET.get('search', '').strip()

    qs = (
        MSKInjuryAssessment.objects
        .select_related('player', 'phase')
        .order_by('-date')
    )

    if search_query:
        qs = qs.filter(
            models.Q(player__name__icontains=search_query) |
            models.Q(physiotherapist_name__icontains=search_query)
        )

    paginator = Paginator(qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'test_name': test_name,
        'page_obj': page_obj,
        'total_results': paginator.count,
        'search_query': search_query,
    }
    return render(request, 'player_app/tests/msk_injury_assessment_list.html', context)

from django.http import Http404
# Organization Player Tests view for all test types
@login_required
def organization_player_tests(request, player_id):
    TEST_MODEL_MAP = {
        # Single-value tests
        '10m': TenMeterTest,
        '20m': TwentyMeterTest,
        '40m': FortyMeterTest,
        'yoyo': YoYoTest,
        'sbj': SBJTest,
        'run_a_3': RunA3Test,
        'run_a_3x6': None,
        '1_mile': OneMileTest,
        'push_ups': PushUpsTest,
        '2_km': TwoKmTest,

        # Asymmetry tests - MATCH YOUR TEMPLATE KEYS
        's_l_glute_bridges': SLGluteBridges,
        'sl_lunge_calf_raises': SLLungeCalfRaises,
        'mb_rotational_throws': MBRotationalThrows,
        'copenhagen': CopenhagenTest,
        's_l_hop': SLHopTest,

        # Complex tests
        'cmj_scores': CMJTest,
        'anthropometry': AnthropometryTest,
        'blood_work': BloodTest,
        'dexa_scan_test': DexaScanTest,
        'msk_injury_assessment': MSKInjuryAssessment,
    }

    player = get_object_or_404(Player, pk=player_id)
    
    
    # Org-safety check
    user_organization = getattr(request.user, 'organization', None)
    if user_organization and player.organization_id != user_organization.id:
        raise Http404("Player not found")

    context = {"player": player}

    # Query each test type
    for test_name, model_cls in TEST_MODEL_MAP.items():
        if model_cls:
            try:
                # Add this debug for SLGluteBridges specifically
                if test_name == 's_l_glute_bridges':
                    print("SLG fields:", [f.name for f in SLGluteBridges._meta.fields])
                    print("Direct query:", SLGluteBridges.objects.filter(player=player).count())

                tests = model_cls.objects.filter(player=player).order_by('-date', '-created_at')[:10]

                
                # Key will now be exactly: tests_s_l_glute_bridges, etc.
                key = f"tests_{test_name}"
                context[key] = tests
                
            except Exception:
                context[f"tests_{test_name}"] = []
        else:
            context[f"tests_{test_name}"] = []
    # Add this before return render() to debug
    print("=== CONTEXT KEYS ===")
    print(SLGluteBridges.objects.filter(player=player))
    for key in context.keys():
        if key.startswith('tests_'):
            print(f"{key}: {len(context[key])} items")
    print("====================")

    return render(request, 'player_app/tests/organization_player_tests.html', context)

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Q
from django.utils import timezone
from datetime import datetime
import json

@login_required
def daily_wellness_create_view(request):
    """Daily wellness form with dynamic player filtering"""
    user_org = getattr(request.user, "organization", None)
    
    all_players = Player.objects.filter(organization=user_org).order_by('name')
    phases = CampTournament.objects.filter(organization=user_org, is_deleted=False)
    session_phase_id = request.session.get('phase_id_test')
    
    context = {
        "players": all_players,
        "phases": phases,
        "session_phase_id": session_phase_id,
        "range_1_10": range(1, 11),
        "range_0_10": range(0, 11),
    }
    
    if request.method == "POST":
        return handle_wellness_submission(request, user_org)
    
    return render(request, "player_app/tests/daily_wellness_form.html", context)


@csrf_exempt
@login_required
def filter_players_ajax(request):
    """AJAX endpoint: Filter players participating in specific phase ON SELECTED DATE"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'POST only'})
    
    try:
        phase_id = request.POST.get('phase_id')
        date_str = request.POST.get('date')
        user_org = getattr(request.user, "organization", None)
        
        if not phase_id or not date_str:
            return JsonResponse({
                'success': True, 
                'players': [],
                'message': 'Please select phase and date',
                'count': 0
            })
        
        selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        phase = get_object_or_404(
            CampTournament, 
            id=phase_id, 
            organization=user_org, 
            is_deleted=False
        )
        

        active_players = Player.objects.filter(
            camps=phase,
            organization=user_org,
        ).distinct().order_by('name')
        

        
        players_data = []
        for player in active_players:
            players_data.append({
                'id': player.id,
                'name': player.name,
                'age': getattr(player, 'age', 'N/A'),
                'role': getattr(player, 'role', 'N/A'),
                'status': getattr(player, 'player_status', 'Active'),
            })
        
        return JsonResponse({
            'success': True,
            'players': players_data,
            'count': len(players_data),
            'message': f'{len(players_data)} players available'
        })
        
    except ValueError:
        return JsonResponse({'success': False, 'error': 'Invalid date format'})
    except Exception as e:
        print(f"ERROR: {e}")
        return JsonResponse({'success': False, 'error': str(e)})


def handle_wellness_submission(request, user_org):
    """Extracted form submission logic with validation"""
    try:

        

        phase_id = request.POST.get("phase_id") or request.POST.get("phase-id")
        player_id = request.POST.get("player_id")
        date = request.POST.get("date")
        

        
        if not all([phase_id, player_id, date]):
            messages.error(request, 'Phase, Player, and Date are required')
            return redirect('daily_wellness_create_view')
        
        # Get objects
        player = get_object_or_404(Player, id=player_id, organization=user_org)
        phase = get_object_or_404(
            CampTournament, 
            id=phase_id, 
            organization=user_org, 
            is_deleted=False
        )
        

        if not player.camps.filter(id=phase.id).exists():
            messages.error(request, f'{player.name} is not registered for {phase.name}')
            return redirect('daily_wellness_create_view')
        
        # Parse form data with safe defaults
        end_date = request.POST.get("enddate") or None
        urine_color = request.POST.get("urine_color", "")
        soreness_level = int(request.POST.get("soreness_level", 0))
        fatigue_level = int(request.POST.get("fatigue_level", 0))
        sleep_hours = float(request.POST.get("sleep_hours", 0))
        has_pain_raw = request.POST.get("has_pain", "false")
        has_pain = has_pain_raw.lower() == "true"
        pain_comment = request.POST.get("pain_comment", "").strip()
        motivation_level = int(request.POST.get("motivation_level", 0))
        balls_bowled = int(request.POST.get("balls_bowled", 0))
        training_session_types = request.POST.getlist("training_session_types")
        total_rpe = int(request.POST.get("total_rpe", 0))
        

        # Create wellness test
        wellness_test = DailyWellnessTest.objects.create(
            player=player,
            phase=phase,
            date=date,
            end_date=end_date,
            urine_color=urine_color,
            soreness_level=soreness_level,
            fatigue_level=fatigue_level,
            sleep_hours=sleep_hours,
            has_pain=has_pain,
            pain_comment=pain_comment,
            motivation_level=motivation_level,
            balls_bowled=balls_bowled,
            training_session_types=training_session_types,
            total_rpe=total_rpe,
            created_by=request.user,

        )
        

        messages.success(request, f'✅ Wellness report submitted for {player.name}!')
        
        if 'phase_id_test' in request.session:
            return redirect('phase_tests_view', id=request.session.get('phase_id_test'))
        return redirect('daily_wellness_results_view')
        
    except ValueError as e:

        messages.error(request, 'Please check numeric values (0-10 scale)')
        return redirect('daily_wellness_create_view')
    except Exception as e:

        messages.error(request, f'Error saving: {str(e)}')
        return redirect('daily_wellness_create_view')



@login_required
def daily_wellness_results_view(request):
    """Results view with search and pagination (unchanged logic, improved)"""
    user_org = getattr(request.user, "organization", None)
    
    if not user_org:
        messages.error(request, "No organization access")
        return redirect('dashboard')
    
    # Base queryset
    qs = DailyWellnessTest.objects.select_related("player", "phase").filter(
        player__organization=user_org
    )
    
    # Search by player name
    search_query = request.GET.get("search", "").strip()
    if search_query:
        qs = qs.filter(Q(player__name__icontains=search_query))
    
    total_results = qs.count()
    
    # Pagination
    page_number = request.GET.get("page", 1)
    paginator = Paginator(qs.order_by("-date", "-id"), 25)
    page_obj = paginator.get_page(page_number)
    
    context = {
        "page_obj": page_obj,
        "search_query": search_query,
        "total_results": total_results,
    }
    return render(
        request,
        "player_app/tests/daily_wellness_results.html",
        context,
    )


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime, timedelta
from django.utils import timezone
from django.db.models import Count, Q


def player_test_select(request):
    user_organization = getattr(request.user, 'organization', None)
    players = Player.objects.filter(organization=user_organization)

    if request.method == 'POST':
        player_id = request.POST.get('player')
        return redirect('player_test', player_id=player_id)

    return render(request, 'player_app/record/player_test.html', {
        'players': players,
    })



@csrf_exempt
@require_http_methods(["POST"])
def fetch_players(request):
    """Fetch players who have taken specific test within date range"""
    try:
        test_name = request.POST.get('test_name')
        start_date = request.POST.get('start_date', '')
        end_date = request.POST.get('end_date', '')
        
        players_data = []
        
        # Map test names to models (extend this for all your models)
        test_models = {
            '10m': TenMeterTest,
            '20m': TwentyMeterTest,
            '40m': FortyMeterTest,
            '1 Mile': OneMileTest,
            '2 KM': TwoKmTest,
            'Push-ups': PushUpsTest,
            'YoYo': YoYoTest,
            'Run A 3': RunA3Test,
            'SBJ': SBJTest,
            'S/L Hop': SLHopTest,
            'Copenhagen': CopenhagenTest,
            'MB Rotational Throws':MBRotationalThrows,
            'SL Lunge Calf Raises':SLLungeCalfRaises,
            'S/L Glute Bridges':SLGluteBridges,
            'CMJ Scores':CMJTest,
            'Anthropometry Test':AnthropometryTest,
            'DEXA Scan Test':DexaScanTest,
            'Blood Test':BloodTest,
            'MSK Injury Assessment':MSKInjuryAssessment,
            'run_a_3x6':RunA3x6Test,
            'Daily Wellness Test':DailyWellnessTest,
            # Add all other tests...
        }
        
        if test_name not in test_models:
            return JsonResponse({'success': False, 'error': 'Invalid test name'})
        
        Model = test_models[test_name]
        
        # Build queryset with date filtering
        queryset = Model.objects.all()
        
        if start_date and end_date:
            queryset = queryset.filter(
                date__range=[start_date, end_date]
            )
        
        # Get players with test count
        player_stats = queryset.values('player_id', 'player__name').annotate(
            total_tests=Count('id')
        ).order_by('-total_tests')
        
        for stat in player_stats:
            players_data.append({
                'id': stat['player_id'],
                'name': stat['player__name'],
                'total_tests': stat['total_tests'],
                'tests': list(queryset.filter(player_id=stat['player_id']).values('id', 'date')[:10])
            })
        
        return JsonResponse({
            'success': True,
            'players': players_data[:50]  # Limit to 50 players
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@csrf_exempt
@require_http_methods(["POST"])
def fetch_report(request):
    try:
        test_name = request.POST.get('test_name')
        player_id = request.POST.get('player_id')
        num_tests = int(request.POST.get('num_tests', 5))
        start_date = request.POST.get('start_date') or None
        end_date = request.POST.get('end_date') or None

        TEST_MODELS = {
            '10m': TenMeterTest,
            '20m': TwentyMeterTest,
            '40m': FortyMeterTest,
            '1 Mile': OneMileTest,
            '2 KM': TwoKmTest,
            'Push-ups': PushUpsTest,
            'YoYo': YoYoTest,
            'Run A 3': RunA3Test,
            'SBJ': SBJTest,
            'S/L Hop': SLHopTest,
            'Copenhagen': CopenhagenTest,
            'MB Rotational Throws':MBRotationalThrows,
            'SL Lunge Calf Raises':SLLungeCalfRaises,
            'S/L Glute Bridges':SLGluteBridges,
            'CMJ Scores':CMJTest,
            'Anthropometry Test':AnthropometryTest,
            'DEXA Scan Test':DexaScanTest,
            'Blood Test':BloodTest,
            'MSK Injury Assessment':MSKInjuryAssessment,
            'run_a_3x6':RunA3x6Test,
            'Daily Wellness Test':DailyWellnessTest,
        }

        Model = TEST_MODELS.get(test_name)
        if not Model:
            return JsonResponse({'success': False, 'error': 'Invalid test name'})

        qs = Model.objects.filter(player_id=player_id)
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        tests = qs.order_by('-date')[:num_tests]
        if not tests.exists():
            return JsonResponse({'success': False, 'error': 'No tests found'})

        player_info = {
            'id': player_id,
            'name': tests.first().player.name,
        }

        chart_data = {
            'labels': [t.date.strftime('%Y-%m-%d') for t in tests],
            'scores': [t.best for t in tests],   # using best
        }

        stats_data = []
        for t in tests:
            stats_data.append({
                'date': t.date.strftime('%Y-%m-%d'),
                'phase': t.phase.name if t.phase else '',
                'best': t.best,
                'notes': t.notes or '',
                'distance_covered': t.distance_covered,
                'predicted_vo2max': t.predicted_vo2max,
                'reported_by': t.reported_by.first_name if t.reported_by else '',
                'gender': t.gender,
                'category': t.category,
                'reported_by_designation': t.reported_by_designation,
                'min': t.min,
                'max': t.max,
                'individual_average': t.individual_average,
            })

        return JsonResponse({
            'success': True,
            'player_info': player_info,
            'chart_data': chart_data,
            'stats_data': stats_data,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

from django.shortcuts import render
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods

from django.db.models import Sum
@csrf_protect
@require_http_methods(["GET", "POST"])
def player_report(request):
    if request.method == "GET":
        session_settings = request.session.get('report_settings', {})

        context = {
            'session_settings': session_settings,
            'has_session_settings': bool(session_settings),
            'min_max_formula': session_settings.get('min_max_formula', 'all_players'),
            'min_is_better': session_settings.get('min_is_better', False),
            'grp_avg_option': session_settings.get('grp_avg_option', None),
        }

        return render(request, "player_app/record/player_test.html", context)

    # This is equivalent to your fetchReport JS:
    test_name = request.POST.get("test_name")
    player_id = request.POST.get("player_id")
    num_tests = int(request.POST.get("num_tests") or 0)
    start_date = request.POST.get("start_date") or None
    end_date = request.POST.get("end_date") or None

    if not test_name or not player_id or not num_tests:
        return render(
            request,
            "player_app/record/player_test.html",
            {"error": "Please select test, player and number of tests"},
        )

    session_settings = request.session.get('report_settings', {})
    min_max_formula = session_settings.get('min_max_formula', 'all_players')
    min_is_better = session_settings.get('min_is_better', False)
    grp_avg_option = session_settings.get('grp_avg_option', None)

    TEST_MODELS = {
        "10m": TenMeterTest,
        "20m": TwentyMeterTest,
        "40m": FortyMeterTest,
        "1 Mile": OneMileTest,
        "2 KM": TwoKmTest,
        "Push-ups": PushUpsTest,
        "YoYo": YoYoTest,
        "Run A 3": RunA3Test,
        "SBJ": SBJTest,
    }

    TEST_OTHERS = {
        "CMJ Scores": CMJTest,      # Add these
        "Anthropometry Test": AnthropometryTest,
        "Blood Test": BloodTest,
        "DEXA Scan Test": DexaScanTest,
        "MSK Injury Assessment": MSKInjuryAssessment,
    }

    TEST_SPECIAL = {
        "S/L Glute Bridges": SLGluteBridges,
        "SL Lunge Calf Raises": SLLungeCalfRaises,
        "MB Rotational Throws": MBRotationalThrows,
        "Copenhagen": CopenhagenTest,
        "S/L Hop": SLHopTest,
    }
    SINGLE_TEST = {
        "run_a_3x6":RunA3x6Test,
    }
    DAILY_WELLNESS = {
        "Daily Wellness Test":DailyWellnessTest,
    }
    if test_name in TEST_MODELS:
        Model = TEST_MODELS.get(test_name)
        if not Model:
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "Invalid test name"},
            )

        qs = Model.objects.filter(player_id=player_id)
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        tests = qs.order_by("-date")[:num_tests]
        if not tests.exists():
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "No tests found for this player and date range"},
            )

        first_test = tests.first()
        player = first_test.player

        player_info = Player.objects.get(id=player_id)

        if min_max_formula == 'all_players':
            min_value = [t.min for t in tests if t.min is not None]
            max_value = [t.max for t in tests if t.max is not None]
            max_value = max(max_value) if max_value else None
            min_value = min(min_value) if min_value else None

        elif min_max_formula == "all_players_by_gender":
            min_value = GenderAggregate.objects.filter(
                test=test_name, gender=player.gender
            ).values_list("min", flat=True)
            max_value = GenderAggregate.objects.filter(
                test=test_name, gender=player.gender
            ).values_list("max", flat=True)
            min_value = min(min_value) if min_value else None
            max_value = max(max_value) if max_value else None

        elif min_max_formula == "category_based":
            min_value = CategoryAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("min", flat=True)
            max_value = CategoryAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("max", flat=True)
            print("CATEGORY BASED MIN/MAX RAW:", min_value, max_value)
            min_value = min(min_value) if min_value else None
            max_value = max(max_value) if max_value else None
            print("CATEGORY BASED MIN/MAX:", min_value, max_value)

        if min_is_better:
            min_value, max_value = max_value, min_value

        group_average = qs.aggregate(avg=Avg('individual_average'))['avg'] or 0

        if grp_avg_option == "all_players_date":
            group_average = qs.aggregate(avg=Avg('individual_average'))['avg'] or 0
            
        
        elif grp_avg_option == "all_players_gender_date":
            group_average = GenderAggregate.objects.filter(
                test=test_name, gender=player.gender
            ).values_list("average", flat=True)[0]

        elif grp_avg_option == "category_based_date":
            group_average = CategoryAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("average", flat=True)[0]
        
        elif grp_avg_option == "camp_or_tournament":
            group_average = CampAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("average",flat=True)[0]

        individual_averages = [
                t.individual_average for t in tests if t.individual_average is not None
            ]

        individual_averages = tests.aggregate(total_avg=Sum('best'))['total_avg'] or 0
        individual_averages = individual_averages / len(tests) if tests else 0
        first_part = individual_averages - min_value
        second_part = max_value - min_value
        normalized_scores = first_part / second_part * 100 if second_part != 0 else None


        # Prepare chart data
        chart_data = {
            "labels": [t.date.strftime("%Y-%m-%d") for t in tests],
            "scores": [t.best for t in tests],
        }

        stats_data = []
        for t in tests:
            stats_data.append(
                {
                    "date": t.date,
                    "phase": t.phase.name if t.phase else "",
                    "best": t.best,
                    "notes": t.notes or "",
                    "distance_covered": t.distance_covered,
                    "predicted_vo2max": t.predicted_vo2max,
                    "reported_by": t.reported_by.first_name if t.reported_by else "",
                    "gender": t.gender,
                    "category": t.category,
                    "reported_by_designation": t.reported_by_designation,
                    "min": t.min,
                    "max": t.max,
                    "individual_average": t.individual_average,
                }
            )

        context = {
            "player_info": player_info,
            "stats_data": stats_data,
            "chart_data": chart_data,         # if you later render chart server‑side or pass to JS
            "selected_test": test_name,
            "num_tests": num_tests,
            "start_date": start_date,
            "end_date": end_date,
            "normalized_scores": normalized_scores,
            "individual_averages": individual_averages,
            "min_value": min_value,
            "max_value": max_value,
            "group_average": group_average,
            'session_settings': session_settings,
            'has_session_settings': bool(session_settings),
            'min_max_formula': session_settings.get('min_max_formula', 'all_players'),
            'min_is_better': session_settings.get('min_is_better', False),
            'grp_avg_option': session_settings.get('grp_avg_option', None),
            }
        return render(request, "player_app/record/player_test.html", context)
    
    elif test_name in TEST_OTHERS:
        Model = TEST_OTHERS.get(test_name)
        if not Model:
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "Invalid test name"},
            )

        # Query data
        qs = Model.objects.filter(player_id=player_id)
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])
        
        tests = qs.order_by("-date")[:num_tests]
        
        if not tests.exists():
            return render(request, "player_app/record/player_test.html", {"error": "No tests found"})

        # DYNAMIC TABLE DATA
        field_names = [f.name for f in Model._meta.get_fields() if getattr(f, "concrete", False)]
        table_data = []
        
        for obj in tests:
            row = {}
            for name in field_names:
                value = getattr(obj, name, None)
                if value is None:
                    row[name] = "-"
                elif hasattr(value, 'strftime'):
                    row[name] = value.strftime("%Y-%m-%d")
                elif hasattr(value, 'first_name'):
                    row[name] = getattr(value, 'first_name', str(value))
                elif hasattr(value, 'name'):
                    row[name] = getattr(value, 'name', str(value))
                else:
                    row[name] = str(value)
            table_data.append(row)

        table_json = json.dumps({'fields': field_names, 'data': table_data})

        context = {
            "table_json": table_json,
            "selected_test": test_name,
            "player_id": player_id,
            'session_settings': session_settings,
            'has_session_settings': bool(session_settings),
            'min_max_formula': session_settings.get('min_max_formula', 'all_players'),
            'min_is_better': session_settings.get('min_is_better', False),
            'grp_avg_option': session_settings.get('grp_avg_option', None),
        }
        return render(request, "player_app/record/player_test.html", context)

    elif test_name in TEST_SPECIAL:
        Model = TEST_SPECIAL.get(test_name)
        if not Model:
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "Invalid test name"},
            )

        # Query data
        qs = Model.objects.filter(player_id=player_id)
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        tests = qs.order_by("-date")[:num_tests]
        if not tests.exists():
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "No tests found for this player and date range"},
            )

        first_test = tests.first()
        player = first_test.player

        player_info = Player.objects.get(id=player_id)

        if min_max_formula == 'all_players':
            min_value_left = [t.left_min for t in tests if t.left_min is not None]
            max_value_left = [t.left_max for t in tests if t.left_max is not None]

            min_value_right = [t.right_min for t in tests if t.right_min is not None]
            max_value_right = [t.right_max for t in tests if t.right_max is not None]

            

        elif min_max_formula == "all_players_by_gender":
            min_value_left = GenderAggregate.objects.filter(test=test_name, gender=player.gender).values_list("left_min", flat=True)
            max_value_left = GenderAggregate.objects.filter(test=test_name, gender=player.gender).values_list("left_max", flat=True)

            min_value_right = GenderAggregate.objects.filter(test=test_name, gender=player.gender).values_list("right_min", flat=True)
            max_value_right = GenderAggregate.objects.filter(test=test_name, gender=player.gender).values_list("right_max", flat=True)

        
        elif min_max_formula == "category_based":
            min_value_left = CategoryAggregate.objects.filter(test=test_name, category=player.age_category).values_list("left_min", flat=True)
            max_value_left = CategoryAggregate.objects.filter(test=test_name, category=player.age_category).values_list("left_max", flat=True)
            
            min_value_right = CategoryAggregate.objects.filter(test=test_name, category=player.age_category).values_list("right_min", flat=True)
            max_value_right = CategoryAggregate.objects.filter(test=test_name, category=player.age_category).values_list("right_max", flat=True)


        if min_is_better:
            min_value_right, max_value_right = max_value_right, min_value_right
            min_value_left, max_value_left = max_value_left, min_value_left

        individual_averages_left = [
            t.individual_average_left for t in tests if t.individual_average_left is not None
        ]
        
        group_average_right = []
        group_average_left = []

        if grp_avg_option == "all_players_date":
            group_average_right = qs.aggregate(avg=Avg('individual_average_right'))['avg'] or 0
            group_average_left = qs.aggregate(avg=Avg('individual_average_left'))['avg'] or 0
        
        elif grp_avg_option == "all_players_gender_date":
            group_average_right = Model.objects.filter(gender=player.gender).values_list("right", flat=True)[0]  
            group_average_left = Model.objects.filter(gender=player.gender).values_list("left",flat=True)[0]
            

        elif grp_avg_option == "category_based_date":
            group_average_right = Model.objects.filter(category=player.age_category).values_list("right", flat=True)[0]
            group_average_left = Model.objects.filter(category=player.age_category).values_list("left",flat=True)[0]

        elif grp_avg_option == "camp_or_tournament":
            group_average_right = Model.objects.filter(category=player.age_category).values_list("right",flat=True)[0]
            group_average_left = Model.objects.filter(category=player.age_category).values_list("left",flat=True)[0]


        individual_averages_left = [t.individual_average_left for t in tests if t.individual_average_left is not None]
        first_part = individual_averages_left[0] - min_value_left[0]
        second_part = max_value_left[0] - min_value_left[0]
        normalized_scores_left = first_part / second_part * 100 if second_part != 0 else None
        # group_average_left = qs.aggregate(avg=Avg('individual_average_left'))['avg'] or 0
        
        individual_averages_right = [
            t.individual_average_right for t in tests if t.individual_average_right is not None
        ]

        individual_averages_right = [t.individual_average_right for t in tests if t.individual_average_right is not None]
        first_part = individual_averages_right[0] - min_value_right[0]
        second_part = max_value_right[0] - min_value_right[0]
        normalized_scores_right = first_part / second_part * 100 if second_part != 0 else None
        # group_average_right = qs.aggregate(avg=Avg('individual_average_right'))['avg'] or 0
    

    

        stats_data_special = []
        for t in tests:
            stats_data_special.append(
                {
                    "date": t.date,
                    "phase": t.phase.name if t.phase else "",
                    "right": t.right,
                    "left": t.left,
                    "difference": t.difference,
                    "ratio": t.ratio,
                    "gender": t.gender,
                    "category": t.category,
                    "notes": t.notes or "",
                    "reported_by": t.reported_by.first_name if t.reported_by else "",
                }
            )

        context = {
            "player_info_new": player_info,
            "stats_data_special": stats_data_special,
            
            "selected_test": test_name,
            "num_tests": num_tests,
            "start_date": start_date,
            "end_date": end_date,
            "normalized_scores_left": normalized_scores_left,
            "individual_averages_left": individual_averages_left[0],
            "min_value_left": min_value_left[0],
            "max_value_left": max_value_left[0],
            "group_average_left": group_average_left,
            "normalized_scores_right": normalized_scores_right,
            "individual_averages_right": individual_averages_right[0],
            "min_value_right": min_value_right[0],
            "max_value_right": max_value_right[0],
            "group_average_right": group_average_right,
            'session_settings': session_settings,
            'has_session_settings': bool(session_settings),
            'min_max_formula': session_settings.get('min_max_formula', 'all_players'),
            'min_is_better': session_settings.get('min_is_better', False),
            'grp_avg_option': session_settings.get('grp_avg_option', None),
        }
        return render(request, "player_app/record/player_test.html", context)
    

    elif test_name in SINGLE_TEST:
        Model = SINGLE_TEST.get(test_name)
        if not Model:
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "Invalid test name"},
            )

        qs = Model.objects.filter(player_id=player_id)
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        tests = qs.order_by("-date")[:num_tests]
        if not tests.exists():
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "No tests found for this player and date range"},
            )

        first_test = tests.first()
        player = first_test.player

        player_info = Player.objects.get(id=player_id)

        if min_max_formula == 'all_players':
            min_value = [t.min for t in tests if t.min is not None]
            max_value = [t.max for t in tests if t.max is not None]
            max_value = max(max_value)

        elif min_max_formula == "all_players_by_gender":
            min_value = GenderAggregate.objects.filter(
                test=test_name, gender=player.gender
            ).values_list("min", flat=True)
            max_value = GenderAggregate.objects.filter(
                test=test_name, gender=player.gender
            ).values_list("max", flat=True)
        
        elif min_max_formula == "category_based":
            min_value = CategoryAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("min", flat=True)
            max_value = CategoryAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("max", flat=True)


        

        if min_is_better:
            min_value, max_value = max_value, min_value

        group_average = qs.aggregate(avg=Avg('individual_average'))['avg'] or 0
        if grp_avg_option == "all_players_date":
            group_average = qs.aggregate(avg=Avg('individual_average'))['avg'] or 0
        
        elif grp_avg_option == "all_players_gender_date":
            group_average = GenderAggregate.objects.filter(
                test=test_name, gender=player.gender
            ).values_list("average", flat=True)[0]

        elif grp_avg_option == "category_based_date":
            group_average = CategoryAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("average", flat=True)[0]
        
        elif grp_avg_option == "camp_or_tournament":
            group_average = CampAggregate.objects.filter(
                test=test_name, category=player.age_category
            ).values_list("average",flat=True)[0]

        individual_averages = [t.individual_average for t in tests if t.individual_average is not None]
        first_part = individual_averages[0] - min_value[0]
        second_part = max_value - min_value[0]
        normalized_scores = first_part / second_part * 100 if second_part != 0 else None


        # Prepare chart data
        chart_data = {
            "labels": [t.date.strftime("%Y-%m-%d") for t in tests],
            
        }

        stats_data = []
        for t in tests:
            stats_data.append(
                {
                    "date": t.date,
                    "phase": t.phase.name if t.phase else "",
                    "trial1": t.trial1,
                    "trial2": t.trial2,
                    "trial3": t.trial3,
                    "trial4": t.trial4,
                    "trial5": t.trial5,
                    "trial6": t.trial6,
                    "average": t.average,
                    "notes": t.notes or "",
                    "reported_by": t.reported_by.first_name if t.reported_by else "",
                    "gender": t.gender,
                    "category": t.category,
                    "reported_by_designation": t.reported_by_designation,
                    "min": t.min,
                    "max": t.max,
                    "individual_average": t.individual_average,
                }
            )

        context = {
            "player_info": player_info,
            "stats_data_single": stats_data,
            "chart_data": chart_data,         # if you later render chart server‑side or pass to JS
            "selected_test": test_name,
            "num_tests": num_tests,
            "start_date": start_date,
            "end_date": end_date,
            "normalized_scores": normalized_scores,
            "individual_averages": individual_averages[0],
            "min_value": min_value[0],
            "max_value": max_value,
            "group_average": group_average,
            'session_settings': session_settings,
            'has_session_settings': bool(session_settings),
            'min_max_formula': session_settings.get('min_max_formula', 'all_players'),
            'min_is_better': session_settings.get('min_is_better', False),
            'grp_avg_option': session_settings.get('grp_avg_option', None),
            }
        return render(request, "player_app/record/player_test.html", context)
    
    elif test_name in DAILY_WELLNESS:
        Model = DAILY_WELLNESS.get(test_name)
        if not Model:
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "Invalid test name"},
            )

        qs = Model.objects.filter(player_id=player_id)
        if start_date and end_date:
            qs = qs.filter(date__range=[start_date, end_date])

        tests = qs.order_by("-date")[:num_tests]
        if not tests.exists():
            return render(
                request,
                "player_app/record/player_test.html",
                {"error": "No tests found for this player and date range"},
            )
        print(tests)
        stats_data_wellness = []
        for t in tests:
            stats_data_wellness.append(
                {
                    "date": t.date,
                    "phase": t.phase.name if t.phase else "",
                    "urine_color": t.urine_color,
                    "fatigue_level": t.fatigue_level,
                    "soreness_level": t.soreness_level,          
                    "sleep_hours": t.sleep_hours,
                    "has_pain": t.has_pain,
                    "pain_comment": t.pain_comment,
                    "motivation_level": t.motivation_level,
                    "balls_bowled": t.balls_bowled,
                    "training_session_types": t.training_session_types,
                    "total_rpe": t.total_rpe,
                    "reported_by": t.created_by.first_name if t.created_by else "",
                }
            )

        context = {
            "stats_data_wellness": tests,
            "selected_test": test_name,
            'session_settings': session_settings,
            'has_session_settings': bool(session_settings),
            'min_max_formula': session_settings.get('min_max_formula', 'all_players'),
            'min_is_better': session_settings.get('min_is_better', False),
            'grp_avg_option': session_settings.get('grp_avg_option', None),
        }
        return render(request, "player_app/record/player_test.html", context)
    

def multi_test_report_view(request):
    """Standalone Multi Test Report View"""
    players = Player.objects.filter(organization=request.user.organization).order_by('name')
    
    context = {
        'players': players,
        'selected_tests': [],
        'selected_player_id': '',
        'num_tests': '',
        'error': None,
    }
    
    if request.method == 'POST':
        try:
            # Extract form data
            date_option = request.POST.get('date_option')
            selected_tests = request.POST.getlist('tests')  # ['10m', 'YoYo', 'SBJ']
            player_id = request.POST.get('player_id')
            num_tests = request.POST.get('num_tests')
            
            # Validation
            if not selected_tests:
                context['error'] = 'Please select at least one test'
                return render(request, 'multi_test_report.html', context)
            
            if not player_id or not num_tests:
                context['error'] = 'Please select player and number of tests'
                return render(request, 'multi_test_report.html', context)
            
            # Get player
            player = get_object_or_404(
                Player, 
                id=player_id, 
                organization=request.user.organization
            )
            
            # Calculate date range
            start_date, end_date = _get_date_range(date_option, request.POST)
            
            # Get multi-test stats
            stats_data = _get_multi_test_stats(
                player, selected_tests, int(num_tests), start_date, end_date
            )
            
            # Player info
            player_info = _get_player_info(player)
            
            # Update context
            context.update({
                'player_info': player_info,
                'stats_data': stats_data,
                'selected_tests': selected_tests,
                'selected_player_id': player_id,
                'num_tests': num_tests,
                'error': None,
            })
            
        except Exception as e:
            context['error'] = f'Error: {str(e)}'
    
    return render(request, 'player_app/record/multi_test_report.html', context)

def _get_date_range(date_option, post_data):
    """Calculate date range based on option"""
    today = datetime.now().date()
    
    if date_option == 'range':
        start_date = post_data.get('start_date')
        end_date = post_data.get('end_date')
        return start_date, end_date
    elif date_option == '1month':
        return (today - timedelta(days=30), today)
    elif date_option == '3months':
        return (today - timedelta(days=90), today)
    elif date_option == '6months':
        return (today - timedelta(days=180), today)
    elif date_option == 'fy':
        # Financial Year Apr 1 - Mar 31
        year = datetime.now().year
        if datetime.now().month >= 4:
            start = datetime(year, 4, 1).date()
            end = datetime(year + 1, 3, 31).date()
        else:
            start = datetime(year - 1, 4, 1).date()
            end = datetime(year, 3, 31).date()
        return (start, end)
    else:
        return (today - timedelta(days=365), today)

def _get_multi_test_stats(player, test_names, num_tests, start_date=None, end_date=None):
    """Get stats for multiple tests"""
    all_stats = []
    
    for test_name in test_names:
        # Filter tests
        queryset = TestAndResult.objects.filter(
            player=player,
            test=test_name
        )
        
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        # Get latest N tests
        latest_tests = queryset.order_by('-date')[:num_tests]
        
        for test_result in latest_tests:
            stat_row = {
                'test': test_name,
                'date': test_result.date.strftime('%Y-%m-%d'),
                'phase': getattr(test_result, 'phase', ''),
                'best': getattr(test_result, 'best', ''),
                'notes': getattr(test_result, 'notes', ''),
                'reported_by': str(getattr(test_result, 'reported_by', '')),
            }
            
            # Test-specific fields
            if hasattr(test_result, 'distance_covered'):
                stat_row['distance_covered'] = getattr(test_result, 'distance_covered', '')
            if hasattr(test_result, 'predicted_vo2max'):
                stat_row['predicted_vo2max'] = getattr(test_result, 'predicted_vo2max', '')
            
            all_stats.append(stat_row)
    
    return all_stats

def _get_player_info(player):
    """Get formatted player info"""
    return {
        'name': player.name,
        'age': player.age,
        'gender': getattr(player, 'gender', ''),
        'role': getattr(player, 'role', ''),
        'handedness': getattr(player, 'handedness', ''),
        'age_category': getattr(player, 'age_category', ''),
        'batting_style': getattr(player, 'batting_style', ''),
        'bowling_style': getattr(player, 'bowling_style', ''),
        'player_status': getattr(player, 'player_status', ''),
        'image': getattr(player, 'image', None),
    }


@csrf_exempt
def fetch_multi_test_report(request):
    """AJAX endpoint for multi-test report"""
    try:
        selected_tests = request.POST.get('tests', '').split(',')
        player_id = request.POST.get('player_id')
        num_tests = int(request.POST.get('num_tests', 0))
        date_option = request.POST.get('date_option')
        
        player = get_object_or_404(Player, id=player_id, organization=request.user.organization)
        start_date, end_date = _get_date_range(date_option, request.POST)
        
        stats_data = _get_multi_test_stats(player, selected_tests, num_tests, start_date, end_date)
        player_info = _get_player_info(player)
        
        return JsonResponse({
            'success': True,
            # 'player_info': player_info,
            'stats_data': stats_data,
            'selected_tests': selected_tests
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    

def handle_test_preview(request):
    """Preview available test counts for dropdown"""
    try:
        player_id = request.POST.get('player_id')
        selected_tests = [t.strip() for t in request.POST.get('tests', '').split(',') if t.strip()]
        
        if not player_id:
            return JsonResponse({
                'success': True,
                'max_tests': 0,
                'message': 'Please select a player'
            })
        
        player = get_object_or_404(Player, id=player_id, organization=request.user.organization)
        max_count = 0
        test_counts = {}
        
        for test_type in selected_tests:
            count = count_player_tests(player, test_type)
            test_counts[test_type] = count
            max_count = max(max_count, count)
        
        if max_count == 0:
            return JsonResponse({
                'success': True,
                'max_tests': 0,
                'message': f'No test data found for selected tests: {", ".join(selected_tests)}'
            })
        
        return JsonResponse({
            'success': True,
            'max_tests': max_count,
            'test_counts': test_counts,
            'message': f'{max_count} tests available'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def count_player_tests(player, test_type):
    """Count player's tests for specific type"""
    test_counts = {
        '10m': TenMeterTest.objects.filter(player=player).count(),
        '20m': TwentyMeterTest.objects.filter(player=player).count(),
        'YoYo': YoYoTest.objects.filter(player=player).count(),
        'SBJ': SBJTest.objects.filter(player=player).count(),
        'Copenhagen': CopenhagenTest.objects.filter(player=player).count(),
        'S/L Glute Bridges': SLGluteBridges.objects.filter(player=player).count(),
    }
    return test_counts.get(test_type, 0)



def fetch_speed_tests(player, test_type, selected_tests, num_tests, start_date, end_date):  # ← NEW (was _get_test_data)
    """Fetch speed/endurance test data"""
    if test_type not in selected_tests:
        return []
    
    model_map = {
        '10m': TenMeterTest,
        '20m': TwentyMeterTest,
        'YoYo': YoYoTest,
        'SBJ': SBJTest,
    }
    
    model = model_map.get(test_type)
    if not model:
        return []
    
    tests = model.objects.filter(
        player=player,
        date__gte=start_date,
        date__lte=end_date
    ).order_by('-date')[:num_tests]
    
    return [
        {
            'date': test.date.strftime('%Y-%m-%d'),
            'phase': getattr(test, 'phase', 'N/A'),
            'gender': player.gender,
            'category': getattr(player, 'age_category', 'N/A'),
            'distance_covered': getattr(test, 'distance_covered', 'N/A'),
            'predicted_vo2max': getattr(test, 'predicted_vo2max', 'N/A'),
            'best': getattr(test, 'distance_covered', 'N/A'),
        }
        for test in tests
    ]

def fetch_asymmetry_tests(player, test_type, selected_tests, num_tests, start_date, end_date):  # ← NEW (was _get_asymmetry_test_data)
    """Fetch asymmetry test data"""
    if test_type not in selected_tests:
        return []
    
    model_map = {
        'Copenhagen': CopenhagenTest,
        'S/L Glute Bridges': SLGluteBridges,
    }
    
    model = model_map.get(test_type)
    if not model:
        return []
    
    tests = model.objects.filter(
        player=player,
        date__gte=start_date,
        date__lte=end_date
    ).order_by('-date')[:num_tests]
    
    return [
        {
            'date': test.date.strftime('%Y-%m-%d'),
            'phase': getattr(test, 'phase', 'N/A'),
            'right': getattr(test, 'right_side', 'N/A'),
            'left': getattr(test, 'left_side', 'N/A'),
            'difference': getattr(test, 'difference', 'N/A'),
            'ratio': getattr(test, 'ratio', 'N/A'),
            'gender': player.gender,
            'category': getattr(player, 'age_category', 'N/A'),
            'notes': getattr(test, 'notes', ''),
        }
        for test in tests
    ]

def parse_date_filter(date_option, post_data):  # ← NEW (was _get_date_range)
    """Parse date range from filter options"""
    today = timezone.now().date()
    
    if date_option == 'range':
        return post_data.get('start_date'), post_data.get('end_date')
    
    elif date_option == '1month':
        return today - timedelta(days=30), today
    elif date_option == '3months':
        return today - timedelta(days=90), today
    elif date_option == '6months':
        return today - timedelta(days=180), today
    elif date_option == 'fy':
        year = today.year
        if today.month < 4:
            start = timezone.datetime(year-1, 4, 1).date()
            end = timezone.datetime(year, 3, 31).date()
        else:
            start = timezone.datetime(year, 4, 1).date()
            end = timezone.datetime(year+1, 3, 31).date()
        return start, end
    else:
        return today - timedelta(days=30), today

def format_player_data(player):  # ← NEW (was _get_player_info)
    """Format player information for frontend"""
    return {
        'name': player.name,
        'age': player.age,
        'gender': player.gender,
        'role': getattr(player, 'role', 'N/A'),
        'handedness': getattr(player, 'handedness', 'N/A'),
        'age_category': getattr(player, 'age_category', 'N/A'),
        'batting_style': getattr(player, 'batting_style', 'N/A'),
        'bowling_style': getattr(player, 'bowling_style', 'N/A'),
        'player_status': getattr(player, 'player_status', 'Active'),
        'image': player.image.url if player.image else None,
    }


@csrf_exempt
def get_player_test_data(request):  # ← NEW NAME (was fetch_multi_test_report)
    """AJAX endpoint for player test data with preview counts"""
    try:
        action = request.POST.get('action')
        
        if action == 'preview_test_counts':
            return handle_test_preview(request)  # ← NEW
        
        # Main report generation
        selected_tests = [t.strip() for t in request.POST.get('tests', '').split(',') if t.strip()]
        player_id = request.POST.get('player_id')
        num_tests = int(request.POST.get('num_tests', 0))
        date_option = request.POST.get('date_option')
        
        player = get_object_or_404(Player, id=player_id, organization=request.user.organization)
        start_date, end_date = parse_date_filter(date_option, request.POST)  # ← NEW
        
        # Generate individual test data
        sprint10m_data = fetch_speed_tests(player, '10m', selected_tests, num_tests, start_date, end_date)  # ← NEW
        sprint20m_data = fetch_speed_tests(player, '20m', selected_tests, num_tests, start_date, end_date)
        yoyo_data = fetch_speed_tests(player, 'YoYo', selected_tests, num_tests, start_date, end_date)
        sbj_data = fetch_speed_tests(player, 'SBJ', selected_tests, num_tests, start_date, end_date)
        copenhagen_data = fetch_asymmetry_tests(player, 'Copenhagen', selected_tests, num_tests, start_date, end_date)  # ← NEW
        glute_data = fetch_asymmetry_tests(player, 'S/L Glute Bridges', selected_tests, num_tests, start_date, end_date)
        
        player_details = format_player_data(player)  # ← NEW
        
        return JsonResponse({
            'success': True,
            'player_info': player_details,
            'test10m': sprint10m_data,
            'test20m': sprint20m_data,
            'testYoYo': yoyo_data,
            'testSBJ': sbj_data,
            'testCopenhagen': copenhagen_data,
            'testGluteBridges': glute_data,
            'selected_tests': selected_tests,
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


    

from django.db.models import Q
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from datetime import date

today = date.today()

def multitest(request):
    if request.method=='POST':
        date_option = request.POST.get('date_option')
        selected_tests = request.POST.getlist('tests')  
        player_id = request.POST.get('player_id')
        num_tests = request.POST.get('num_tests')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
      

        date_time = 0
        if date_option == 'range':
           
            if not start_date or not end_date:
                date_time = (today - timedelta(days=30), today)
            date_time = (start_date, end_date)
           
        
        elif date_option == '1month':
            date_time = (today - timedelta(days=30), today)
            start_date = date_time[0]
            end_date = date_time[1]
        
        elif date_option == '3months':  # Fixed: was "3month"
            date_time =(today - timedelta(days=90), today)
            start_date = date_time[0]
            end_date = date_time[1]
        
        elif date_option == '6months':  # Fixed: was "6month"
            date_time = (today - timedelta(days=180), today)
            start_date = date_time[0]
            end_date = date_time[1]
        
        elif date_option == 'fy':
            # Financial Year: April 1 - March 31
            year = datetime.now().year
            if datetime.now().month >= 4:
                start_date = datetime(year, 4, 1).date()
                end_date = datetime(year + 1, 3, 31).date()
            else:
                start_date = datetime(year - 1, 4, 1).date()
                end_date = datetime(year, 3, 31).date()
            date_time = (start_date, end_date)
        
        else:  # Default: last year
            date_time = (today - timedelta(days=365), today)
            start_date = date_time[0]
            end_date = date_time[1]

        
        player = Player.objects.get(id=player_id)
        player_info = Player.objects.get(id=player_id)
        players = Player.objects.filter(organization=request.user.organization).order_by('name')
        test10 = None
        if "10m" in selected_tests:
            test10 = TenMeterTest.objects.filter(player=player,date__range=[start_date, end_date])[:int(num_tests)]
            
        
        test20 = None
        if "20m" in selected_tests:
            test20 = TwentyMeterTest.objects.filter(player=player,date__range=[start_date, end_date])[:int(num_tests)]
            

        testYoYo = None
        if "YoYo" in selected_tests:
            testYoYo = YoYoTest.objects.filter(player=player,date__range=[start_date, end_date])[:int(num_tests)]
            

        testSBJ = None
        if "SBJ" in selected_tests:
            testSBJ = SBJTest.objects.filter(player=player,date__range=[start_date, end_date])[:int(num_tests)]
            

        testCopenhagen = None
        if "Copenhagen" in selected_tests:
            testCopenhagen = CopenhagenTest.objects.filter(player=player,date__range=[start_date, end_date])[:int(num_tests)]
            

        testGluteBridges = None
        if "S/L Glute Bridges" in selected_tests:
            testGluteBridges = SLGluteBridges.objects.filter(player=player,date__range=[start_date, end_date])[:int(num_tests)]
            

        injuries = Injury.objects.filter(player=player,injury_date__range=[start_date, end_date])
        
       
        context = {
            "test10":test10,
            "test20":test20,
            "testYoYo":testYoYo,
            "testSBJ":testSBJ,
            "testCopenhagen":testCopenhagen,
            "testGluteBridges":testGluteBridges,
            'player_info':player_info,
            'players':players,
            'injuries': injuries,
            'total_injuries': injuries.count(),
            'open_injuries': injuries.filter(status='open').count(),
            'closed_injuries': injuries.filter(status='closed').count(),
        }

    return render(request,'player_app/record/multi_test_report.html',context)